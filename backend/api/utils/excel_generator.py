import calendar
import os
import socket
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from smb.SMBConnection import SMBConnection


class ExcelGenerator:
    BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
    OUTPUT_PATH = BASE_DIR / "data" / "excel"

    # Fallback defaults if not provided
    DEFAULT_DEPT_CODE = "K390140R1C"
    DEFAULT_DEPT_NAME = "BG6-RD Center-Automatic System Test R&D Div.1-Dept.1-PTB Sec.1"

    EXCEL_TEMP_ONLY = os.getenv("EXCEL_TEMP_ONLY", "false").lower() == "true"  # Set true to keep files temp-only (e.g., SMB-only uploads)

    # Use SMB_* environment variables (consistent with Django settings)
    SMB_CONFIG = {
        "host": os.getenv("SMB_SERVER") or "",
        "username": os.getenv("SMB_USERNAME") or "",
        "password": os.getenv("SMB_PASSWORD") or "",
        "share_name": os.getenv("SMB_SHARE_NAME") or "",
        "domain": os.getenv("SMB_DOMAIN", "WORKGROUP"),
        "port": int(os.getenv("SMB_PORT", "445")),
        "path": os.getenv("SMB_PATH_PREFIX", "Management\\PTB\\AST_Portal_Overtime\\").replace("\\", "/"),  # Will be set from SMB_PATH_PREFIX env var if provided
    }

    _connection_cache = None
    _connection_cache_time = None
    _connection_cache_timeout = 300

    # Cache for DB-loaded SMB config (avoid DB query on every call)
    _smb_config_cache_time: float | None = None
    _smb_config_cache_ttl = 60  # reload from DB every 60 seconds

    @classmethod
    def _load_smb_config(cls):
        """
        Load SMB config from the database (SMBConfiguration model) first,
        falling back to environment variables if no DB record exists.
        Cached for 60 seconds to avoid DB queries on every call.
        """
        now = time.monotonic()
        if cls._smb_config_cache_time and (now - cls._smb_config_cache_time) < cls._smb_config_cache_ttl:
            return  # Use cached config
        try:
            from ..models import SMBConfiguration

            cls.SMB_CONFIG = SMBConfiguration.get_active_config()
            cls._smb_config_cache_time = now
        except Exception:
            pass  # Keep env-based defaults

    @classmethod
    def invalidate_smb_cache(cls):
        """Force next _load_smb_config() to reload from DB."""
        cls._smb_config_cache_time = None

    @classmethod
    def _smb_configured(cls) -> bool:
        """Return True if all SMB config pieces are present (host/user/password/share)."""
        cls._load_smb_config()
        return all(cls.SMB_CONFIG.get(key) for key in ("host", "username", "password", "share_name"))

    @classmethod
    def get_period_folder(cls, date: datetime | None = None) -> str:
        return cls.get_local_period_folder(date)

    @classmethod
    def get_local_period_folder(cls, date: datetime | None = None) -> str:
        if date is None:
            date = datetime.now()

        current_period_start = date.replace(day=26)
        if date.day < 26:
            current_period_start = (date.replace(day=1) - timedelta(days=1)).replace(day=26)

        next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)
        folder_name = f"{current_period_start.strftime('%Y-%m-%d')}_{next_period_end.strftime('%Y-%m-%d')}"

        period_path = cls.OUTPUT_PATH / folder_name
        period_path.mkdir(parents=True, exist_ok=True)
        return str(period_path)

    @classmethod
    def get_smb_period_folder(cls, date: datetime | None = None) -> str | None:
        if not cls._smb_configured():
            return None

        if date is None:
            date = datetime.now()

        current_period_start = date.replace(day=26)
        if date.day < 26:
            current_period_start = (date.replace(day=1) - timedelta(days=1)).replace(day=26)

        next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)
        folder_name = f"{current_period_start.strftime('%Y-%m-%d')}_{next_period_end.strftime('%Y-%m-%d')}"

        base_path = cls.SMB_CONFIG["path"].rstrip("/")
        if base_path:
            return f"{base_path}/{folder_name}"
        return folder_name

    @classmethod
    def get_smb_connection(cls, use_cache: bool = True):
        if not cls._smb_configured():
            return None

        current_time = time.time()
        if use_cache and cls._connection_cache and cls._connection_cache_time and (current_time - cls._connection_cache_time) < cls._connection_cache_timeout:
            try:
                cls._connection_cache.listShares()
                return cls._connection_cache
            except Exception:
                cls._connection_cache = None
                cls._connection_cache_time = None

        client_name = socket.gethostname()
        conn = SMBConnection(
            username=cls.SMB_CONFIG["username"],
            password=cls.SMB_CONFIG["password"],
            my_name=client_name,
            remote_name=cls.SMB_CONFIG["host"],
            use_ntlm_v2=True,
            is_direct_tcp=True,
        )

        if conn.connect(cls.SMB_CONFIG["host"], 445, timeout=15):
            if use_cache:
                cls._connection_cache = conn
                cls._connection_cache_time = current_time
            return conn

        conn.close()
        return None

    @classmethod
    def clear_connection_cache(cls):
        if cls._connection_cache:
            try:
                cls._connection_cache.close()
            except Exception:
                pass
        cls._connection_cache = None
        cls._connection_cache_time = None

    @classmethod
    def ensure_folder_exists(cls, relative_path: str, conn=None):
        if not relative_path:
            return

        should_close = False
        if conn is None:
            conn = cls.get_smb_connection()
            should_close = conn is not None

        if conn is None:
            return

        try:
            try:
                conn.listPath(cls.SMB_CONFIG["share_name"], relative_path)
            except Exception:
                try:
                    conn.createDirectory(cls.SMB_CONFIG["share_name"], relative_path)
                except Exception as dir_error:
                    msg = str(dir_error).lower()
                    if "already exists" not in msg and "file exists" not in msg:
                        print(f"Warning: could not create SMB directory {relative_path}: {dir_error}")
        finally:
            if should_close:
                conn.close()

    @classmethod
    def _format_hours(cls, value):
        try:
            hours_value = float(value) if isinstance(value, str) else value
            if hours_value == int(hours_value):
                return str(int(hours_value))
            if hours_value * 10 % 10 == 0:
                return str(hours_value).rstrip("0").rstrip(".")
            return str(hours_value)
        except Exception:
            return str(value)

    @classmethod
    def create_ot_form(cls, data, date_info, dept_code=None, dept_name=None):
        # Use provided dept info or fallback to defaults
        dept_code = dept_code or cls.DEFAULT_DEPT_CODE
        dept_name = dept_name or cls.DEFAULT_DEPT_NAME

        wb = Workbook()
        ws = wb.active

        columns = {
            "A": 4.14,
            "B": 13.43,
            "C": 21.14,
            "D": 22.29,
            "E": 17.86,
            "F": 14.14,
            "G": 15.14,
            "H": 13.57,
            "I": 15.71,
            "J": 1.00,
            "K": 17.14,
            "L": 14.43,
            "M": 15.43,
            "N": 18.57,
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        ws.merge_cells("A1:N1")
        ws["A1"] = "Form Lembur (加 班 申 请 单)"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(name="Arial", size=20)

        ws.merge_cells("A2:C2")
        ws["A2"] = f"Departemen\n(部门代码)：\n{dept_code}\n{dept_name}"
        ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["A2"].font = Font(name="Arial", size=11)

        ws["D2"] = "Klasifikasi Karyawan\n(人员分类):"
        ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
        ws["E2"] = "☐ Band0(1-2職等)\n☑ Band1(3-5職等)"
        ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

        ch_date = f"{date_info.year} 年 {date_info.month:02d} 月 {date_info.day:02d} 日"
        weekday = calendar.day_name[date_info.weekday()].upper()
        ws.merge_cells("F2:H2")
        ws["F2"] = f"Tanggal Lembur （加班日期）：\n                           {ch_date}\nHari （星期）: {weekday}"
        ws["F2"].alignment = Alignment(vertical="center", wrap_text=True)

        overtime_type = "Jenis Lembur （加班类别） :\n"
        if data and len(data) > 0:
            first_request = data[0]
            is_weekend = first_request.get("is_weekend", False)
            is_holiday = first_request.get("is_holiday", False)

            if is_holiday:
                overtime_type += "☐ Saat Hari Kerja (工作日延长加班)\n☐ Saat Hari Libur (休息日加班)\n☑ Saat Tanggal Merah (法定假日加班)"
            elif is_weekend:
                overtime_type += "☐ Saat Hari Kerja (工作日延长加班)\n☑ Saat Hari Libur (休息日加班)\n☐ Saat Tanggal Merah (法定假日加班)"
            else:
                overtime_type += "☑ Saat Hari Kerja (工作日延长加班)\n☐ Saat Hari Libur (休息日加班)\n☐ Saat Tanggal Merah (法定假日加班)"

        ws.merge_cells("I2:N2")
        ws["I2"] = overtime_type
        ws["I2"].alignment = Alignment(vertical="center", wrap_text=True)

        headers = [
            ("A3", "No"),
            ("B3", "No Karyawan\n(工号)"),
            ("C3", "Nama\n(姓名)"),
            ("D3", "Alasan Lembur\n(申请加班事由)"),
            ("E3", "Jam Lembur (Waktu)\n(预计加班\n起止时间)"),
            ("F3", "Durasi Lembur\n(预计加班时数)"),
            ("G3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)"),
            ("H3", "Tanda Tangan Karyawan\n(员工签名)"),
            ("I3", "Tanda Tangan Supervisor\n(课级主管签名)"),
            ("K3", "Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)"),
            ("L3", "Konfirmasi Durasi Lembur\n(实际加班时数)"),
            ("M3", "Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)"),
            ("N3", "Tanda Tangan Karyawan\n(员工签名)"),
        ]

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        for row in range(4, 34):
            for col in "ABCDEFGHIKLMN":
                cell = f"{col}{row}"
                ws[cell] = ""
                ws[cell].border = thin_border
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
            ws[f"A{row}"] = row - 3

        current_row = 4
        for item in data or []:
            ws[f"B{current_row}"] = item.get("employee_id")
            ws[f"C{current_row}"] = item.get("employee_name")

            ws[f"D{current_row}"] = item.get("reason")
            ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"E{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"F{current_row}"] = f"{item.get('total_hours')} hour(s)"
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee takes break time (original format)
            ws[f"G{current_row}"] = "V" if item.get("has_break") else "-"
            ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"K{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
            ws[f"K{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"L{current_row}"] = f"{item.get('total_hours')} hour(s)"
            ws[f"L{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee takes break time (original format)
            ws[f"M{current_row}"] = "V" if item.get("has_break") else "-"
            ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        ws.merge_cells("A34:N34")
        ws["A34"] = "Keterangan："
        ws["A34"].font = Font(name="Arial", size=12)

        notes = [
            ("A35:N35", "1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。"),
            ("A36:N36", "2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。"),
            ("A37:N37", "3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。"),
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            anchor = cells.split(":")[0]
            ws[anchor] = text
            ws[anchor].font = Font(name="Arial", size=12)
            ws[anchor].alignment = Alignment(vertical="center", wrap_text=True)

        ws["M38"] = "Form No.:PTB-TB004-001 Rev.01"
        ws["M38"].font = Font(name="Arial", size=10)

        return wb

    @classmethod
    def create_ot_summary(cls, data, date_info, dept_code=None, dept_name=None):
        """Create overtime summary workbook (original format)"""
        # Use provided dept info or fallback to defaults (not used in summary but kept for consistency)
        dept_code = dept_code or cls.DEFAULT_DEPT_CODE
        dept_name = dept_name or cls.DEFAULT_DEPT_NAME

        wb = Workbook()
        ws = wb.active

        # Original column widths
        columns = {
            "A": 11.5,  # Work ID
            "B": 12.0,  # Overtime Type
            "C": 15.0,  # Overtime Start Date
            "D": 9.75,  # Start Time
            "E": 9.75,  # End Time
            "F": 9.75,  # Break
            "G": 6.0,  # Hours
            "H": 25.0,  # Reason
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Original headers
        headers = [
            ("A1", "Work ID"),
            ("B1", "Overtime Type"),
            ("C1", "Overtime Start Date"),
            ("D1", "Start Time"),
            ("E1", "End Time"),
            ("F1", "Meal/Rest"),
            ("G1", "Hours"),
            ("H1", "Reason"),
        ]

        header_style = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        current_row = 2
        for item in data or []:
            # Determine overtime type based on weekend/holiday (1=weekday, 2=weekend, 3=holiday)
            ot_type = 1  # Default: weekday
            if item.get("is_holiday", False):
                ot_type = 3  # Holiday
            elif item.get("is_weekend", False):
                ot_type = 2  # Weekend

            # Format date from date_info
            formatted_date = date_info.strftime("%Y-%m-%d")

            ws[f"A{current_row}"] = item.get("employee_id")
            ws[f"B{current_row}"] = ot_type
            ws[f"C{current_row}"] = formatted_date
            ws[f"D{current_row}"] = item.get("time_start")
            ws[f"E{current_row}"] = item.get("time_end")
            ws[f"F{current_row}"] = "Y" if item.get("has_break") else "N"
            ws[f"G{current_row}"] = cls._format_hours(item.get("total_hours"))
            ws[f"H{current_row}"] = item.get("reason")

            # Apply formatting
            for col in "ABCDEFGH":
                cell = f"{col}{current_row}"
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
                if col in "ABCDEFG":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        return wb

    @classmethod
    def create_monthly_ot_form(cls, monthly_data, period_start, period_end, dept_code=None, dept_name=None):
        # Use provided dept info or fallback to defaults
        dept_code = dept_code or cls.DEFAULT_DEPT_CODE
        dept_name = dept_name or cls.DEFAULT_DEPT_NAME

        wb = Workbook()
        ws = wb.active

        columns = {
            "A": 4.14,
            "B": 13.88,
            "C": 13.93,
            "D": 21.14,
            "E": 22.29,
            "F": 15.22,
            "G": 16.55,
            "H": 14.55,
            "I": 15.14,
            "J": 13.57,
            "K": 15.71,
            "L": 1.00,
            "M": 16.55,
            "N": 14.55,
            "O": 15.43,
            "P": 18.57,
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        ws.merge_cells("A1:P1")
        ws["A1"] = "Form Lembur Bulanan (月度加班申请单)"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(name="Arial", size=20, bold=True)

        ws.merge_cells("A2:C2")
        ws["A2"] = f"Departemen\n(部门代码)：\n{dept_code}\n{dept_name}"
        ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["A2"].font = Font(name="Arial", size=11)

        ws["D2"] = "Klasifikasi Karyawan\n(人员分类):"
        ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
        ws["E2"] = "☐ Band0(1-2職等)\n☑ Band1(3-5職等)"
        ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

        period_str = f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')}"
        ws.merge_cells("F2:H2")
        ws["F2"] = f"Periode Lembur (加班期间):\n{period_str}"
        ws["F2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells("I2:P2")
        ws["I2"] = ""

        headers = [
            ("A3", "No"),
            ("B3", "Tanggal Lembur\n(加班日期)"),
            ("C3", "No Karyawan\n(工号)"),
            ("D3", "Nama\n(姓名)"),
            ("E3", "Alasan Lembur\n(申请加班事由)"),
            ("F3", "Jenis Lembur\n(加班类别)"),
            ("G3", "Jam Lembur (Waktu)\n(预计加班\n起止时间)"),
            ("H3", "Durasi Lembur\n(预计加班时数)"),
            ("I3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)"),
            ("J3", "Tanda Tangan Karyawan\n(员工签名)"),
            ("K3", "Tanda Tangan Supervisor\n(课级主管签名)"),
            ("M3", "Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)"),
            ("N3", "Konfirmasi Durasi Lembur\n(实际加班时数)"),
            ("O3", "Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)"),
            ("P3", "Tanda Tangan Karyawan\n(员工签名)"),
        ]

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        data_row_count = len(monthly_data) if monthly_data else 0
        required_rows = max(30, data_row_count)
        last_data_row = 4 + required_rows - 1

        for row in range(4, last_data_row + 1):
            for col in "ABCDEFGHIJKLMNOP":
                cell = f"{col}{row}"
                ws[cell] = ""
                ws[cell].border = thin_border
                ws[cell].font = Font(name="Arial", size=11)
                if col in "ABCD":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws[cell].alignment = Alignment(vertical="center")
            ws[f"A{row}"] = row - 3

        current_row = 4
        for item in monthly_data or []:
            ws[f"B{current_row}"] = item.get("request_date")
            ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"C{current_row}"] = item.get("employee_id")
            ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"D{current_row}"] = item.get("employee_name")
            ws[f"D{current_row}"].alignment = Alignment(vertical="center")

            ws[f"E{current_row}"] = item.get("reason")
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            if item.get("is_holiday", False):
                overtime_type = "Holiday"
            elif item.get("is_weekend", False):
                overtime_type = "Weekend"
            else:
                overtime_type = "Weekday"

            ws[f"F{current_row}"] = overtime_type
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"G{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
            ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"H{current_row}"] = f"{item.get('total_hours')} hour(s)"
            ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"I{current_row}"] = "V" if item.get("has_break") else "-"
            ws[f"I{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"M{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
            ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"N{current_row}"] = f"{item.get('total_hours')} hour(s)"
            ws[f"N{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"O{current_row}"] = "V" if item.get("has_break") else "-"
            ws[f"O{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        footer_row = last_data_row + 1
        notes_start_row = footer_row + 1
        form_number_row = notes_start_row + 3

        ws.merge_cells(f"A{footer_row}:P{footer_row}")
        ws[f"A{footer_row}"] = "Keterangan:"
        ws[f"A{footer_row}"].font = Font(name="Arial", size=12)

        notes = [
            (f"A{notes_start_row}:P{notes_start_row}", "1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。"),
            (f"A{notes_start_row + 1}:P{notes_start_row + 1}", "2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。"),
            (f"A{notes_start_row + 2}:P{notes_start_row + 2}", "3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。"),
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            anchor = cells.split(":")[0]
            ws[anchor] = text
            ws[anchor].font = Font(name="Arial", size=12)
            ws[anchor].alignment = Alignment(vertical="center", wrap_text=True)

        ws[f"O{form_number_row}"] = "Form No.:PTB-TB004-001 Rev.01"
        ws[f"O{form_number_row}"].font = Font(name="Arial", size=10)

        return wb

    @classmethod
    def create_monthly_ot_summary(cls, monthly_data, period_start, period_end, dept_code=None, dept_name=None):
        # Use provided dept info or fallback to defaults (not used in summary but kept for consistency)
        dept_code = dept_code or cls.DEFAULT_DEPT_CODE
        dept_name = dept_name or cls.DEFAULT_DEPT_NAME

        wb = Workbook()
        ws = wb.active

        columns = {
            "A": 11.5,
            "B": 12.0,
            "C": 15.0,
            "D": 9.75,
            "E": 9.75,
            "F": 9.75,
            "G": 6.0,
            "H": 25.0,
            "I": 25.0,
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        headers = [
            ("A1", "Work ID"),
            ("B1", "Overtime Type"),
            ("C1", "Overtime Start Date"),
            ("D1", "Start Time"),
            ("E1", "End Time"),
            ("F1", "Meal/Rest"),
            ("G1", "Hours"),
            ("H1", "Reason"),
            ("I1", "Detail"),
        ]

        header_style = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        current_row = 2
        for item in monthly_data or []:
            ot_type = 1
            if item.get("is_holiday", False):
                ot_type = 3
            elif item.get("is_weekend", False):
                ot_type = 2

            ws[f"A{current_row}"] = item.get("employee_id")
            ws[f"B{current_row}"] = ot_type
            ws[f"C{current_row}"] = item.get("request_date")
            ws[f"D{current_row}"] = item.get("time_start")
            ws[f"E{current_row}"] = item.get("time_end")
            ws[f"F{current_row}"] = "Y" if item.get("has_break") else "N"
            ws[f"G{current_row}"] = cls._format_hours(item.get("total_hours"))
            ws[f"H{current_row}"] = item.get("reason")
            ws[f"I{current_row}"] = item.get("detail", "")

            for col in "ABCDEFGHI":
                cell = f"{col}{current_row}"
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
                if col in "ABCDEFG":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        return wb

    @classmethod
    def create_ot_form_multi_sheet(cls, data_by_department, date_info):
        """
        Create daily OT form with multiple sheets (one per department)

        Args:
            data_by_department: Dict like {dept_code: {dept_code, dept_name, data: [...]}}
            date_info: Date object

        Returns:
            Workbook with one sheet per department
        """
        wb = Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create a sheet for each department
        for dept_code, dept_info in sorted(data_by_department.items()):
            ws = wb.create_sheet(title=dept_code[:31])  # Excel sheet name limit is 31 chars
            data = dept_info["data"]
            dept_name = dept_info["dept_name"]

            # Set column widths
            columns = {
                "A": 4.14,
                "B": 13.43,
                "C": 21.14,
                "D": 22.29,
                "E": 17.14,
                "F": 14.43,
                "G": 15.14,
                "H": 13.57,
                "I": 15.71,
                "J": 1.00,
                "K": 17.14,
                "L": 14.43,
                "M": 15.43,
                "N": 18.57,
            }
            for col, width in columns.items():
                ws.column_dimensions[col].width = width

            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 80
            ws.row_dimensions[3].height = 92

            # Title and department header
            ws.merge_cells("A1:N1")
            ws["A1"] = "Form Lembur (加 班 申 请 单)"
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(name="Arial", size=20)

            ws.merge_cells("A2:C2")
            ws["A2"] = f"Departemen\n(部门代码)：\n{dept_code}\n{dept_name}"
            ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
            ws["A2"].font = Font(name="Arial", size=11)

            ws["D2"] = "Klasifikasi Karyawan\n(人员分类):"
            ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
            ws["E2"] = "☐ Band0(1-2職等)\n☑ Band1(3-5職等)"
            ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

            ch_date = f"{date_info.year} 年 {date_info.month:02d} 月 {date_info.day:02d} 日"
            weekday = calendar.day_name[date_info.weekday()].upper()
            ws.merge_cells("F2:H2")
            ws["F2"] = f"Tanggal Lembur （加班日期）：\n                           {ch_date}\nHari （星期）: {weekday}"
            ws["F2"].alignment = Alignment(vertical="center", wrap_text=True)

            overtime_type = "Jenis Lembur （加班类别） :\n"
            if data and len(data) > 0:
                first_request = data[0]
                is_weekend = first_request.get("is_weekend", False)
                is_holiday = first_request.get("is_holiday", False)

                if is_holiday:
                    overtime_type += "☐ Saat Hari Kerja (工作日延长加班)\n☐ Saat Hari Libur (休息日加班)\n☑ Saat Tanggal Merah (法定假日加班)"
                elif is_weekend:
                    overtime_type += "☐ Saat Hari Kerja (工作日延长加班)\n☑ Saat Hari Libur (休息日加班)\n☐ Saat Tanggal Merah (法定假日加班)"
                else:
                    overtime_type += "☑ Saat Hari Kerja (工作日延长加班)\n☐ Saat Hari Libur (休息日加班)\n☐ Saat Tanggal Merah (法定假日加班)"

            ws.merge_cells("I2:N2")
            ws["I2"] = overtime_type
            ws["I2"].alignment = Alignment(vertical="center", wrap_text=True)

            # Headers
            headers = [
                ("A3", "No"),
                ("B3", "No Karyawan\n(工号)"),
                ("C3", "Nama\n(姓名)"),
                ("D3", "Alasan Lembur\n(申请加班事由)"),
                ("E3", "Jam Lembur (Waktu)\n(预计加班\n起止时间)"),
                ("F3", "Durasi Lembur\n(预计加班时数)"),
                ("G3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)"),
                ("H3", "Tanda Tangan Karyawan\n(员工签名)"),
                ("I3", "Tanda Tangan Supervisor\n(课级主管签名)"),
                ("K3", "Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)"),
                ("L3", "Konfirmasi Durasi Lembur\n(实际加班时数)"),
                ("M3", "Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)"),
                ("N3", "Tanda Tangan Karyawan\n(员工签名)"),
            ]

            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_font = Font(name="Arial", size=11)

            for cell, text in headers:
                ws[cell] = text
                ws[cell].alignment = header_style
                ws[cell].font = header_font
                ws[cell].border = thin_border

            # Data rows
            for row in range(4, 34):
                for col in "ABCDEFGHIKLMN":
                    cell = f"{col}{row}"
                    ws[cell] = ""
                    ws[cell].border = thin_border
                    ws[cell].font = Font(name="Arial", size=11)
                    ws[cell].alignment = Alignment(vertical="center")
                ws[f"A{row}"] = row - 3

            # Fill data
            current_row = 4
            for item in data or []:
                ws[f"B{current_row}"] = item.get("employee_id")
                ws[f"C{current_row}"] = item.get("employee_name")
                ws[f"D{current_row}"] = item.get("reason")
                ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"E{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
                ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"F{current_row}"] = f"{item.get('total_hours')} hour(s)"
                ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Show 'V' if employee takes break time (original format)
                ws[f"G{current_row}"] = "V" if item.get("has_break") else "-"
                ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"K{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
                ws[f"K{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws[f"L{current_row}"] = f"{item.get('total_hours')} hour(s)"
                ws[f"L{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Show 'V' if employee takes break time (original format)
                ws[f"M{current_row}"] = "V" if item.get("has_break") else "-"
                ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

            # Notes
            ws.merge_cells("A34:N34")
            ws["A34"] = "Keterangan："
            ws["A34"].font = Font(name="Arial", size=12)

            notes = [
                ("A35:N35", "1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。"),
                ("A36:N36", "2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。"),
                ("A37:N37", "3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。"),
            ]

            for cells, text in notes:
                ws.merge_cells(cells)
                anchor = cells.split(":")[0]
                ws[anchor] = text
                ws[anchor].font = Font(name="Arial", size=12)
                ws[anchor].alignment = Alignment(vertical="center", wrap_text=True)

            ws["M38"] = "Form No.:PTB-TB004-001 Rev.01"
            ws["M38"].font = Font(name="Arial", size=10)

        return wb

    @classmethod
    def create_ot_summary_multi_sheet(cls, data_by_department, date_info):
        """Create daily OT summary workbook with one sheet per department (original format).

        Expects grouped data structure from export_daily_data_by_department.
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for dept_code, dept_info in sorted(data_by_department.items()):
            ws = wb.create_sheet(title=dept_code[:31])
            data = dept_info["data"]

            # Original column widths
            columns = {
                "A": 11.5,  # Work ID
                "B": 12.0,  # Overtime Type
                "C": 15.0,  # Overtime Start Date
                "D": 9.75,  # Start Time
                "E": 9.75,  # End Time
                "F": 9.75,  # Break
                "G": 6.0,  # Hours
                "H": 25.0,  # Reason
            }
            for col, width in columns.items():
                ws.column_dimensions[col].width = width

            # Original headers
            headers = [
                ("A1", "Work ID"),
                ("B1", "Overtime Type"),
                ("C1", "Overtime Start Date"),
                ("D1", "Start Time"),
                ("E1", "End Time"),
                ("F1", "Meal/Rest"),
                ("G1", "Hours"),
                ("H1", "Reason"),
            ]

            header_style = Alignment(horizontal="center", vertical="center")
            header_font = Font(name="Arial", size=11)

            for cell, text in headers:
                ws[cell] = text
                ws[cell].alignment = header_style
                ws[cell].font = header_font

            current_row = 2
            for item in data or []:
                # Determine overtime type based on weekend/holiday (1=weekday, 2=weekend, 3=holiday)
                ot_type = 1  # Default: weekday
                if item.get("is_holiday", False):
                    ot_type = 3  # Holiday
                elif item.get("is_weekend", False):
                    ot_type = 2  # Weekend

                # Format date from date_info
                formatted_date = date_info.strftime("%Y-%m-%d")

                ws[f"A{current_row}"] = item.get("employee_id")
                ws[f"B{current_row}"] = ot_type
                ws[f"C{current_row}"] = formatted_date
                ws[f"D{current_row}"] = item.get("time_start")
                ws[f"E{current_row}"] = item.get("time_end")
                ws[f"F{current_row}"] = "Y" if item.get("has_break") else "N"
                ws[f"G{current_row}"] = cls._format_hours(item.get("total_hours"))
                ws[f"H{current_row}"] = item.get("reason")

                for col in "ABCDEFGH":
                    cell = f"{col}{current_row}"
                    ws[cell].font = Font(name="Arial", size=11)
                    ws[cell].alignment = Alignment(vertical="center")
                    if col in "ABCDEFG":
                        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

        return wb

    @classmethod
    def create_monthly_ot_form_multi_sheet(cls, data_by_department, period_start, period_end):
        """Create monthly OT form workbook with one sheet per department (original format).

        Expects grouped data from export_monthly_data_by_department.
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for dept_code, dept_info in sorted(data_by_department.items()):
            ws = wb.create_sheet(title=dept_code[:31])
            data = dept_info["data"]
            dept_name = dept_info["dept_name"]

            # Column widths setup (original format)
            columns = {
                "A": 4.14,
                "B": 13.88,
                "C": 13.93,
                "D": 21.14,
                "E": 22.29,
                "F": 15.22,
                "G": 16.55,
                "H": 14.55,
                "I": 15.14,
                "J": 13.57,
                "K": 15.71,
                "L": 1.00,
                "M": 16.55,
                "N": 14.55,
                "O": 15.43,
                "P": 18.57,
            }
            for col, width in columns.items():
                ws.column_dimensions[col].width = width

            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 80
            ws.row_dimensions[3].height = 92

            # Title
            ws.merge_cells("A1:P1")
            ws["A1"] = "Form Lembur Bulanan (月度加班申请单)"
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(name="Arial", size=20, bold=True)

            # Department info
            ws.merge_cells("A2:C2")
            ws["A2"] = f"Departemen\n(部门代码)：\n{dept_code}\n{dept_name}"
            ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
            ws["A2"].font = Font(name="Arial", size=11)

            # Employee classification
            ws["D2"] = "Klasifikasi Karyawan\n(人员分类):"
            ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
            ws["E2"] = "☐ Band0(1-2職等)\n☑ Band1(3-5職等)"
            ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

            # Period information
            period_str = f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')}"
            ws.merge_cells("F2:H2")
            ws["F2"] = f"Periode Lembur (加班期间):\n{period_str}"
            ws["F2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Empty right section
            ws.merge_cells("I2:P2")
            ws["I2"] = ""

            # Column headers (original format)
            headers = [
                ("A3", "No"),
                ("B3", "Tanggal Lembur\n(加班日期)"),
                ("C3", "No Karyawan\n(工号)"),
                ("D3", "Nama\n(姓名)"),
                ("E3", "Alasan Lembur\n(申请加班事由)"),
                ("F3", "Jenis Lembur\n(加班类别)"),
                ("G3", "Jam Lembur (Waktu)\n(预计加班\n起止时间)"),
                ("H3", "Durasi Lembur\n(预计加班时数)"),
                ("I3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)"),
                ("J3", "Tanda Tangan Karyawan\n(员工签名)"),
                ("K3", "Tanda Tangan Supervisor\n(课级主管签名)"),
                ("M3", "Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)"),
                ("N3", "Konfirmasi Durasi Lembur\n(实际加班时数)"),
                ("O3", "Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)"),
                ("P3", "Tanda Tangan Karyawan\n(员工签名)"),
            ]

            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_font = Font(name="Arial", size=11)

            for cell, text in headers:
                ws[cell] = text
                ws[cell].alignment = header_style
                ws[cell].font = header_font
                ws[cell].border = thin_border

            # Calculate required rows
            data_row_count = len(data) if data else 0
            required_rows = max(30, data_row_count)
            last_data_row = 4 + required_rows - 1

            # Create empty rows with formatting
            for row in range(4, last_data_row + 1):
                for col in "ABCDEFGHIJKLMNOP":
                    cell = f"{col}{row}"
                    ws[cell] = ""
                    ws[cell].border = thin_border
                    ws[cell].font = Font(name="Arial", size=11)
                    if col in "ABCD":
                        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        ws[cell].alignment = Alignment(vertical="center")
                ws[f"A{row}"] = row - 3

            # Fill in monthly data
            current_row = 4
            for item in data or []:
                ws[f"B{current_row}"] = item.get("request_date")
                ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"C{current_row}"] = item.get("employee_id")
                ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"D{current_row}"] = item.get("employee_name")
                ws[f"D{current_row}"].alignment = Alignment(vertical="center")

                ws[f"E{current_row}"] = item.get("reason")
                ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Overtime type based on weekend/holiday status
                if item.get("is_holiday", False):
                    overtime_type = "Holiday"
                elif item.get("is_weekend", False):
                    overtime_type = "Weekend"
                else:
                    overtime_type = "Weekday"

                ws[f"F{current_row}"] = overtime_type
                ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"G{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
                ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"H{current_row}"] = f"{item.get('total_hours')} hour(s)"
                ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Show 'V' if employee takes break time (original format)
                ws[f"I{current_row}"] = "V" if item.get("has_break") else "-"
                ws[f"I{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Actual time, duration and break status
                ws[f"M{current_row}"] = f"{item.get('time_start')} - {item.get('time_end')}"
                ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                ws[f"N{current_row}"] = f"{item.get('total_hours')} hour(s)"
                ws[f"N{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Show 'V' if employee takes break time (original format)
                ws[f"O{current_row}"] = "V" if item.get("has_break") else "-"
                ws[f"O{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

            # Dynamic footer positioning
            footer_row = last_data_row + 1
            notes_start_row = footer_row + 1
            form_number_row = notes_start_row + 3

            ws.merge_cells(f"A{footer_row}:P{footer_row}")
            ws[f"A{footer_row}"] = "Keterangan:"
            ws[f"A{footer_row}"].font = Font(name="Arial", size=12)

            notes = [
                (f"A{notes_start_row}:P{notes_start_row}", "1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。"),
                (f"A{notes_start_row + 1}:P{notes_start_row + 1}", "2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。"),
                (f"A{notes_start_row + 2}:P{notes_start_row + 2}", "3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。"),
            ]

            for cells, text in notes:
                ws.merge_cells(cells)
                anchor = cells.split(":")[0]
                ws[anchor] = text
                ws[anchor].font = Font(name="Arial", size=12)
                ws[anchor].alignment = Alignment(vertical="center", wrap_text=True)

            ws[f"O{form_number_row}"] = "Form No.:PTB-TB004-001 Rev.01"
            ws[f"O{form_number_row}"].font = Font(name="Arial", size=10)

        return wb

    @classmethod
    def create_monthly_ot_summary_multi_sheet(cls, data_by_department, period_start, period_end):
        """Create monthly OT summary workbook with one sheet per department.

        Expects grouped data from export_monthly_data_by_department.
        """
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for dept_code, dept_info in sorted(data_by_department.items()):
            ws = wb.create_sheet(title=dept_code[:31])
            data = dept_info["data"]

            columns = {"A": 11.5, "B": 12.0, "C": 15.0, "D": 9.75, "E": 9.75, "F": 9.75, "G": 6.0, "H": 25.0, "I": 25.0}
            for col, width in columns.items():
                ws.column_dimensions[col].width = width

            headers = [
                ("A1", "Work ID"),
                ("B1", "Overtime Type"),
                ("C1", "Overtime Start Date"),
                ("D1", "Start Time"),
                ("E1", "End Time"),
                ("F1", "Meal/Rest"),
                ("G1", "Hours"),
                ("H1", "Reason"),
                ("I1", "Detail"),
            ]

            header_style = Alignment(horizontal="center", vertical="center")
            header_font = Font(name="Arial", size=11)

            for cell, text in headers:
                ws[cell] = text
                ws[cell].alignment = header_style
                ws[cell].font = header_font

            current_row = 2
            for item in data or []:
                ot_type = 1
                if item.get("is_holiday", False):
                    ot_type = 3
                elif item.get("is_weekend", False):
                    ot_type = 2

                ws[f"A{current_row}"] = item.get("employee_id")
                ws[f"B{current_row}"] = ot_type
                ws[f"C{current_row}"] = item.get("request_date")
                ws[f"D{current_row}"] = item.get("time_start")
                ws[f"E{current_row}"] = item.get("time_end")
                ws[f"F{current_row}"] = "Y" if item.get("has_break") else "N"
                ws[f"G{current_row}"] = cls._format_hours(item.get("total_hours"))
                ws[f"H{current_row}"] = item.get("reason")
                ws[f"I{current_row}"] = item.get("detail", "")

                for col in "ABCDEFGHI":
                    cell = f"{col}{current_row}"
                    ws[cell].font = Font(name="Arial", size=11)
                    ws[cell].alignment = Alignment(vertical="center")
                    if col in "ABCDEFG":
                        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

                current_row += 1

        return wb

    @classmethod
    def generate_excel_files(cls, data, date_obj, upload: bool = True, dept_code=None, dept_name=None):
        local_paths = cls.generate_all_excel_files(data, None, date_obj, upload=upload, dept_code=dept_code, dept_name=dept_name)
        return local_paths.get("daily_form"), local_paths.get("daily_summary")

    @classmethod
    def generate_monthly_excel_files(cls, monthly_data, date_obj, upload: bool = True, dept_code=None, dept_name=None):
        local_paths = cls.generate_all_excel_files(None, monthly_data, date_obj, upload=upload, dept_code=dept_code, dept_name=dept_name)
        return local_paths.get("monthly_form"), local_paths.get("monthly_summary")

    @classmethod
    def generate_all_excel_files(cls, daily_data, monthly_data, date_obj, upload: bool = True, temp_only: bool = False, dept_code=None, dept_name=None):
        """Generate Excel files with multi-department sheet support and auto-detection.

        Accepts grouped inputs (preferred) or single-department data; ungrouped payloads are wrapped
        to maintain backward compatibility. When grouped, one workbook is produced with a sheet per
        department for each report type.
        """
        if not daily_data and not monthly_data:
            return {}

        date_obj = date_obj or datetime.now()
        date_str = date_obj.strftime("%Y%m%d")

        temp_only = temp_only or cls.EXCEL_TEMP_ONLY

        if temp_only:
            local_period_path = tempfile.mkdtemp(prefix="excel_tmp_")
        else:
            local_period_path = cls.get_local_period_folder(date_obj)

        remote_period_path = cls.get_smb_period_folder(date_obj) if upload else None

        local_paths = {}
        files_to_upload = []

        # Determine if we have multi-department data or single-department
        # Multi-dept data has structure: {dept_code: {dept_code, dept_name, data: [...]}}

        if daily_data:
            # Check if this is already grouped by department
            if daily_data and isinstance(daily_data, dict) and "data" not in daily_data and dept_code is None:
                # It's grouped by department
                data_by_dept = daily_data
            else:
                # Single department or ungrouped - wrap it
                data_by_dept = {
                    dept_code or cls.DEFAULT_DEPT_CODE: {
                        "dept_code": dept_code or cls.DEFAULT_DEPT_CODE,
                        "dept_name": dept_name or cls.DEFAULT_DEPT_NAME,
                        "data": daily_data if isinstance(daily_data, list) else (daily_data.get("data") if isinstance(daily_data, dict) else []),
                    }
                }

            daily_form_path = os.path.join(local_period_path, f"{date_str}OT.xlsx")
            daily_summary_path = os.path.join(local_period_path, f"{date_str}OTSummary.xlsx")

            wb_form = cls.create_ot_form_multi_sheet(data_by_dept, date_obj)
            wb_summary = cls.create_ot_summary_multi_sheet(data_by_dept, date_obj)

            wb_form.save(daily_form_path)
            wb_summary.save(daily_summary_path)

            local_paths["daily_form"] = daily_form_path
            local_paths["daily_summary"] = daily_summary_path

            if remote_period_path:
                files_to_upload.append((daily_form_path, f"{remote_period_path}/{date_str}OT.xlsx"))
                files_to_upload.append((daily_summary_path, f"{remote_period_path}/{date_str}OTSummary.xlsx"))

        if monthly_data:
            # Check if this is already grouped by department
            if monthly_data and isinstance(monthly_data, dict) and "data" not in monthly_data and dept_code is None:
                # It's grouped by department
                data_by_dept = monthly_data
            else:
                # Single department or ungrouped - wrap it
                data_by_dept = {
                    dept_code or cls.DEFAULT_DEPT_CODE: {
                        "dept_code": dept_code or cls.DEFAULT_DEPT_CODE,
                        "dept_name": dept_name or cls.DEFAULT_DEPT_NAME,
                        "data": monthly_data if isinstance(monthly_data, list) else (monthly_data.get("data") if isinstance(monthly_data, dict) else []),
                    }
                }

            current_period_start = date_obj.replace(day=26)
            if date_obj.day < 26:
                current_period_start = (date_obj.replace(day=1) - timedelta(days=1)).replace(day=26)
            next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)

            monthly_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OT.xlsx"
            monthly_summary_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OTSummary.xlsx"

            monthly_form_path = os.path.join(local_period_path, monthly_filename)
            monthly_summary_path = os.path.join(local_period_path, monthly_summary_filename)

            wb_monthly_form = cls.create_monthly_ot_form_multi_sheet(data_by_dept, current_period_start, next_period_end)
            wb_monthly_summary = cls.create_monthly_ot_summary_multi_sheet(data_by_dept, current_period_start, next_period_end)

            wb_monthly_form.save(monthly_form_path)
            wb_monthly_summary.save(monthly_summary_path)

            local_paths["monthly_form"] = monthly_form_path
            local_paths["monthly_summary"] = monthly_summary_path

            if remote_period_path:
                files_to_upload.append((monthly_form_path, f"{remote_period_path}/{monthly_filename}"))
                files_to_upload.append((monthly_summary_path, f"{remote_period_path}/{monthly_summary_filename}"))

        if upload and files_to_upload and cls._smb_configured():
            conn = cls.get_smb_connection()
            if conn:
                try:
                    cls.ensure_folder_exists(remote_period_path, conn)
                    for local_path, remote_path in files_to_upload:
                        with open(local_path, "rb") as file_obj:
                            conn.storeFile(cls.SMB_CONFIG["share_name"], remote_path, file_obj)
                finally:
                    conn.close()

        if upload and temp_only:
            for path in local_paths.values():
                try:
                    if path and os.path.exists(path):
                        os.remove(path)
                except Exception as exc:
                    print(f"Error cleaning temp file {path}: {exc}")
            try:
                if os.path.isdir(local_period_path):
                    for root, _, files in os.walk(local_period_path, topdown=False):
                        for name in files:
                            try:
                                os.remove(os.path.join(root, name))
                            except Exception:
                                pass
                        try:
                            os.rmdir(root)
                        except Exception:
                            pass
            except Exception as exc:
                print(f"Error cleaning temp directory {local_period_path}: {exc}")

        return local_paths

    @classmethod
    def delete_files_batch(cls, files_to_delete):
        if not files_to_delete:
            return

        for path in files_to_delete:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                    print(f"Deleted local file: {path}")
                except Exception as exc:
                    print(f"Error deleting local file {path}: {exc}")

        remote_candidates = [path for path in files_to_delete if path and not os.path.exists(path)]
        if not remote_candidates or not cls._smb_configured():
            return

        conn = cls.get_smb_connection()
        if conn is None:
            return
        try:
            for remote_path in remote_candidates:
                try:
                    conn.deleteFiles(cls.SMB_CONFIG["share_name"], remote_path)
                    print(f"Deleted remote file: {remote_path}")
                except Exception as exc:
                    print(f"Error deleting remote file {remote_path}: {exc}")
        finally:
            conn.close()
