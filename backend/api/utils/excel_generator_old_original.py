import calendar
import os
import socket
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from smb.SMBConnection import SMBConnection


class ExcelHandler:
    def save_file(self, wb, filepath):
        """Save Excel file with enhanced error handling"""
        print(f"Attempting to save file: {filepath}")
        temp_path = f"{filepath}.tmp"

        try:
            # Save to temp file first
            print(f"Saving to temporary file: {temp_path}")
            wb.save(temp_path)

            # Ensure directory exists
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            # Move temp file to final location
            print("Moving temp file to final location")
            os.replace(temp_path, filepath)

            print(f"File saved successfully: {filepath}")
            return True

        except Exception as e:
            print(f"Save failed: {str(e)}")
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                    print("Temporary file cleaned up")
                except Exception:
                    pass
            return False


class ExcelGenerator:
    # Load environment variables
    load_dotenv()

    # Constants
    DEPT_CODE = "K390140R1C"
    DEPT_NAME = "BG6-RD Center-Automatic System Test R&D Div.1-Dept.1-PTB Sec.1"

    # SMB Configuration
    SMB_CONFIG = {"host": os.getenv("CIFS_HOST"), "username": os.getenv("CIFS_USER"), "password": os.getenv("CIFS_PASSWORD"), "share_name": os.getenv("CIFS_SHARE"), "path": os.getenv("CIFS_PATH").replace("\\", "/")}

    # Connection caching
    _connection_cache = None
    _connection_cache_time = None
    _connection_cache_timeout = 300  # 5 minutes

    @classmethod
    def get_smb_connection(cls, use_cache=True):
        """Create SMB connection with caching and optimization"""
        current_time = time.time()

        # Check if we have a valid cached connection
        if use_cache and cls._connection_cache and cls._connection_cache_time and (current_time - cls._connection_cache_time) < cls._connection_cache_timeout:
            try:
                # Test if connection is still alive with a quick operation
                cls._connection_cache.listShares()
                return cls._connection_cache
            except Exception:
                cls._connection_cache = None
                cls._connection_cache_time = None

        try:
            client_name = socket.gethostname()

            # Optimized connection - try the most common working method first
            conn = SMBConnection(username=cls.SMB_CONFIG["username"], password=cls.SMB_CONFIG["password"], my_name=client_name, remote_name=cls.SMB_CONFIG["host"], use_ntlm_v2=True, is_direct_tcp=True)

            if conn.connect(cls.SMB_CONFIG["host"], 445, timeout=15):  # Reduced timeout
                if use_cache:
                    cls._connection_cache = conn
                    cls._connection_cache_time = current_time
                return conn
            else:
                conn.close()
                raise Exception("Connection returned False")

        except Exception as e:
            print(f"SMB connection error: {str(e)}")
            raise

    @classmethod
    def get_period_folder(cls, date=None):
        """Generate folder name based on period dates - optimized version"""
        if date is None:
            date = datetime.now()

        current_period_start = date.replace(day=26)
        if date.day < 26:
            current_period_start = (date.replace(day=1) - timedelta(days=1)).replace(day=26)

        next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)
        folder_name = f"{current_period_start.strftime('%Y-%m-%d')}_{next_period_end.strftime('%Y-%m-%d')}"

        # Construct path using forward slashes for SMB
        smb_base_path = cls.SMB_CONFIG["path"].rstrip("/")
        relative_path = f"{smb_base_path}/{folder_name}"

        return relative_path

    @classmethod
    def ensure_folder_exists(cls, relative_path, conn=None):
        """Ensure monthly period folder exists (base path already exists)"""
        should_close_conn = False
        if conn is None:
            conn = cls.get_smb_connection()
            should_close_conn = True

        try:
            # Only check/create the period folder since base path exists
            try:
                conn.listPath(cls.SMB_CONFIG["share_name"], relative_path)
            except Exception:
                # Period folder doesn't exist, create it
                try:
                    conn.createDirectory(cls.SMB_CONFIG["share_name"], relative_path)
                except Exception as dir_error:
                    error_msg = str(dir_error).lower()
                    if not ("already exists" in error_msg or "file exists" in error_msg):
                        print(f"Warning: Could not create period directory {relative_path}: {str(dir_error)}")

        finally:
            if should_close_conn:
                conn.close()

    @classmethod
    def generate_excel_files(cls, data, date_obj):
        """Optimized Excel generation with connection reuse"""
        if not data:
            return None, None

        try:
            date_str = date_obj.strftime("%Y%m%d")
            period_path = cls.get_period_folder(date_obj)

            with tempfile.TemporaryDirectory() as temp_dir:
                excel_path = os.path.join(temp_dir, f"{date_str}OT.xlsx")
                summary_path = os.path.join(temp_dir, f"{date_str}OTSummary.xlsx")

                # Create and save workbooks
                wb_form = cls.create_ot_form(data, date_obj)
                wb_summary = cls.create_ot_summary(data, date_obj)

                wb_form.save(excel_path)
                wb_summary.save(summary_path)

                # Upload with single connection
                conn = cls.get_smb_connection()
                try:
                    # Ensure folder exists using the same connection
                    cls.ensure_folder_exists(period_path, conn)

                    # Upload files
                    remote_excel = f"{period_path}/{date_str}OT.xlsx"
                    remote_summary = f"{period_path}/{date_str}OTSummary.xlsx"

                    with open(excel_path, "rb") as file:
                        conn.storeFile(cls.SMB_CONFIG["share_name"], remote_excel, file)

                    with open(summary_path, "rb") as file:
                        conn.storeFile(cls.SMB_CONFIG["share_name"], remote_summary, file)

                    return remote_excel, remote_summary

                finally:
                    conn.close()

        except Exception as e:
            print(f"Excel generation failed: {str(e)}")
            raise

    @classmethod
    def generate_monthly_excel_files(cls, monthly_data, date_obj):
        """Optimized monthly Excel generation with connection reuse"""
        if not monthly_data:
            return None, None

        try:
            # Get period dates for file naming
            current_period_start = date_obj.replace(day=26)
            if date_obj.day < 26:
                current_period_start = (date_obj.replace(day=1) - timedelta(days=1)).replace(day=26)

            next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)

            # Create filename in format ~YYYY_MM_DD-YYYY_MM_DDOT.xlsx
            monthly_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OT.xlsx"
            monthly_summary_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OTSummary.xlsx"

            period_path = cls.get_period_folder(date_obj)

            with tempfile.TemporaryDirectory() as temp_dir:
                monthly_excel_path = os.path.join(temp_dir, monthly_filename)
                monthly_summary_path = os.path.join(temp_dir, monthly_summary_filename)

                # Create and save monthly workbooks
                wb_monthly_form = cls.create_monthly_ot_form(monthly_data, current_period_start, next_period_end)
                wb_monthly_summary = cls.create_monthly_ot_summary(monthly_data, current_period_start, next_period_end)

                wb_monthly_form.save(monthly_excel_path)
                wb_monthly_summary.save(monthly_summary_path)

                # Upload with single connection
                conn = cls.get_smb_connection()
                try:
                    # Ensure folder exists using the same connection
                    cls.ensure_folder_exists(period_path, conn)

                    # Upload files
                    remote_monthly_excel = f"{period_path}/{monthly_filename}"
                    remote_monthly_summary = f"{period_path}/{monthly_summary_filename}"

                    with open(monthly_excel_path, "rb") as file:
                        conn.storeFile(cls.SMB_CONFIG["share_name"], remote_monthly_excel, file)

                    with open(monthly_summary_path, "rb") as file:
                        conn.storeFile(cls.SMB_CONFIG["share_name"], remote_monthly_summary, file)

                    return remote_monthly_excel, remote_monthly_summary

                finally:
                    conn.close()

        except Exception as e:
            print(f"Monthly Excel generation failed: {str(e)}")
            raise

    @classmethod
    def is_file_locked(cls, filepath):
        """Check if Excel file is currently open"""
        if not Path(filepath).exists():
            return False

        try:
            # Try to open file in write mode
            with open(filepath, "a+b") as f:
                f.seek(0)  # Try to seek to validate write access
                return False
        except (OSError, PermissionError):
            print(f"File {filepath} is locked or in use")
            return True

    @classmethod
    def _delete_files(cls, date_obj):
        """Delete existing Excel files for given date"""
        try:
            excel_file = f"{cls.NETWORK_PATH}{date_obj.strftime('%Y%m%d')}OT.xlsx"
            summary_file = f"{cls.NETWORK_PATH}{date_obj.strftime('%Y%m%d')}OTSummary.xlsx"

            for filepath in [excel_file, summary_file]:
                if os.path.exists(filepath) and not cls.is_file_locked(filepath):
                    os.remove(filepath)
                    print(f"Deleted: {filepath}")
        except Exception as e:
            print(f"Error deleting files: {str(e)}")

    @classmethod
    def create_ot_form(cls, data, date_obj):
        """Create overtime form workbook"""
        wb = Workbook()
        ws = wb.active

        # Column widths setup
        columns = {"A": 4.14, "B": 13.43, "C": 21.14, "D": 22.29, "E": 22.86, "F": 12.84, "G": 15.14, "H": 13.57, "I": 15.71, "J": 1.00, "K": 17.14, "L": 14.43, "M": 15.43, "N": 18.57}
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        # Title
        ws.merge_cells("A1:N1")
        ws["A1"] = "Form Lembur (? ? ? ? ?)"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(name="Arial", size=20)

        # Department info
        ws.merge_cells("A2:C2")
        ws["A2"] = f"Departemen\n(????):\n{cls.DEPT_CODE}\n{cls.DEPT_NAME}"
        ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["A2"].font = Font(name="Arial", size=11)

        # Employee classification
        ws["D2"] = "Klasifikasi Karyawan\n(????):"
        ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
        ws["E2"] = "? Band0(1-2??)\n? Band1(3-5??)"
        ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

        # Date information
        ch_date = f"{date_obj.year} ? {date_obj.month:02d} ? {date_obj.day:02d} ?"
        weekday = calendar.day_name[date_obj.weekday()].upper()
        ws.merge_cells("F2:H2")
        ws["F2"] = f"Tanggal Lembur (????):\n                           {ch_date}\nHari (??): {weekday}"
        ws["F2"].alignment = Alignment(vertical="center", wrap_text=True)

        print("Excel data received:", data)
        print("First request:", data[0] if data else "No data")

        print("DEBUG - Weekend Check:")
        print(f"Date: {date_obj}")
        print(f"Weekday number: {date_obj.weekday()}")  # 0=Monday, 5=Saturday, 6=Sunday
        print(f"Is weekend?: {date_obj.weekday() >= 5}")

        if data and len(data) > 0:
            first_request = data[0]
            print("First request data:", first_request)
            print("is_weekend value:", first_request.get("is_weekend"))
            print("is_holiday value:", first_request.get("is_holiday"))

        # Overtime type based on is_weekend/is_holiday
        overtime_type = "Jenis Lembur (????) :\n"
        if data and len(data) > 0:
            first_request = data[0]
            is_weekend = first_request.get("is_weekend", False)
            is_holiday = first_request.get("is_holiday", False)

            if is_holiday:
                overtime_type += "? Saat Hari Kerja (???????)\n? Saat Hari Libur (?????)\n? Saat Tanggal Merah (??????)"
            elif is_weekend:
                overtime_type += "? Saat Hari Kerja (???????)\n? Saat Hari Libur (?????)\n? Saat Tanggal Merah (??????)"
            else:
                overtime_type += "? Saat Hari Kerja (???????)\n? Saat Hari Libur (?????)\n? Saat Tanggal Merah (??????)"

        ws.merge_cells("I2:N2")
        ws["I2"] = overtime_type
        ws["I2"].alignment = Alignment(vertical="center", wrap_text=True)

        # Column headers
        headers = [
            ("A3", "No"),
            ("B3", "No Karyawan\n(??)"),
            ("C3", "Nama\n(??)"),
            ("D3", "Alasan Lembur\n(??????)"),
            ("E3", "Jam Lembur (Waktu)\n(????\n????)"),
            ("F3", "Durasi Lembur\n(??????)"),
            ("G3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(????????V)"),
            ("H3", "Tanda Tangan Karyawan\n(????)"),
            ("I3", "Tanda Tangan Supervisor\n(??????)"),
            ("K3", "Konfirmasi Jam Lembur (Waktu)\n(????\n????)"),
            ("L3", "Konfirmasi Durasi Lembur\n(??????)"),
            ("M3", "Jam Istirahat (Jika ada, gunakan V)\n(????????V)"),
            ("N3", "Tanda Tangan Karyawan\n(????)"),
        ]

        # Apply header formatting
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        # Empty rows with formatting
        for row in range(4, 34):
            for col in "ABCDEFGHIKLMN":
                cell = f"{col}{row}"
                ws[cell] = ""
                ws[cell].border = thin_border
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
            ws[f"A{row}"] = row - 3  # Add row numbers

        # Fill in data
        current_row = 4
        for item in data:
            ws[f"B{current_row}"] = item["employee_id"]
            ws[f"C{current_row}"] = item["employee_name"]

            # Centered cells
            ws[f"D{current_row}"] = item["reason"]
            ws[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"E{current_row}"] = f"{item['time_start']} - {item['time_end']}"
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"F{current_row}"] = f"{item['total_hours']} hour(s)"
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee take break time
            ws[f"G{current_row}"] = "V" if item["has_break"] else "-"
            ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Actual time, duration and break status
            ws[f"K{current_row}"] = f"{item['time_start']} - {item['time_end']}"
            ws[f"K{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"L{current_row}"] = f"{item['total_hours']} hour(s)"
            ws[f"L{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee take break time
            ws[f"M{current_row}"] = "V" if item["has_break"] else "-"
            ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        # Footer notes
        ws.merge_cells("A34:N34")
        ws["A34"] = "Keterangan:"
        ws["A34"].font = Font(name="Arial", size=12)

        notes = [
            ("A35:N35", "1?Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(?????????,???????????)?"),
            ("A36:N36", "2?Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (??????,?????????)?"),
            ("A37:N37", "3?Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (????????????) ?"),
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            ws[cells.split(":")[0]] = text
            ws[cells.split(":")[0]].font = Font(name="Arial", size=12)
            ws[cells.split(":")[0]].alignment = Alignment(vertical="center", wrap_text=True)

        # Form number
        ws["M38"] = "Form No.:PTB-TB004-001 Rev.01"
        ws["M38"].font = Font(name="Arial", size=10)

        # Save workbook
        return wb

    @classmethod
    def create_ot_summary(cls, data, date_obj):
        """Create overtime summary workbook"""
        wb = Workbook()
        ws = wb.active

        # Set column widths
        columns = {
            "A": 11.5,  # Work ID
            "B": 12.0,  # Overtime Type
            "C": 15.0,  # Overtime Start Date
            "D": 9.75,  # Start Time
            "E": 9.75,  # End Time
            "F": 9.75,  # Break
            "G": 6.0,  # Hours
            "H": 25.0,  # Reason
        }

        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Headers with formatting
        headers = [("A1", "Work ID"), ("B1", "Overtime Type"), ("C1", "Overtime Start Date"), ("D1", "Start Time"), ("E1", "End Time"), ("F1", "Break"), ("G1", "Hours"), ("H1", "Reason")]

        header_style = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        # Fill data
        current_row = 2
        for item in data:
            # Determine overtime type based on weekend/holiday
            ot_type = 1  # Default: weekday
            if item.get("is_holiday", False):
                ot_type = 3  # Holiday
            elif item.get("is_weekend", False):
                ot_type = 2  # Weekend (Saturday)

            # Format date from the date_obj
            formatted_date = date_obj.strftime("%Y-%m-%d")

            ws[f"A{current_row}"] = item["employee_id"]
            ws[f"B{current_row}"] = ot_type
            ws[f"C{current_row}"] = formatted_date
            ws[f"D{current_row}"] = item["time_start"]
            ws[f"E{current_row}"] = item["time_end"]
            ws[f"F{current_row}"] = "Y" if item["has_break"] else "N"

            # Format total_hours according to requirements
            try:
                # Convert to float if it's a string
                if isinstance(item["total_hours"], str):
                    hours_value = float(item["total_hours"])
                else:
                    hours_value = item["total_hours"]

                # Apply formatting rules
                if hours_value == int(hours_value):
                    # For whole numbers (3.00 ? 3)
                    formatted_hours = str(int(hours_value))
                elif hours_value * 10 % 10 == 0:
                    # For values with .5 (3.50 ? 3.5)
                    formatted_hours = str(hours_value).rstrip("0").rstrip(".")
                else:
                    # For other values (3.05, 3.55 ? keep as is)
                    formatted_hours = str(hours_value)
            except (ValueError, TypeError):
                # Fallback if conversion fails
                formatted_hours = str(item["total_hours"])

            ws[f"G{current_row}"] = formatted_hours
            ws[f"H{current_row}"] = item["reason"]

            # Formatting
            for col in "ABCDEFGH":
                cell = f"{col}{current_row}"
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
                if col in "ABCDEFG":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        # Save
        return wb

    @classmethod
    def create_monthly_ot_form(cls, monthly_data, period_start, period_end):
        """Create monthly overtime form workbook with additional columns"""
        wb = Workbook()
        ws = wb.active

        # Column widths setup
        columns = {"A": 4.14, "B": 13.88, "C": 13.93, "D": 21.14, "E": 22.29, "F": 15.22, "G": 16.55, "H": 14.55, "I": 15.14, "J": 13.57, "K": 15.71, "L": 1.00, "M": 16.55, "N": 14.55, "O": 15.43, "P": 18.57}
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        # Title
        ws.merge_cells("A1:P1")
        ws["A1"] = "Form Lembur Bulanan (???????)"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(name="Arial", size=20, bold=True)

        # Department info
        ws.merge_cells("A2:C2")
        ws["A2"] = f"Departemen\n(????):\n{cls.DEPT_CODE}\n{cls.DEPT_NAME}"
        ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        ws["A2"].font = Font(name="Arial", size=11)

        # Employee classification
        ws["D2"] = "Klasifikasi Karyawan\n(????):"
        ws["D2"].alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
        ws["E2"] = "? Band0(1-2??)\n? Band1(3-5??)"
        ws["E2"].alignment = Alignment(vertical="center", wrap_text=True)

        # Period information
        period_str = f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')}"
        ws.merge_cells("F2:H2")
        ws["F2"] = f"Periode Lembur (????):\n{period_str}"
        ws["F2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Empty right section for monthly form
        ws.merge_cells("I2:P2")
        ws["I2"] = ""

        # Column headers
        headers = [
            ("A3", "No"),
            ("B3", "Tanggal Lembur\n(????)"),
            ("C3", "No Karyawan\n(??)"),
            ("D3", "Nama\n(??)"),
            ("E3", "Alasan Lembur\n(??????)"),
            ("F3", "Jenis Lembur\n(????)"),
            ("G3", "Jam Lembur (Waktu)\n(????\n????)"),
            ("H3", "Durasi Lembur\n(??????)"),
            ("I3", "Jam istirahat saat lembur (jika perlu, gunakan V)\n(????????V)"),
            ("J3", "Tanda Tangan Karyawan\n(????)"),
            ("K3", "Tanda Tangan Supervisor\n(??????)"),
            ("M3", "Konfirmasi Jam Lembur (Waktu)\n(????\n????)"),
            ("N3", "Konfirmasi Durasi Lembur\n(??????)"),
            ("O3", "Jam Istirahat (Jika ada, gunakan V)\n(????????V)"),
            ("P3", "Tanda Tangan Karyawan\n(????)"),
        ]

        # Apply header formatting
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        header_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        # Calculate required rows based on data size
        # Minimum 30 rows, but expand if needed for data
        data_row_count = len(monthly_data) if monthly_data else 0
        min_rows = 30
        required_rows = max(min_rows, data_row_count)
        last_data_row = 4 + required_rows - 1  # Row 4 is first data row

        print(f"Monthly Excel: Data count={data_row_count}, Required rows={required_rows}, Last data row={last_data_row}")

        # Create empty rows with formatting (dynamic row count)
        for row in range(4, last_data_row + 1):
            for col in "ABCDEFGHIJKLMNOP":
                cell = f"{col}{row}"
                ws[cell] = ""
                ws[cell].border = thin_border
                ws[cell].font = Font(name="Arial", size=11)
                # Center columns A, B, C, D; others just vertical center
                if col in "ABCD":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws[cell].alignment = Alignment(vertical="center")
            ws[f"A{row}"] = row - 3  # Add row numbers

        # Fill in monthly data
        current_row = 4
        for item in monthly_data:
            ws[f"B{current_row}"] = item["request_date"]
            ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"C{current_row}"] = item["employee_id"]
            ws[f"C{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"D{current_row}"] = item["employee_name"]
            ws[f"D{current_row}"].alignment = Alignment(vertical="center")

            ws[f"E{current_row}"] = item["reason"]
            ws[f"E{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Overtime type based on weekend/holiday status
            if item.get("is_holiday", False):
                overtime_type = "Holiday"
            elif item.get("is_weekend", False):
                overtime_type = "Weekend"
            else:
                overtime_type = "Weekday"

            ws[f"F{current_row}"] = overtime_type
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"G{current_row}"] = f"{item['time_start']} - {item['time_end']}"
            ws[f"G{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"H{current_row}"] = f"{item['total_hours']} hour(s)"
            ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee take break time
            ws[f"I{current_row}"] = "V" if item["has_break"] else "-"
            ws[f"I{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Actual time, duration and break status
            ws[f"M{current_row}"] = f"{item['time_start']} - {item['time_end']}"
            ws[f"M{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            ws[f"N{current_row}"] = f"{item['total_hours']} hour(s)"
            ws[f"N{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Show 'V' if employee take break time
            ws[f"O{current_row}"] = "V" if item["has_break"] else "-"
            ws[f"O{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        # Dynamic footer positioning - place after all data rows
        footer_row = last_data_row + 1
        notes_start_row = footer_row + 1
        form_number_row = notes_start_row + 3

        # Footer notes (dynamic positioning)
        ws.merge_cells(f"A{footer_row}:P{footer_row}")
        ws[f"A{footer_row}"] = "Keterangan:"
        ws[f"A{footer_row}"].font = Font(name="Arial", size=12)

        notes = [
            (f"A{notes_start_row}:P{notes_start_row}", "1?Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(?????????,???????????)?"),
            (f"A{notes_start_row + 1}:P{notes_start_row + 1}", "2?Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (??????,?????????)?"),
            (f"A{notes_start_row + 2}:P{notes_start_row + 2}", "3?Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (????????????)?"),
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            ws[cells.split(":")[0]] = text
            ws[cells.split(":")[0]].font = Font(name="Arial", size=12)
            ws[cells.split(":")[0]].alignment = Alignment(vertical="center", wrap_text=True)

        # Form number (dynamic positioning)
        ws[f"O{form_number_row}"] = "Form No.:PTB-TB004-001 Rev.01"
        ws[f"O{form_number_row}"].font = Font(name="Arial", size=10)

        return wb

    @classmethod
    def create_monthly_ot_summary(cls, monthly_data, period_start, period_end):
        """Create monthly overtime summary workbook with additional Detail column"""
        wb = Workbook()
        ws = wb.active

        # Set column widths
        columns = {
            "A": 11.5,  # Work ID
            "B": 12.0,  # Overtime Type
            "C": 15.0,  # Overtime Start Date
            "D": 9.75,  # Start Time
            "E": 9.75,  # End Time
            "F": 9.75,  # Break
            "G": 6.0,  # Hours
            "H": 25.0,  # Reason
            "I": 25.0,  # Detail
        }

        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Headers with formatting
        headers = [("A1", "Work ID"), ("B1", "Overtime Type"), ("C1", "Overtime Start Date"), ("D1", "Start Time"), ("E1", "End Time"), ("F1", "Break"), ("G1", "Hours"), ("H1", "Reason"), ("I1", "Detail")]

        header_style = Alignment(horizontal="center", vertical="center")
        header_font = Font(name="Arial", size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        # Fill monthly data
        current_row = 2
        for item in monthly_data:
            # Determine overtime type based on weekend/holiday
            ot_type = 1  # Default: weekday
            if item.get("is_holiday", False):
                ot_type = 3  # Holiday
            elif item.get("is_weekend", False):
                ot_type = 2  # Weekend (Saturday)

            ws[f"A{current_row}"] = item["employee_id"]
            ws[f"B{current_row}"] = ot_type
            ws[f"C{current_row}"] = item["request_date"]
            ws[f"D{current_row}"] = item["time_start"]
            ws[f"E{current_row}"] = item["time_end"]
            ws[f"F{current_row}"] = "Y" if item["has_break"] else "N"

            # Format total_hours
            try:
                if isinstance(item["total_hours"], str):
                    hours_value = float(item["total_hours"])
                else:
                    hours_value = item["total_hours"]

                if hours_value == int(hours_value):
                    formatted_hours = str(int(hours_value))
                elif hours_value * 10 % 10 == 0:
                    formatted_hours = str(hours_value).rstrip("0").rstrip(".")
                else:
                    formatted_hours = str(hours_value)
            except (ValueError, TypeError):
                formatted_hours = str(item["total_hours"])

            ws[f"G{current_row}"] = formatted_hours
            ws[f"H{current_row}"] = item["reason"]
            ws[f"I{current_row}"] = item.get("detail", "")

            # Formatting
            for col in "ABCDEFGHI":
                cell = f"{col}{current_row}"
                ws[cell].font = Font(name="Arial", size=11)
                ws[cell].alignment = Alignment(vertical="center")
                if col in "ABCDEFG":
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1

        return wb

    @classmethod
    def generate_all_excel_files(cls, daily_data, monthly_data, date_obj):
        """Optimized batch generation of both daily and monthly files using single connection"""
        if not daily_data and not monthly_data:
            return None, None, None, None

        try:
            # Prepare file names and paths
            date_str = date_obj.strftime("%Y%m%d")
            period_path = cls.get_period_folder(date_obj)

            # Monthly file names
            current_period_start = date_obj.replace(day=26)
            if date_obj.day < 26:
                current_period_start = (date_obj.replace(day=1) - timedelta(days=1)).replace(day=26)
            next_period_end = (current_period_start + timedelta(days=32)).replace(day=25)

            monthly_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OT.xlsx"
            monthly_summary_filename = f"~{current_period_start.strftime('%Y_%m_%d')}-{next_period_end.strftime('%Y_%m_%d')}OTSummary.xlsx"

            # Results placeholders
            daily_results = (None, None)
            monthly_results = (None, None)

            with tempfile.TemporaryDirectory() as temp_dir:
                files_to_upload = []

                # Generate daily files if data exists
                if daily_data:
                    excel_path = os.path.join(temp_dir, f"{date_str}OT.xlsx")
                    summary_path = os.path.join(temp_dir, f"{date_str}OTSummary.xlsx")

                    wb_form = cls.create_ot_form(daily_data, date_obj)
                    wb_summary = cls.create_ot_summary(daily_data, date_obj)

                    wb_form.save(excel_path)
                    wb_summary.save(summary_path)

                    remote_excel = f"{period_path}/{date_str}OT.xlsx"
                    remote_summary = f"{period_path}/{date_str}OTSummary.xlsx"

                    files_to_upload.extend([(excel_path, remote_excel), (summary_path, remote_summary)])

                    daily_results = (remote_excel, remote_summary)

                # Generate monthly files if data exists
                if monthly_data:
                    monthly_excel_path = os.path.join(temp_dir, monthly_filename)
                    monthly_summary_path = os.path.join(temp_dir, monthly_summary_filename)

                    wb_monthly_form = cls.create_monthly_ot_form(monthly_data, current_period_start, next_period_end)
                    wb_monthly_summary = cls.create_monthly_ot_summary(monthly_data, current_period_start, next_period_end)

                    wb_monthly_form.save(monthly_excel_path)
                    wb_monthly_summary.save(monthly_summary_path)

                    remote_monthly_excel = f"{period_path}/{monthly_filename}"
                    remote_monthly_summary = f"{period_path}/{monthly_summary_filename}"

                    files_to_upload.extend([(monthly_excel_path, remote_monthly_excel), (monthly_summary_path, remote_monthly_summary)])

                    monthly_results = (remote_monthly_excel, remote_monthly_summary)

                # Upload all files with single connection
                if files_to_upload:
                    conn = cls.get_smb_connection()
                    try:
                        # Ensure folder exists once
                        cls.ensure_folder_exists(period_path, conn)

                        # Upload all files
                        for local_path, remote_path in files_to_upload:
                            with open(local_path, "rb") as file:
                                conn.storeFile(cls.SMB_CONFIG["share_name"], remote_path, file)

                    finally:
                        conn.close()

                return daily_results[0], daily_results[1], monthly_results[0], monthly_results[1]

        except Exception as e:
            print(f"Batch Excel generation failed: {str(e)}")
            raise

    @classmethod
    def clear_connection_cache(cls):
        """Clear the connection cache and close existing connection"""
        if cls._connection_cache:
            try:
                cls._connection_cache.close()
            except Exception:
                pass
            cls._connection_cache = None
            cls._connection_cache_time = None
            print("Connection cache cleared")

    @classmethod
    def delete_files_batch(cls, files_to_delete):
        """Delete multiple files using single connection"""
        if not files_to_delete:
            return

        try:
            conn = cls.get_smb_connection()
            try:
                for file_path in files_to_delete:
                    try:
                        conn.deleteFiles(cls.SMB_CONFIG["share_name"], file_path)
                        print(f"Deleted: {file_path}")
                    except Exception as e:
                        print(f"Error deleting {file_path}: {str(e)}")
            finally:
                conn.close()
        except Exception as e:
            print(f"Error in batch file deletion: {str(e)}")
