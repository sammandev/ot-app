import logging
from datetime import datetime

from django.contrib.auth import authenticate, get_user_model
from django.core.cache import cache
from django.db import models, transaction
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import AuthenticationFailed, PermissionDenied
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle
from rest_framework.views import APIView
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from rest_framework_simplejwt.tokens import RefreshToken

from .cookies import clear_auth_cookies, set_auth_cookies
from .models import (
    Asset,
    BoardPresence,
    CalendarEvent,
    Department,
    Employee,
    EmployeeLeave,
    ExternalUser,
    Holiday,
    Notification,
    OvertimeLimitConfig,
    OvertimeRegulation,
    OvertimeRegulationDocument,
    OvertimeRequest,
    PersonalNote,
    Project,
    PurchaseRequest,
    ReleaseNote,
    SMBConfiguration,
    SystemConfiguration,
    TaskActivity,
    TaskAttachment,
    TaskComment,
    TaskGroup,
    TaskReminder,
    TaskSubtask,
    TaskTimeLog,
    UserActivityLog,
    UserReport,
    UserSession,
)
from .pagination import CalendarEventPagination, EmployeePagination, OvertimeRequestPagination, ProjectPagination
from .permissions import IsSuperAdmin, ResourcePermission
from .serializers import (
    AssetSerializer,
    AssetSummarySerializer,
    BoardPresenceSerializer,
    CalendarEventSerializer,
    DepartmentSerializer,
    EmployeeLeaveSerializer,
    EmployeeSerializer,
    HolidaySerializer,
    NotificationSerializer,
    OvertimeLimitConfigSerializer,
    OvertimeRegulationDocumentSerializer,
    OvertimeRegulationSerializer,
    OvertimeSerializer,
    PersonalNoteSerializer,
    ProjectSerializer,
    PurchaseRequestSerializer,
    ReleaseNoteSerializer,
    SMBConfigurationSerializer,
    SystemConfigurationSerializer,
    TaskActivitySerializer,
    TaskAttachmentSerializer,
    TaskCommentSerializer,
    TaskGroupSerializer,
    TaskReminderSerializer,
    TaskSubtaskSerializer,
    TaskTimeLogSerializer,
    UserActivityLogSerializer,
    UserReportAdminSerializer,
    UserReportSerializer,
)
from .serializers_access import UserAccessSerializer, UserAccessUpdateSerializer
from .services.bulk_service import BulkImportExportService
from .services.cache_service import CacheService, cache_invalidate_on_change, cached_list
from .services.employee_service import get_employee_queryset
from .services.external_auth import ExternalAuthService
from .services.overtime_service import get_overtime_queryset
from .services.project_service import get_project_queryset
from .utils.excel_generator import ExcelGenerator

logger = logging.getLogger(__name__)
User = get_user_model()


def is_superadmin_user(user):
    """
    Check if a user is a superadmin (either developer or superadmin role).
    Uses the 'role' field on ExternalUser.
    """
    if not user:
        return False

    role = getattr(user, "role", "") or ""
    return role in ("developer", "superadmin")


def is_developer_user(user):
    """
    Check if a user is a developer (highest privilege, cannot be modified).
    Uses the 'role' field on ExternalUser.
    """
    if not user:
        return False

    role = getattr(user, "role", "") or ""
    return role == "developer"


def is_ptb_admin(user):
    """
    Check if a user has PTB admin privileges.
    """
    if not user:
        return False
    return bool(getattr(user, "is_ptb_admin", False))


def get_employee_for_user(user, raise_if_not_found=True):
    """
    Get the Employee record associated with the current user.

    Matching strategy (in order of priority):
    1. Match Employee.emp_id with User.worker_id
    2. Match Employee.name with User's full name (first_name + last_name)
    3. Match Employee.name with User.username (handles firstname_lastname format)

    If raise_if_not_found is True and employee is not found, raises ValidationError.
    If raise_if_not_found is False and employee is not found, returns None.

    This function also tries to auto-create an Employee record if the user has
    a valid worker_id but no matching employee exists.

    Uses per-request caching on the user object to avoid repeated DB lookups.
    """
    if not user:
        if raise_if_not_found:
            raise serializers.ValidationError({"error": "User not authenticated. Please log in again."})
        return None

    # Per-request cache: avoid re-querying within the same request
    cache_attr = "_employee_cache"
    if hasattr(user, cache_attr):
        cached = getattr(user, cache_attr)
        if cached is not None or not raise_if_not_found:
            return cached

    employee = None

    # Build a single consolidated query with Q objects to replace
    # the sequential per-strategy lookup (avoids up to 5 queries).
    from django.db.models import Q

    q_filter = Q()

    # Strategy 1: emp_id == worker_id
    worker_id = getattr(user, "worker_id", None)
    if worker_id:
        q_filter |= Q(emp_id=worker_id)

    # Strategy 2: name == full_name
    full_name = f"{user.first_name} {user.last_name}".strip()
    if full_name:
        q_filter |= Q(name__iexact=full_name)

    # Strategy 3: name == username-derived name
    username = getattr(user, "username", None)
    name_from_username = None
    if username:
        name_from_username = username.replace("_", " ")
        if name_from_username.lower() != (full_name or "").lower():
            q_filter |= Q(name__iexact=name_from_username)

    if q_filter:
        # Run a single query with all conditions OR'd
        candidates = list(Employee.objects.filter(q_filter)[:10])

        # Pick the best match in priority order
        for candidate in candidates:
            if worker_id and candidate.emp_id == worker_id:
                employee = candidate
                break
        if not employee and full_name:
            for candidate in candidates:
                if candidate.name and candidate.name.lower() == full_name.lower():
                    employee = candidate
                    break
        if not employee and name_from_username:
            for candidate in candidates:
                if candidate.name and candidate.name.lower() == name_from_username.lower():
                    employee = candidate
                    break
            # Partial match fallback (contains)
            if not employee and name_from_username:
                for candidate in candidates:
                    if candidate.name and name_from_username.lower() in candidate.name.lower():
                        employee = candidate
                        break

    if employee:
        setattr(user, cache_attr, employee)
        return employee

    # Auto-provision: create a minimal Employee record so the user can use task
    # features (comments, checklists, time logs).  Admins should review and
    # complete the record (department, supervisor, etc.) via the Employee admin page.
    if hasattr(user, "worker_id") and user.worker_id:
        try:
            create_name = full_name if full_name else user.username.replace("_", " ")

            employee = Employee.objects.create(
                emp_id=user.worker_id,
                name=create_name,
                is_enabled=True,
            )
            logger.warning(f"Auto-provisioned Employee for user {user.username} (worker_id={user.worker_id}). Admin should review this record.")
            setattr(user, cache_attr, employee)
            return employee
        except Exception as e:
            logger.error(f"Failed to auto-provision employee for user {user.username}: {e}")

    if raise_if_not_found:
        worker_id = getattr(user, "worker_id", "N/A")
        raise serializers.ValidationError({"error": f"Employee not found for current user (username: {user.username}, worker_id: {worker_id}). Please contact your administrator to create an employee record."})
    setattr(user, cache_attr, None)
    return None


class ProjectViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "projects"
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    pagination_class = ProjectPagination
    filter_backends = [SearchFilter]
    search_fields = ["name"]
    ordering_fields = ["id", "name", "created_at"]
    ordering = ["-id"]

    def get_queryset(self):
        return get_project_queryset()

    @swagger_auto_schema(
        operation_summary="List all projects",
        responses={200: ProjectSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(name="page", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Page number (default: 1)"),
            openapi.Parameter(name="page_size", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Items per page (default: 30, max: 100)"),
            openapi.Parameter(name="search", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Search by project name"),
            openapi.Parameter(name="ordering", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Order by field (use - for descending): -name, id, etc."),
        ],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        try:
            # Print request data for debugging
            logger.debug("Received project data: %s", request.data)

            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            logger.warning("Validation errors: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error creating project: {str(e)}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                self.perform_update(serializer)
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EmployeeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "employees"
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    pagination_class = EmployeePagination
    filter_backends = [SearchFilter]
    search_fields = ["name", "emp_id"]
    ordering_fields = ["id", "name", "emp_id", "created_at"]
    ordering = ["-id"]

    def get_queryset(self):
        return get_employee_queryset()

    @swagger_auto_schema(
        operation_summary="List all employees",
        responses={200: EmployeeSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(name="page", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Page number (default: 1)"),
            openapi.Parameter(name="page_size", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Items per page (default: 50, max: 200)"),
            openapi.Parameter(name="search", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Search by name or employee ID"),
            openapi.Parameter(name="ordering", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Order by field (use - for descending): -name, emp_id, etc."),
        ],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        try:
            # Print request data for debugging
            logger.debug("Received employee data: %s", request.data)

            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            logger.warning("Validation errors: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error creating employee: {str(e)}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                self.perform_update(serializer)
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @swagger_auto_schema(
        operation_summary="Bulk import employees from CSV",
        operation_description="Upload a CSV file to import multiple employees at once. Maximum 1000 rows.",
        manual_parameters=[
            openapi.Parameter(name="file", in_=openapi.IN_FORM, type=openapi.TYPE_FILE, required=True, description="CSV file containing employee data"),
            openapi.Parameter(name="update_existing", in_=openapi.IN_FORM, type=openapi.TYPE_BOOLEAN, required=False, description="Update existing records (default: false)"),
        ],
        consumes=["multipart/form-data"],
        responses={
            200: openapi.Response(
                description="Import successful",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "created": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "updated": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "errors": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                        "total_rows": openapi.Schema(type=openapi.TYPE_INTEGER),
                    },
                ),
            ),
            400: "Bad request - invalid file or format",
        },
    )
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def bulk_import(self, request):
        """Import employees from CSV file."""
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        update_existing = request.data.get("update_existing", "false").lower() == "true"

        try:
            results = BulkImportExportService.import_from_csv(file_obj=file_obj, serializer_class=EmployeeSerializer, max_rows=1000, update_existing=update_existing, lookup_field="emp_id")

            # Invalidate cache after bulk import
            CacheService.invalidate_cache(["employees"])

            logger.info(f"Bulk import completed: {results['created']} created, {results['updated']} updated, {len(results['errors'])} errors")

            return Response(results, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Bulk import failed: {str(e)}")
            return Response({"error": "Invalid request data. Please check your input and try again."}, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(
        operation_summary="Export employees to CSV",
        operation_description="Download all employees as a CSV file",
        manual_parameters=[openapi.Parameter(name="fields", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, description="Comma-separated list of fields to export (default: all)")],
        responses={200: "CSV file"},
    )
    @action(detail=False, methods=["get"])
    def export(self, request):
        """Export employees to CSV file."""
        # Get fields to export
        fields_param = request.query_params.get("fields", "")
        if fields_param:
            fields = [f.strip() for f in fields_param.split(",")]
        else:
            fields = ["id", "emp_id", "name", "department.name", "exclude_from_reports"]

        # Get queryset
        queryset = self.get_queryset()

        # Export to CSV
        csv_file = BulkImportExportService.export_to_csv(queryset=queryset, fields=fields)

        # Create response
        response = HttpResponse(csv_file.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="employees_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        logger.info(f"Exported {queryset.count()} employees to CSV")

        return response


class DepartmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "departments"
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [SearchFilter]
    search_fields = ["code", "name"]
    ordering_fields = ["id", "code", "name", "created_at"]
    ordering = ["code"]

    @swagger_auto_schema(
        operation_summary="List all departments",
        responses={200: DepartmentSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(name="search", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Search by department code or name"),
            openapi.Parameter(name="ordering", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Order by field (use - for descending): -code, name, etc."),
        ],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="Get employees in a department",
        responses={200: EmployeeSerializer(many=True)},
    )
    @action(detail=True, methods=["get"])
    def employees(self, request, pk=None):
        """Get all employees in this department"""
        department = self.get_object()
        employees = Employee.objects.filter(department=department).select_related("department").order_by("name")
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)

    @swagger_auto_schema(
        operation_summary="Remove employee from department",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "employee_id": openapi.Schema(type=openapi.TYPE_INTEGER, description="Employee ID to remove"),
            },
            required=["employee_id"],
        ),
        responses={200: openapi.Response("Employee removed from department")},
    )
    @action(detail=True, methods=["post"])
    def remove_employee(self, request, pk=None):
        """Remove an employee from this department (PTB admin only)"""
        # Check if user is PTB admin
        if not request.user.is_ptb_admin and not request.user.is_superuser:
            return Response({"error": "Only PTB admins can remove employees from departments"}, status=status.HTTP_403_FORBIDDEN)

        department = self.get_object()
        employee_id = request.data.get("employee_id")

        if not employee_id:
            return Response({"error": "employee_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id, department=department)
            employee.department = None
            employee.save()
            return Response({"message": f"Employee {employee.name} removed from department {department.code}"})
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found in this department"}, status=status.HTTP_404_NOT_FOUND)

    def create(self, request, *args, **kwargs):
        try:
            logger.debug("Received department data: %s", request.data)
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            logger.warning("Validation errors: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error creating department: {str(e)}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                self.perform_update(serializer)
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def generate_recurring_events(parent_event, assigned_to_ids=None):
    """
    Generate recurring event instances based on the parent event's repeat_frequency.
    - hourly: 24 hours from start
    - daily: 1 year from start
    - weekly: 1 year from start
    - monthly: 1 year from start
    - yearly: next 5 years
    """
    from datetime import timedelta

    from dateutil.relativedelta import relativedelta

    if not parent_event.is_repeating or not parent_event.repeat_frequency:
        return []

    frequency = parent_event.repeat_frequency
    start = parent_event.start
    end = parent_event.end
    duration = end - start  # Duration of each event

    # Calculate 1 year from start date for daily, weekly, monthly recurrence
    one_year_from_start = start + relativedelta(years=1)

    child_events = []
    current_start = start

    if frequency == "hourly":
        # Generate for 24 hours from start
        end_time = start + timedelta(hours=24)
        current_start = start + timedelta(hours=1)
        while current_start < end_time:
            child_events.append(
                {
                    "start": current_start,
                    "end": current_start + duration,
                }
            )
            current_start += timedelta(hours=1)

    elif frequency == "daily":
        # Generate for 1 year from start
        current_start = start + timedelta(days=1)
        while current_start <= one_year_from_start:
            child_events.append(
                {
                    "start": current_start,
                    "end": current_start + duration,
                }
            )
            current_start += timedelta(days=1)

    elif frequency == "weekly":
        # Generate for 1 year from start
        current_start = start + timedelta(weeks=1)
        while current_start <= one_year_from_start:
            child_events.append(
                {
                    "start": current_start,
                    "end": current_start + duration,
                }
            )
            current_start += timedelta(weeks=1)

    elif frequency == "monthly":
        # Generate for 1 year from start
        current_start = start + relativedelta(months=1)
        while current_start <= one_year_from_start:
            child_events.append(
                {
                    "start": current_start,
                    "end": current_start + duration,
                }
            )
            current_start += relativedelta(months=1)

    elif frequency == "yearly":
        # Generate for next 5 years
        for i in range(1, 6):
            current_start = start + relativedelta(years=i)
            child_events.append(
                {
                    "start": current_start,
                    "end": current_start + duration,
                }
            )

    # Create child events in database using bulk_create for performance
    events_to_create = []
    for event_data in child_events:
        events_to_create.append(
            CalendarEvent(
                title=parent_event.title,
                event_type=parent_event.event_type,
                status=parent_event.status,
                description=parent_event.description,
                start=event_data["start"],
                end=event_data["end"],
                all_day=parent_event.all_day,
                location=parent_event.location,
                color=parent_event.color,
                is_repeating=False,  # Child events are not repeating themselves
                repeat_frequency=None,
                parent_event=parent_event,
                created_by=parent_event.created_by,
                meeting_url=parent_event.meeting_url,
                project=parent_event.project,
                leave_type=parent_event.leave_type,
                applied_by=parent_event.applied_by,
                agent=parent_event.agent,
            )
        )

    created_events = CalendarEvent.objects.bulk_create(events_to_create, batch_size=100)

    # Set assigned_to for all created events in bulk
    if assigned_to_ids and created_events:
        through_model = CalendarEvent.assigned_to.through
        through_objects = []
        for event in created_events:
            for emp_id in assigned_to_ids:
                through_objects.append(through_model(calendarevent_id=event.id, employee_id=emp_id))
        if through_objects:
            through_model.objects.bulk_create(through_objects, batch_size=500)

    logger.info(f"Generated {len(created_events)} recurring events for parent event {parent_event.id}")
    return created_events


def update_recurring_child_events(parent_event, update_data, assigned_to_ids=None):
    """
    Update all child events when parent is updated.
    Propagates time changes relative to original.
    Uses bulk_update for efficiency instead of saving one-by-one.
    """
    child_events = list(parent_event.child_events.all())

    if not child_events:
        return

    # Fields to propagate to children
    propagate_fields = ["title", "event_type", "status", "description", "all_day", "location", "color", "meeting_url", "project", "leave_type", "applied_by", "agent"]

    # Calculate time shift if start/end changed
    time_shift = None
    if "start" in update_data or "end" in update_data:
        # Get the original times before update
        original_start = parent_event.start
        new_start = update_data.get("start", original_start)
        if isinstance(new_start, str):
            new_start = parse_datetime(new_start)
        time_shift = new_start - original_start

    # Collect fields that actually changed for bulk_update
    update_fields = []
    if time_shift:
        update_fields.extend(["start", "end"])
    for field in propagate_fields:
        if field in update_data:
            update_fields.append(field)

    for child in child_events:
        # Update time if shifted
        if time_shift:
            child.start = child.start + time_shift
            child.end = child.end + time_shift

        # Update propagate fields
        for field in propagate_fields:
            if field in update_data:
                setattr(child, field, update_data[field])

    # Bulk update all children at once instead of individual saves
    if update_fields:
        from .models import CalendarEvent

        CalendarEvent.objects.bulk_update(child_events, update_fields, batch_size=100)

    # Update assigned_to if provided (M2M requires per-object set)
    if assigned_to_ids is not None:
        for child in child_events:
            child.assigned_to.set(assigned_to_ids)

    logger.info(f"Updated {len(child_events)} child events for parent event {parent_event.id}")


def create_event_notification(event, notification_type="created"):
    """
    Create notifications for event participants.
    Excludes PTB admins from receiving task notifications - they only get leave notifications.
    For leave events, also notifies agents and PTB admins.
    For group tasks, also notifies group members who aren't directly assigned.
    """
    from .models import ExternalUser, Notification
    from .signals import notify_leave_event_participants, send_websocket_notification

    notifications = []

    # Handle leave events specially - notify agents and PTB admins
    if event.event_type == "leave":
        is_update = notification_type == "updated"
        leave_notifications = notify_leave_event_participants(event, is_update=is_update)
        if leave_notifications:
            notifications.extend(leave_notifications)
        return notifications

    # Get all assigned employees and find their ExternalUser accounts
    assigned_employees = list(event.assigned_to.all())

    # Batch lookup: fetch all matching ExternalUsers in one query instead of N+1
    emp_ids = [e.emp_id for e in assigned_employees if e.emp_id]
    user_map = {}
    if emp_ids:
        for eu in ExternalUser.objects.filter(worker_id__in=emp_ids):
            user_map[eu.worker_id.lower()] = eu

    # Track which employee IDs already got notified (to avoid duplicates with group members)
    notified_user_ids = set()

    # Check if this task belongs to a group
    group = getattr(event, "group", None)
    group_name = group.name if group else None

    notifs_to_create = []
    ws_payloads = []
    for employee in assigned_employees:
        if not employee.emp_id:
            continue

        external_user = user_map.get(employee.emp_id.lower())
        if not external_user:
            logger.debug(f"No ExternalUser found for employee {employee.emp_id}")
            continue

        # Skip PTB admins - they shouldn't receive task notifications
        if external_user.is_ptb_admin:
            logger.debug(f"Skipping notification for PTB admin {employee.emp_id}")
            continue

        notified_user_ids.add(external_user.id)

        if notification_type == "created":
            title = f"New {event.event_type.title()}: {event.title}"
            if group_name:
                message = f"You have been assigned to a new {event.event_type} '{event.title}' for group '{group_name}' on {event.start.strftime('%Y-%m-%d %H:%M')}."
            else:
                message = f"You have been assigned to a new {event.event_type} '{event.title}' on {event.start.strftime('%Y-%m-%d %H:%M')}."
        elif notification_type == "updated":
            title = f"{event.event_type.title()} Updated: {event.title}"
            message = f"The {event.event_type} '{event.title}' has been updated."
        elif notification_type == "deleted":
            title = f"{event.event_type.title()} Deleted: {event.title}"
            message = f"The {event.event_type} '{event.title}' has been cancelled."
        else:
            continue

        notifs_to_create.append(
            Notification(
                recipient=external_user,
                title=title,
                message=message,
                event=event if notification_type != "deleted" else None,
                event_type=event.event_type,
            )
        )
        ws_payloads.append(
            (
                external_user.id,
                {
                    "title": title,
                    "message": message,
                    "event_type": event.event_type,
                    "event_id": event.id if notification_type != "deleted" else None,
                    "is_read": False,
                },
            )
        )

    # Notify group members who are NOT already assigned (only for task creation)
    if group and notification_type == "created" and event.event_type == "task":
        group_members = list(group.members.all())
        group_emp_ids = [e.emp_id for e in group_members if e.emp_id]
        group_user_map = {}
        if group_emp_ids:
            for eu in ExternalUser.objects.filter(worker_id__in=group_emp_ids):
                group_user_map[eu.worker_id.lower()] = eu

        for member in group_members:
            if not member.emp_id:
                continue
            external_user = group_user_map.get(member.emp_id.lower())
            if not external_user:
                continue
            # Skip if already notified as an assignee
            if external_user.id in notified_user_ids:
                continue
            # Skip PTB admins
            if external_user.is_ptb_admin:
                continue
            notified_user_ids.add(external_user.id)

            title = f"New Group Task: {event.title}"
            message = f"A new task '{event.title}' has been created for your group '{group_name}' on {event.start.strftime('%Y-%m-%d %H:%M')}."

            notifs_to_create.append(
                Notification(
                    recipient=external_user,
                    title=title,
                    message=message,
                    event=event,
                    event_type="task",
                )
            )
            ws_payloads.append(
                (
                    external_user.id,
                    {
                        "title": title,
                        "message": message,
                        "event_type": "task",
                        "event_id": event.id,
                        "is_read": False,
                    },
                )
            )

    if notifs_to_create:
        try:
            created = Notification.objects.bulk_create(notifs_to_create)
            notifications.extend(created)
            for notif, (uid, payload) in zip(created, ws_payloads, strict=True):
                payload["id"] = notif.id
                payload["created_at"] = notif.created_at.isoformat()
                send_websocket_notification(uid, payload)
        except Exception as e:
            logger.error(f"Error bulk creating notifications for event {event.id}: {e}")

    logger.info(f"Created {len(notifications)} notifications for event {event.id}")
    return notifications


class CalendarEventViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "calendar"
    queryset = CalendarEvent.objects.all()
    serializer_class = CalendarEventSerializer
    pagination_class = CalendarEventPagination
    ordering_fields = ["id", "start", "end", "created_at"]
    ordering = ["-start"]

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return CalendarEvent.objects.none()
        if not self.request.user.is_authenticated:
            return CalendarEvent.objects.none()

        queryset = (
            CalendarEvent.objects.select_related(
                "created_by",
                "project",
                "group",
                "applied_by",
                "agent",
            )
            .prefetch_related(
                "assigned_to",
                "subtasks",
            )
            .annotate(
                _subtask_count=models.Count("subtasks"),
                _subtask_completed=models.Count("subtasks", filter=models.Q(subtasks__is_completed=True)),
            )
        )

        # Filter by event_type if provided (for separating calendar and task board data)
        event_type = self.request.query_params.get("event_type")
        if event_type:
            queryset = queryset.filter(event_type=event_type)

        # Filter to user's own events + holidays (for reminders)
        my_events = self.request.query_params.get("my_events")
        if my_events and my_events.lower() == "true":
            user = self.request.user
            worker_id = getattr(user, "worker_id", None)
            if worker_id:
                # Use subquery to avoid JOIN + DISTINCT overhead on M2M
                from django.db.models import Exists, OuterRef

                assigned_subquery = CalendarEvent.assigned_to.through.objects.filter(
                    calendarevent_id=OuterRef("pk"),
                    employee__emp_id=worker_id,
                )
                queryset = queryset.filter(models.Q(created_by__emp_id=worker_id) | Exists(assigned_subquery) | models.Q(event_type="holiday"))
            else:
                queryset = queryset.filter(event_type="holiday")

        return queryset

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.filter_queryset(self.get_queryset())

            # Use pagination instead of returning all events at once
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            import traceback

            logger.error(f"CalendarEvent list error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            # Explicitly support partial updates for PATCH
            partial = kwargs.pop("partial", request.method.upper() == "PATCH")
            instance = self.get_object()
            logger.info(f"CalendarEvent update - ID: {instance.id}, Data: {request.data}")

            # Check if this is a parent repeating event
            is_parent_repeating = instance.is_repeating and instance.parent_event is None

            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            if serializer.is_valid():
                self.perform_update(serializer)

                # If parent repeating event, update all children
                if is_parent_repeating:
                    assigned_to_ids = request.data.get("assigned_to", None)
                    update_recurring_child_events(instance, request.data, assigned_to_ids)

                # Create notification for update
                create_event_notification(instance, "updated")

                # Broadcast to task board if this is a task
                if instance.event_type == "task":
                    try:
                        from .consumers import broadcast_task_updated

                        broadcast_task_updated(serializer.data, request.user.username)
                    except Exception as ws_error:
                        logger.debug(f"WebSocket broadcast failed (non-critical): {ws_error}")

                logger.info(f"CalendarEvent {instance.id} updated successfully")
                return Response(serializer.data)
            logger.warning(f"CalendarEvent update validation failed: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback

            logger.error(f"CalendarEvent update error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.error(f"Request data: {request.data}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            task_id = instance.id
            event_type = instance.event_type
            logger.info(f"CalendarEvent delete - ID: {task_id}")

            # No notification for deleted events per user request

            # If this is a parent event, child events will be deleted by CASCADE
            self.perform_destroy(instance)

            # Broadcast to task board if this was a task
            if event_type == "task":
                try:
                    from .consumers import broadcast_task_deleted

                    broadcast_task_deleted(task_id, request.user.username)
                except Exception as ws_error:
                    logger.debug(f"WebSocket broadcast failed (non-critical): {ws_error}")

            logger.info(f"CalendarEvent {task_id} deleted successfully")
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            import traceback

            logger.error(f"CalendarEvent delete error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def perform_create(self, serializer):
        # Auto-set created_by to the current user's Employee record if not provided
        if not serializer.validated_data.get("created_by"):
            user = self.request.user
            worker_id = getattr(user, "worker_id", None)
            if worker_id:
                employee = Employee.objects.filter(emp_id=worker_id).first()
                if employee:
                    serializer.save(created_by=employee)
                    return
        return super().perform_create(serializer)

    def perform_update(self, serializer):
        return super().perform_update(serializer)

    def perform_destroy(self, instance):
        return super().perform_destroy(instance)

    def create(self, request, *args, **kwargs):
        try:
            logger.info(f"CalendarEvent create - Data: {request.data}")
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                # Use perform_create to auto-set created_by
                self.perform_create(serializer)
                event = serializer.instance

                # Re-fetch the event with related fields to include employee_name
                event.refresh_from_db()
                fresh_serializer = self.get_serializer(event)

                # Generate recurring events if this is a repeating event
                if event.is_repeating and event.repeat_frequency:
                    assigned_to_ids = request.data.get("assigned_to", [])
                    generate_recurring_events(event, assigned_to_ids)

                # Create notification for new event
                create_event_notification(event, "created")

                # Broadcast to task board if this is a task
                if event.event_type == "task":
                    try:
                        from .consumers import broadcast_task_created

                        broadcast_task_created(fresh_serializer.data)
                    except Exception as ws_error:
                        logger.debug(f"WebSocket broadcast failed (non-critical): {ws_error}")

                logger.info(f"CalendarEvent created successfully: {fresh_serializer.data.get('id')}")
                return Response(fresh_serializer.data, status=status.HTTP_201_CREATED)
            logger.warning(f"CalendarEvent create validation errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback

            logger.error(f"CalendarEvent create error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.error(f"Request data: {request.data}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HolidayViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing holidays on the calendar.
    Separate from CalendarEvent for simpler holiday management.
    """

    permission_classes = [IsAuthenticated]
    queryset = Holiday.objects.all()
    serializer_class = HolidaySerializer
    pagination_class = None  # No pagination for holidays
    ordering = ["date"]

    def get_queryset(self):
        queryset = Holiday.objects.select_related("created_by")

        # Filter by year if provided
        year = self.request.query_params.get("year")
        if year:
            try:
                year = int(year)
                queryset = queryset.filter(date__year=year)
            except ValueError:
                pass

        # Filter by month if provided
        month = self.request.query_params.get("month")
        if month:
            try:
                month = int(month)
                queryset = queryset.filter(date__month=month)
            except ValueError:
                pass

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        return queryset.order_by("date")

    def perform_create(self, serializer):
        """Set the created_by field to the current user"""
        user = self.request.user
        if isinstance(user, ExternalUser):
            instance = serializer.save(created_by=user)
        else:
            # Try to find matching ExternalUser
            ext_user = ExternalUser.objects.filter(username=user.username).first()
            instance = serializer.save(created_by=ext_user)
        # Broadcast calendar update
        try:
            from .consumers import broadcast_calendar_update

            broadcast_calendar_update("created", "holiday", HolidaySerializer(instance).data)
        except Exception as e:
            logger.debug(f"Calendar broadcast failed (non-critical): {e}")

    def perform_update(self, serializer):
        instance = serializer.save()
        try:
            from .consumers import broadcast_calendar_update

            broadcast_calendar_update("updated", "holiday", HolidaySerializer(instance).data)
        except Exception as e:
            logger.debug(f"Calendar broadcast failed (non-critical): {e}")

    def perform_destroy(self, instance):
        holiday_id = instance.id
        instance.delete()
        try:
            from .consumers import broadcast_calendar_update

            broadcast_calendar_update("deleted", "holiday", {"id": holiday_id})
        except Exception as e:
            logger.debug(f"Calendar broadcast failed (non-critical): {e}")


class EmployeeLeaveViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing employee leaves.
    Includes agent coverage tracking.
    """

    permission_classes = [IsAuthenticated]
    queryset = EmployeeLeave.objects.all()
    serializer_class = EmployeeLeaveSerializer
    pagination_class = None  # No pagination for leaves
    ordering = ["date", "employee__name"]

    def get_queryset(self):
        queryset = EmployeeLeave.objects.select_related("employee", "employee__department", "created_by").prefetch_related(models.Prefetch("agents", queryset=Employee.objects.select_related("department")))

        # Filter by year if provided
        year = self.request.query_params.get("year")
        if year:
            try:
                year = int(year)
                queryset = queryset.filter(date__year=year)
            except ValueError:
                pass

        # Filter by month if provided
        month = self.request.query_params.get("month")
        if month:
            try:
                month = int(month)
                queryset = queryset.filter(date__month=month)
            except ValueError:
                pass

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)

        # Filter by employee
        employee_id = self.request.query_params.get("employee")
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        # Filter by specific date
        date = self.request.query_params.get("date")
        if date:
            queryset = queryset.filter(date=date)

        return queryset.order_by("date", "employee__name")

    def perform_create(self, serializer):
        """Set the created_by field to the current user and notify PTB admins"""
        from django.db import transaction

        from .consumers import send_notification_to_user

        user = self.request.user
        if isinstance(user, ExternalUser):
            leave = serializer.save(created_by=user)
        else:
            # Try to find matching ExternalUser
            ext_user = ExternalUser.objects.filter(username=user.username).first()
            leave = serializer.save(created_by=ext_user)

        # Capture values needed for the on_commit callback
        leave_id = leave.id
        request_user = user

        def _send_leave_notifications():
            """Send notifications after transaction commits (M2M data is guaranteed flushed)."""
            from .consumers import broadcast_calendar_update

            leave_obj = EmployeeLeave.objects.get(pk=leave_id)
            employee_name = leave_obj.employee.name
            leave_date = leave_obj.date.strftime("%B %d, %Y")
            created_by_name = request_user.username if hasattr(request_user, "username") else "Unknown"

            # Collect agent names for richer notification messages
            agent_employees = list(leave_obj.agents.all())
            agent_names = [a.name for a in agent_employees] if agent_employees else []
            agent_names_str = ", ".join(agent_names) if agent_names else "None assigned"

            title = "New Leave Request"
            message = f"{employee_name} has taken leave on {leave_date}.\nAgent(s): {agent_names_str}\nSubmitted by {created_by_name}."

            # Bulk-create notifications for PTB admins
            ptb_admins = list(ExternalUser.objects.filter(is_ptb_admin=True, is_active=True))
            admin_notifs = [Notification(recipient=admin, title=title, message=message, event_type="leave") for admin in ptb_admins]
            if admin_notifs:
                created_admin = Notification.objects.bulk_create(admin_notifs)
                for notif, admin in zip(created_admin, ptb_admins, strict=True):
                    send_notification_to_user(admin.id, {"id": notif.id, "title": title, "message": message, "event_type": "leave", "event_id": None, "is_read": False, "created_at": notif.created_at.isoformat()})

            # Notify agents assigned to cover this leave (batch user lookup)
            if agent_employees:
                agent_emp_ids = [a.emp_id for a in agent_employees if a.emp_id]
                agent_users = {u.worker_id.lower(): u for u in ExternalUser.objects.filter(worker_id__in=agent_emp_ids, is_active=True)} if agent_emp_ids else {}

                agent_title = "Agent Assignment"
                agent_notifs = []
                agent_ws = []
                for agent_employee in agent_employees:
                    if not agent_employee.emp_id:
                        continue
                    agent_user = agent_users.get(agent_employee.emp_id.lower())
                    if agent_user:
                        agent_message = f"You have been assigned as agent for {employee_name} on {leave_date}."
                        agent_notifs.append(Notification(recipient=agent_user, title=agent_title, message=agent_message, event_type="leave"))
                        agent_ws.append((agent_user.id, agent_title, agent_message))

                if agent_notifs:
                    created_agents = Notification.objects.bulk_create(agent_notifs)
                    for notif, (uid, atitle, amsg) in zip(created_agents, agent_ws, strict=True):
                        send_notification_to_user(uid, {"id": notif.id, "title": atitle, "message": amsg, "event_type": "leave", "event_id": None, "is_read": False, "created_at": notif.created_at.isoformat()})

            # Broadcast calendar update for real-time PTB Calendar
            try:
                broadcast_calendar_update("created", "leave", EmployeeLeaveSerializer(leave_obj).data)
            except Exception as e:
                logger.debug(f"Calendar broadcast failed (non-critical): {e}")

        transaction.on_commit(_send_leave_notifications)

    def perform_update(self, serializer):
        instance = serializer.save()
        try:
            from .consumers import broadcast_calendar_update

            broadcast_calendar_update("updated", "leave", EmployeeLeaveSerializer(instance).data)
        except Exception as e:
            logger.debug(f"Calendar broadcast failed (non-critical): {e}")

    def perform_destroy(self, instance):
        leave_id = instance.id
        instance.delete()
        try:
            from .consumers import broadcast_calendar_update

            broadcast_calendar_update("deleted", "leave", {"id": leave_id})
        except Exception as e:
            logger.debug(f"Calendar broadcast failed (non-critical): {e}")


class OvertimeRequestViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = OvertimeRequest.objects.all()
    serializer_class = OvertimeSerializer
    pagination_class = OvertimeRequestPagination

    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = [
        "employee__name",
        "project__name",
        "request_date",
    ]
    ordering_fields = ["id", "request_date", "employee_name", "project_name", "total_hours", "status", "created_at"]
    ordering = ["-request_date", "-created_at"]

    @swagger_auto_schema(
        operation_summary="List overtime requests",
        operation_description="Get list of overtime requests. Admins see all, regular users see only their own.",
        manual_parameters=[
            openapi.Parameter(name="page", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Page number (default: 1)"),
            openapi.Parameter(name="page_size", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Items per page (default: 20, max: 50)"),
            openapi.Parameter(name="search", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Search by employee name, project, or date"),
            openapi.Parameter(
                name="employee",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="Filter by employee ID",
            ),
            openapi.Parameter(
                name="project",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_INTEGER,
                required=False,
                description="Filter by project ID",
            ),
            openapi.Parameter(
                name="request_date",
                in_=openapi.IN_QUERY,
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=False,
                description="Filter by specific date (YYYY-MM-DD)",
            ),
            openapi.Parameter(name="start_date", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, format=openapi.FORMAT_DATE, required=False, description="Filter by date range start (YYYY-MM-DD)"),
            openapi.Parameter(name="end_date", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, format=openapi.FORMAT_DATE, required=False, description="Filter by date range end (YYYY-MM-DD)"),
            openapi.Parameter(name="ordering", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Order by field (use - for descending): -request_date, employee, etc."),
        ],
        responses={
            200: OvertimeSerializer(many=True),
            400: openapi.Response(
                description="Bad Request",
                examples={"application/json": {"error": "Invalid parameters"}},
            ),
        },
    )
    def list(self, request, *args, **kwargs):
        # Return fresh data to avoid stale entries
        return super().list(request, *args, **kwargs)

    def _get_permission_queryset(self):
        """Return queryset with only permission checks applied (no query param filters).

        Used by stats endpoints that apply their own filters to avoid double-filtering.
        """
        queryset = get_overtime_queryset()

        # Skip filtering during schema generation
        if getattr(self, "swagger_fake_view", False):
            return queryset

        user = self.request.user
        if not user.is_authenticated:
            return queryset.none()

        # Apply Permission Checks
        is_god_mode = is_superadmin_user(user)

        if getattr(user, "is_ptb_admin", False) or is_god_mode:
            pass  # Admin sees all
        else:
            # Regular user sees only their own requests
            worker_id = getattr(user, "worker_id", None)
            employee_id = getattr(user, "employee_id", None)

            if worker_id:
                queryset = queryset.filter(employee__emp_id=worker_id)
            elif employee_id:
                queryset = queryset.filter(employee__id=employee_id)
            else:
                return queryset.none()

        return queryset

    def get_queryset(self):
        queryset = self._get_permission_queryset()

        # Apply Filters from Query Params
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        employee = self.request.query_params.get("employee")
        project = self.request.query_params.get("project")
        date = self.request.query_params.get("request_date")

        if start_date and end_date:
            queryset = queryset.filter(request_date__range=[start_date, end_date])
        if employee:
            try:
                queryset = queryset.filter(employee=int(employee))
            except (ValueError, TypeError):
                return queryset.none()
        if project:
            try:
                queryset = queryset.filter(project=int(project))
            except (ValueError, TypeError):
                return queryset.none()
        if date:
            queryset = queryset.filter(request_date=date)

        # Additional filters: status and department_code
        status_filter = self.request.query_params.get("status")
        department_code = self.request.query_params.get("department_code")

        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if department_code:
            queryset = queryset.filter(department_code=department_code)

        return queryset

    @swagger_auto_schema(
        operation_summary="Create overtime request",
        request_body=OvertimeSerializer,
        responses={201: OvertimeSerializer(), 400: "Validation Error"},
    )
    @cache_invalidate_on_change(["overtime_requests", "employees", "projects"])
    def create(self, request, *args, **kwargs):
        cache_key = f"ot_request_{request.data.get('employee')}_{request.data.get('project')}_{request.data.get('request_date')}"

        try:
            logger.info(f"OvertimeRequest create - Data: {request.data}")
            if not cache.add(cache_key, "locked", timeout=30):
                return Response({"error": "Request in progress. Please wait."}, status=409)

            with transaction.atomic():
                # Check existing inside transaction
                # Be defensive: if record exists but has corrupted data, skip it
                existing = None
                try:
                    existing = (
                        OvertimeRequest.objects.select_for_update()
                        .filter(
                            employee_id=request.data.get("employee"),
                            project_id=request.data.get("project"),
                            request_date=request.data.get("request_date"),
                        )
                        .first()
                    )

                    # Verify the existing record is valid by accessing its FK
                    if existing:
                        try:
                            _ = existing.employee.name  # Test FK access
                        except (ValueError, TypeError, Exception) as fk_error:
                            logger.warning(f"Found existing record {existing.id} but FK access failed: {fk_error}")
                            # Record exists but has corrupted data - skip update, create new instead
                            existing = None
                except Exception as query_error:
                    logger.warning(f"Error querying for existing record: {query_error}")
                    existing = None

                if existing:
                    logger.info(f"Existing OvertimeRequest found: {existing.id}, returning 409")
                    return Response(
                        {"error": "An overtime request already exists for this employee, project, and date.", "existing_id": existing.id},
                        status=status.HTTP_409_CONFLICT,
                    )

                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                logger.info(f"OvertimeRequest created successfully: {serializer.data.get('id')}")
                return Response(serializer.data, status=201)
        except Exception as e:
            import traceback

            logger.error(f"OvertimeRequest create error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.error(f"Request data: {request.data}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            cache.delete(cache_key)

    @cache_invalidate_on_change(["overtime_requests", "employees", "projects"])
    def destroy(self, request, *args, **kwargs):
        cache_key = f"ot_delete_{kwargs['pk']}"
        try:
            logger.info(f"OvertimeRequest delete - ID: {kwargs['pk']}")
            if not cache.add(cache_key, "locked", timeout=30):
                return Response({"error": "Delete in progress"}, status=409)

            with transaction.atomic():
                try:
                    instance = self.get_object()
                    # Lock the record within transaction
                    OvertimeRequest.objects.select_for_update().get(pk=instance.pk)
                    response = super().destroy(request, *args, **kwargs)
                    logger.info(f"OvertimeRequest {kwargs['pk']} deleted successfully")
                    return response
                except OvertimeRequest.DoesNotExist:
                    logger.warning(f"OvertimeRequest {kwargs['pk']} not found (already deleted)")
                    return Response({"error": "Request already deleted"}, status=404)
        except Exception as e:
            import traceback

            logger.error(f"OvertimeRequest delete error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            cache.delete(cache_key)

    @swagger_auto_schema(
        operation_summary="Update overtime request",
        request_body=OvertimeSerializer,
        responses={200: OvertimeSerializer(), 404: "Not Found"},
    )
    @cache_invalidate_on_change(["overtime_requests", "employees", "projects"])
    def update(self, request, *args, **kwargs):
        cache_key = f"ot_update_{kwargs['pk']}"
        try:
            logger.info(f"OvertimeRequest update - ID: {kwargs['pk']}, Data: {request.data}")
            if not cache.add(cache_key, "locked", timeout=30):
                return Response({"error": "Update in progress"}, status=409)

            with transaction.atomic():
                instance = self.get_object()
                # Lock the record
                OvertimeRequest.objects.select_for_update().get(pk=instance.pk)

                # Block editing approved or rejected requests (admins can still change status via bulk endpoint)
                if instance.status in ("approved", "rejected"):
                    return Response(
                        {"error": f"Cannot edit an overtime request that has been {instance.status}."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Prevent regular users from changing the status field directly
                user = request.user
                if "status" in request.data and not (getattr(user, "is_ptb_admin", False) or is_superadmin_user(user)):
                    return Response(
                        {"error": "Only admins can change request status. Use the approval workflow."},
                        status=status.HTTP_403_FORBIDDEN,
                    )

                response = super().update(request, *args, **kwargs)
                logger.info(f"OvertimeRequest {kwargs['pk']} updated successfully")
                return response
        except Exception as e:
            import traceback

            logger.error(f"OvertimeRequest update error: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.error(f"Request data: {request.data}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            cache.delete(cache_key)

    @swagger_auto_schema(
        operation_summary="Bulk update overtime request statuses",
        operation_description="Update status of multiple overtime requests at once. Admin only.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["ids", "status"],
            properties={
                "ids": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_INTEGER), description="List of overtime request IDs"),
                "status": openapi.Schema(type=openapi.TYPE_STRING, enum=["approved", "rejected", "pending"], description="New status to set"),
            },
        ),
        responses={200: openapi.Response(description="Bulk update successful", examples={"application/json": {"message": "10 requests updated successfully", "updated_count": 10}})},
    )
    @action(detail=False, methods=["post"], url_path="bulk-update-status")
    @cache_invalidate_on_change(["overtime_requests"])
    def bulk_update_status(self, request):
        """Bulk update status for multiple overtime requests - much faster than individual updates"""
        user = request.user
        if not (getattr(user, "is_ptb_admin", False) or is_superadmin_user(user)):
            return Response({"error": "Permission denied. Admin only."}, status=status.HTTP_403_FORBIDDEN)

        ids = request.data.get("ids", [])
        new_status = request.data.get("status")

        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate ids is a list of integers
        if not isinstance(ids, list) or not all(isinstance(i, int) for i in ids):
            return Response({"error": "IDs must be a list of integers"}, status=status.HTTP_400_BAD_REQUEST)

        if len(ids) > 500:
            return Response({"error": "Cannot update more than 500 requests at once"}, status=status.HTTP_400_BAD_REQUEST)

        if new_status not in ["approved", "rejected", "pending"]:
            return Response({"error": "Invalid status. Must be 'approved', 'rejected', or 'pending'"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Collect affected dates BEFORE updating (for Excel regeneration)
                affected_requests = list(OvertimeRequest.objects.filter(id__in=ids).values("id", "request_date", "employee__emp_id", "employee_name"))
                affected_dates = list({r["request_date"] for r in affected_requests})

                # Build update kwargs with status_changed_by and timestamps
                update_kwargs = {
                    "status": new_status,
                    "status_changed_by": getattr(user, "username", ""),
                }
                if new_status == "approved":
                    update_kwargs["approved_at"] = timezone.now()
                    update_kwargs["rejected_at"] = None
                elif new_status == "rejected":
                    update_kwargs["rejected_at"] = timezone.now()
                    update_kwargs["approved_at"] = None
                else:
                    # pending — clear both timestamps
                    update_kwargs["approved_at"] = None
                    update_kwargs["rejected_at"] = None

                # Lock and update all records in a single query
                updated_count = OvertimeRequest.objects.filter(id__in=ids).update(**update_kwargs)
                logger.info(f"Bulk status update: {updated_count} requests updated to '{new_status}' by user {user.username}")

            # After commit: send notifications to affected employees
            if updated_count > 0 and new_status in ("approved", "rejected"):
                try:
                    from .consumers import send_notification_to_user

                    status_label = "approved" if new_status == "approved" else "rejected"
                    admin_name = getattr(user, "username", "Admin")

                    # Collect unique emp_ids to look up ExternalUser records
                    emp_ids = list({r["employee__emp_id"] for r in affected_requests if r["employee__emp_id"]})
                    ext_users_map = {}
                    if emp_ids:
                        ext_users_map = {u.worker_id.lower(): u for u in ExternalUser.objects.filter(worker_id__in=emp_ids, is_active=True)}

                    notifs_to_create = []
                    ws_send_list = []
                    for req_info in affected_requests:
                        emp_id = (req_info["employee__emp_id"] or "").lower()
                        ext_user = ext_users_map.get(emp_id)
                        if not ext_user:
                            continue
                        date_str = req_info["request_date"].strftime("%B %d, %Y") if hasattr(req_info["request_date"], "strftime") else str(req_info["request_date"])
                        title = f"Overtime Request {status_label.title()}"
                        message = f"Your overtime request for {date_str} has been {status_label} by {admin_name}."
                        notifs_to_create.append(Notification(recipient=ext_user, title=title, message=message, event_type=f"overtime_{status_label}"))
                        ws_send_list.append((ext_user.id, title, message))

                    if notifs_to_create:
                        created_notifs = Notification.objects.bulk_create(notifs_to_create)
                        for notif, (uid, ntitle, nmsg) in zip(created_notifs, ws_send_list, strict=True):
                            send_notification_to_user(
                                uid,
                                {
                                    "id": notif.id,
                                    "title": ntitle,
                                    "message": nmsg,
                                    "event_type": notif.event_type,
                                    "event_id": None,
                                    "is_read": False,
                                    "created_at": notif.created_at.isoformat(),
                                },
                            )
                        logger.info(f"Sent {len(created_notifs)} notifications for {status_label} overtime requests")
                except Exception as notif_err:
                    logger.warning(f"Failed to send notifications after bulk status update: {notif_err}")

            # After commit: regenerate Excel files for affected dates
            # This ensures rejected requests are excluded from reports
            if updated_count > 0 and affected_dates:
                try:
                    from api.tasks import regenerate_excel_after_delete

                    for d in affected_dates:
                        d_str = d.isoformat() if hasattr(d, "isoformat") else str(d)
                        regenerate_excel_after_delete.delay(d_str)
                    logger.info(f"Queued Excel regeneration for {len(affected_dates)} affected dates after bulk status update")
                except Exception as regen_err:
                    logger.warning(f"Failed to queue Excel regeneration after bulk status update: {regen_err}")

            return Response({"message": f"{updated_count} requests updated successfully", "updated_count": updated_count, "status": new_status})
        except Exception as e:
            logger.error(f"Bulk update error: {str(e)}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=["post"])
    def export_files(self, request):
        # Only PTB admins and superadmins can export files
        user = request.user
        if not (getattr(user, "is_ptb_admin", False) or is_superadmin_user(user)):
            return Response({"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN)

        try:
            date = datetime.strptime(request.data["date"], "%Y-%m-%d").date()

            # Use grouped (by-department) variants for consistency with model save()
            daily_data = OvertimeRequest.export_daily_data_by_department(date)
            monthly_data = OvertimeRequest.export_monthly_data_by_department(date)

            # Generate Excel files
            ExcelGenerator.generate_all_excel_files(daily_data, monthly_data, date)

            return Response({"status": "success", "message": "Excel files exported successfully"})

        except Exception as e:
            logger.error(f"Export files error: {e}")
            return Response({"error": "An error occurred while exporting files."}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"])
    def employee_stats(self, request):
        """Get overtime statistics grouped by employee"""
        try:
            start_date = request.query_params.get("start_date")
            end_date = request.query_params.get("end_date")
            status_filter = request.query_params.get("status")
            employee_id = request.query_params.get("employee")

            # Cache key based on user + query params (stats change infrequently)
            user_id = getattr(request.user, "id", "anon")
            cache_key = f"ot_employee_stats:{user_id}:{start_date}:{end_date}:{status_filter}:{employee_id}"
            cached = cache.get(cache_key)
            if cached is not None:
                return Response(cached)

            queryset = self._get_permission_queryset()
            if start_date:
                queryset = queryset.filter(request_date__gte=start_date)
            if end_date:
                queryset = queryset.filter(request_date__lte=end_date)
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            else:
                # Exclude rejected requests by default for stats
                queryset = queryset.exclude(status="rejected")
            if employee_id:
                queryset = queryset.filter(employee=employee_id)

            from django.db.models import Case, Count, DecimalField, Q, Sum, Value, When

            _decimal = DecimalField(max_digits=10, decimal_places=2)
            # NOTE: The annotation name must NOT be 'total_hours' because it would
            # shadow the model field of the same name, causing FieldError when
            # Case/When expressions reference "total_hours" (Django resolves it to
            # the aggregate instead of the DB column).
            stats = (
                queryset.values("employee", "employee_name")
                .annotate(
                    sum_total_hours=Sum("total_hours"),
                    total_requests=Count("id"),
                    weekday_hours=Sum(Case(When(Q(is_weekend=False) & Q(is_holiday=False), then="total_hours"), default=Value(0), output_field=_decimal)),
                    weekend_hours=Sum(Case(When(is_weekend=True, then="total_hours"), default=Value(0), output_field=_decimal)),
                    holiday_hours=Sum(Case(When(is_holiday=True, then="total_hours"), default=Value(0), output_field=_decimal)),
                    approved_hours=Sum(Case(When(status="approved", then="total_hours"), default=Value(0), output_field=_decimal)),
                    pending_hours=Sum(Case(When(status="pending", then="total_hours"), default=Value(0), output_field=_decimal)),
                )
                .order_by("-sum_total_hours")
            )

            # Map 'sum_total_hours' back to 'total_hours' for frontend compatibility
            result = []
            for row in stats:
                row["total_hours"] = row.pop("sum_total_hours", 0)
                result.append(row)

            cache.set(cache_key, result, 120)  # Cache for 2 minutes
            return Response(result)
        except Exception as e:
            logger.error(f"employee_stats error: {e}", exc_info=True)
            return Response({"error": "Failed to compute employee stats. Please try again."}, status=500)

    @action(detail=False, methods=["get"])
    def project_stats(self, request):
        """Get overtime statistics grouped by project"""
        try:
            start_date = request.query_params.get("start_date")
            end_date = request.query_params.get("end_date")
            status_filter = request.query_params.get("status")
            project_id = request.query_params.get("project")

            # Cache key based on user + query params
            user_id = getattr(request.user, "id", "anon")
            cache_key = f"ot_project_stats:{user_id}:{start_date}:{end_date}:{status_filter}:{project_id}"
            cached = cache.get(cache_key)
            if cached is not None:
                return Response(cached)

            queryset = self._get_permission_queryset()
            if start_date:
                queryset = queryset.filter(request_date__gte=start_date)
            if end_date:
                queryset = queryset.filter(request_date__lte=end_date)
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            else:
                queryset = queryset.exclude(status="rejected")
            if project_id:
                queryset = queryset.filter(project=project_id)

            from django.db.models import Count, Sum

            stats = queryset.values("project", "project_name").annotate(total_hours=Sum("total_hours"), total_requests=Count("id"), unique_employees=Count("employee", distinct=True)).order_by("-total_hours")

            result = list(stats)
            cache.set(cache_key, result, 120)  # Cache for 2 minutes
            return Response(result)
        except Exception as e:
            logger.error(f"project_stats error: {e}", exc_info=True)
            return Response({"error": "Failed to compute project stats. Please try again."}, status=500)

    @action(detail=False, methods=["get"])
    def summary_stats(self, request):
        """Get overall overtime summary statistics.

        Accepts optional prev_start_date/prev_end_date to return previous-period
        stats alongside current stats (useful for trend calculations).
        """
        try:
            start_date = request.query_params.get("start_date")
            end_date = request.query_params.get("end_date")
            prev_start_date = request.query_params.get("prev_start_date")
            prev_end_date = request.query_params.get("prev_end_date")

            # Cache key based on user + query params
            user_id = getattr(request.user, "id", "anon")
            cache_key = f"ot_summary_stats:{user_id}:{start_date}:{end_date}:{prev_start_date}:{prev_end_date}"
            cached = cache.get(cache_key)
            if cached is not None:
                return Response(cached)

            base_queryset = self._get_permission_queryset().exclude(status="rejected")

            def _aggregate(qs):
                from django.db.models import Case, Count, DecimalField, Q, Sum, Value, When

                _decimal = DecimalField(max_digits=10, decimal_places=2)
                # Use 'sum_total_hours' to avoid shadowing the 'total_hours' model field.
                # Case/When expressions reference the DB column "total_hours" by name;
                # if an aggregate annotation has the same name, Django resolves to the
                # aggregate → FieldError: "'total_hours' is an aggregate".
                data = qs.aggregate(
                    sum_total_hours=Sum("total_hours"),
                    total_requests=Count("id"),
                    unique_employees=Count("employee", distinct=True),
                    unique_projects=Count("project", distinct=True),
                    weekday_hours=Sum(Case(When(Q(is_weekend=False) & Q(is_holiday=False), then="total_hours"), default=Value(0), output_field=_decimal)),
                    weekend_hours=Sum(Case(When(is_weekend=True, then="total_hours"), default=Value(0), output_field=_decimal)),
                    holiday_hours=Sum(Case(When(is_holiday=True, then="total_hours"), default=Value(0), output_field=_decimal)),
                    approved_hours=Sum(Case(When(status="approved", then="total_hours"), default=Value(0), output_field=_decimal)),
                    pending_hours=Sum(Case(When(status="pending", then="total_hours"), default=Value(0), output_field=_decimal)),
                )
                # Map back to 'total_hours' for frontend compatibility
                data["total_hours"] = data.pop("sum_total_hours", 0)
                return data

            current_qs = base_queryset
            if start_date:
                current_qs = current_qs.filter(request_date__gte=start_date)
            if end_date:
                current_qs = current_qs.filter(request_date__lte=end_date)

            result = _aggregate(current_qs)

            # Compute previous-period stats if dates provided
            if prev_start_date and prev_end_date:
                prev_qs = base_queryset.filter(
                    request_date__gte=prev_start_date,
                    request_date__lte=prev_end_date,
                )
                result["previous"] = _aggregate(prev_qs)

            cache.set(cache_key, result, 120)  # Cache for 2 minutes
            return Response(result)
        except Exception as e:
            logger.error(f"summary_stats error: {e}", exc_info=True)
            return Response({"error": "Failed to compute summary stats. Please try again."}, status=500)


class OvertimeRegulationViewSet(viewsets.ModelViewSet):
    """ViewSet for managing overtime regulations and rules"""

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "regulations"

    queryset = OvertimeRegulation.objects.all()
    serializer_class = OvertimeRegulationSerializer
    filter_backends = [SearchFilter]
    search_fields = ["title", "description", "category"]
    ordering_fields = ["id", "order", "category", "created_at"]
    ordering = ["order", "id"]

    def get_permissions(self):
        """Allow all authenticated users to read regulations (shown on OT Form).
        Only admin/authorized users can create/update/delete."""
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), ResourcePermission()]

    @swagger_auto_schema(
        operation_summary="List all overtime regulations",
        responses={200: OvertimeRegulationSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(name="category", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Filter by category"),
            openapi.Parameter(name="is_active", in_=openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description="Filter by active status"),
            openapi.Parameter(name="search", in_=openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description="Search by title, description, or category"),
            openapi.Parameter(name="page_size", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Items per page"),
        ],
    )
    @cached_list("overtime_regulations", ttl=3600)
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Filter by category
        category = request.query_params.get("category")
        if category:
            queryset = queryset.filter(category=category)

        # Filter by active status
        is_active = request.query_params.get("is_active")
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        # Check if pagination is requested
        page_size = request.query_params.get("page_size")
        if page_size and page_size.isdigit():
            # Use standard pagination
            return super().list(request, *args, **kwargs)

        # Return non-paginated response for compatibility
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data, "count": len(serializer.data)})

    @cache_invalidate_on_change(["overtime_regulations"])
    def perform_create(self, serializer):
        """Create and invalidate cache"""
        return super().perform_create(serializer)

    @cache_invalidate_on_change(["overtime_regulations"])
    def perform_update(self, serializer):
        """Update and invalidate cache"""
        return super().perform_update(serializer)

    @cache_invalidate_on_change(["overtime_regulations"])
    def perform_destroy(self, instance):
        """Delete and invalidate cache"""
        return super().perform_destroy(instance)

    def create(self, request, *args, **kwargs):
        try:
            logger.debug("Received regulation data: %s", request.data)
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            logger.warning("Validation errors: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error creating regulation: {str(e)}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        try:
            logger.debug(f"Updating regulation with data: {request.data}")
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                self.perform_update(serializer)
                return Response(serializer.data)
            logger.warning(f"Validation errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback

            logger.error(f"Error updating regulation: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class OvertimeRegulationDocumentViewSet(viewsets.ModelViewSet):
    """ViewSet for managing overtime regulation PDF documents"""

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "regulations"
    queryset = OvertimeRegulationDocument.objects.filter(is_deleted=False)
    serializer_class = OvertimeRegulationDocumentSerializer
    filter_backends = [SearchFilter]
    search_fields = ["title", "description"]
    ordering_fields = ["id", "version", "created_at"]
    ordering = ["-created_at"]

    def get_permissions(self):
        """Allow all authenticated users to read regulation documents (shown on OT Form).
        Only admin/authorized users can create/upload/delete."""
        if self.action in ("list", "retrieve"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), ResourcePermission()]

    @swagger_auto_schema(
        operation_summary="List all regulation documents",
        responses={200: OvertimeRegulationDocumentSerializer(many=True)},
        manual_parameters=[
            openapi.Parameter(name="is_active", in_=openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description="Filter by active status"),
            openapi.Parameter(name="page_size", in_=openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description="Items per page"),
        ],
    )
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Filter by active status
        is_active = request.query_params.get("is_active")
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == "true")

        # Check if pagination is requested
        page_size = request.query_params.get("page_size")
        if page_size and page_size.isdigit():
            # Use standard pagination
            return super().list(request, *args, **kwargs)

        # Return non-paginated response for compatibility
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data, "count": len(serializer.data)})

    def perform_create(self, serializer):
        """Set uploaded_by to current user."""
        serializer.save(uploaded_by=self.request.user)

    def perform_destroy(self, instance):
        """Soft delete instead of hard delete"""
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        # instance.deleted_by = self.request.user  # Set when auth is implemented
        instance.save()


class OvertimeLimitConfigViewSet(viewsets.ModelViewSet):
    """ViewSet for managing overtime limit configuration (singleton-like)."""

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "regulations"

    queryset = OvertimeLimitConfig.objects.all()
    serializer_class = OvertimeLimitConfigSerializer

    def get_permissions(self):
        """Allow all authenticated users to read limits (used on OT Form).
        Only admin/authorized users can update."""
        if self.action in ("list", "retrieve", "active"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), ResourcePermission()]

    def list(self, request, *args, **kwargs):
        """Return the active config (singleton pattern)."""
        config = OvertimeLimitConfig.get_active()
        serializer = self.get_serializer(config)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def active(self, request):
        """Return the active overtime limit configuration."""
        config = OvertimeLimitConfig.get_active()
        serializer = self.get_serializer(config)
        return Response(serializer.data)

    @action(detail=False, methods=["put", "patch"])
    def update_limits(self, request):
        """Update the active overtime limit configuration."""
        config = OvertimeLimitConfig.get_active()
        serializer = self.get_serializer(config, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ===========================
# Authentication Views
# ===========================


class LoginRateThrottle(AnonRateThrottle):
    """
    Strict rate limit for login endpoints — always active regardless of the
    global THROTTLING_ENABLED toggle.  Prevents brute-force credential stuffing.
    """

    rate = "10/min"
    scope = "login"


class LocalLoginView(APIView):
    """
    Local authentication endpoint for Django users
    Uses Django's built-in authentication and JWT tokens
    """

    permission_classes = [AllowAny]
    throttle_classes = [LoginRateThrottle]

    @swagger_auto_schema(
        operation_summary="Login with local credentials",
        operation_description="Authenticate local Django user and return JWT tokens",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["username", "password"],
            properties={
                "username": openapi.Schema(type=openapi.TYPE_STRING, description="Username"),
                "password": openapi.Schema(type=openapi.TYPE_STRING, format="password", description="Password"),
            },
        ),
        responses={
            200: openapi.Response(
                description="Login successful",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "access": openapi.Schema(type=openapi.TYPE_STRING, description="JWT access token"),
                        "refresh": openapi.Schema(type=openapi.TYPE_STRING, description="JWT refresh token"),
                        "user": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                                "username": openapi.Schema(type=openapi.TYPE_STRING),
                                "email": openapi.Schema(type=openapi.TYPE_STRING),
                                "first_name": openapi.Schema(type=openapi.TYPE_STRING),
                                "last_name": openapi.Schema(type=openapi.TYPE_STRING),
                                "is_staff": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                "is_superuser": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                            },
                        ),
                        "auth_type": openapi.Schema(type=openapi.TYPE_STRING, enum=["local"]),
                    },
                ),
            ),
            400: "Bad request - username or password missing",
            401: "Unauthorized - invalid credentials or inactive account",
        },
        tags=["auth"],
    )
    def post(self, request):
        """
        Authenticate local user and return JWT tokens

        Request Body:
            {
                "username": "user",
                "password": "password"
            }

        Response:
            {
                "access": "jwt_access_token",
                "refresh": "jwt_refresh_token",
                "user": {
                    "id": 1,
                    "username": "user",
                    "email": "user@example.com",
                    "first_name": "First",
                    "last_name": "Last",
                    "is_staff": true,
                    "is_superuser": false
                }
            }
        """
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return Response({"error": "Username and password are required"}, status=status.HTTP_400_BAD_REQUEST)

        # Authenticate user
        user = authenticate(request, username=username, password=password)

        if user is None:
            logger.warning(f"Failed local login attempt for username: {username}")
            return Response({"error": "Invalid username or password"}, status=status.HTTP_401_UNAUTHORIZED)

        if not user.is_active:
            logger.warning(f"Inactive user login attempt: {username}")
            return Response({"error": "User account is disabled"}, status=status.HTTP_401_UNAUTHORIZED)

        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)
        refresh_token = str(refresh)

        # Update last login
        user.last_login = timezone.now()
        user.save(update_fields=["last_login"])

        logger.info(f"Local user logged in successfully: {username}")

        # Note: Local users don't have ExternalUser records, so we skip activity logging for now
        # Activity logging is designed for ExternalUser model

        # Check if this is the superadmin user
        is_super_admin = is_superadmin_user(user)
        permission_updated = getattr(user, "permission_updated_at", None)

        response = Response(
            {
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email or "",
                    "first_name": user.first_name or "",
                    "last_name": user.last_name or "",
                    "is_staff": user.is_staff,
                    "is_superuser": user.is_superuser,
                    # Include worker_id, is_ptb_admin, menu_permissions for consistency
                    "worker_id": getattr(user, "worker_id", None),
                    "is_ptb_admin": getattr(user, "is_ptb_admin", False) or is_super_admin,
                    "role": getattr(user, "role", "developer" if is_super_admin else "user"),
                    "menu_permissions": getattr(user, "menu_permissions", {}),
                    "permission_updated_at": permission_updated.isoformat() if permission_updated else None,
                    "employee_id": None,
                    "department_id": None,
                },
            },
            status=status.HTTP_200_OK,
        )

        # Deliver tokens exclusively via httpOnly cookies
        set_auth_cookies(response, access_token, refresh_token)
        return response


class ExternalLoginView(APIView):
    """
    External authentication endpoint
    Authenticates with external API and stores session
    """

    permission_classes = [AllowAny]
    throttle_classes = [LoginRateThrottle]

    @swagger_auto_schema(
        operation_summary="Login with external credentials",
        operation_description="Authenticate with external API and return tokens",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["username", "password"],
            properties={
                "username": openapi.Schema(type=openapi.TYPE_STRING, description="Username"),
                "password": openapi.Schema(type=openapi.TYPE_STRING, format="password", description="Password"),
            },
        ),
        responses={
            200: openapi.Response(
                description="Login successful",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "access": openapi.Schema(type=openapi.TYPE_STRING, description="External JWT access token"),
                        "refresh": openapi.Schema(type=openapi.TYPE_STRING, description="External JWT refresh token"),
                        "user": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                                "username": openapi.Schema(type=openapi.TYPE_STRING),
                                "email": openapi.Schema(type=openapi.TYPE_STRING),
                                "first_name": openapi.Schema(type=openapi.TYPE_STRING),
                                "last_name": openapi.Schema(type=openapi.TYPE_STRING),
                                "is_staff": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                "is_superuser": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                "is_active": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                "is_ptb_admin": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                "worker_id": openapi.Schema(type=openapi.TYPE_STRING),
                                "employee_id": openapi.Schema(type=openapi.TYPE_INTEGER, nullable=True),
                                "department_id": openapi.Schema(type=openapi.TYPE_INTEGER, nullable=True),
                                "date_joined": openapi.Schema(type=openapi.TYPE_STRING, format="date-time"),
                                "last_login": openapi.Schema(type=openapi.TYPE_STRING, format="date-time", nullable=True),
                                "groups": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_STRING)),
                                "permissions": openapi.Schema(type=openapi.TYPE_OBJECT),
                            },
                        ),
                    },
                ),
            ),
            400: "Bad request - username or password missing",
            401: "Unauthorized - invalid credentials",
            500: "Internal server error - authentication service error",
        },
        tags=["auth"],
    )
    def post(self, request):
        """
        Authenticate with external API and return tokens

        Request Body:
            {
                "username": "user",
                "password": "password"
            }

        Response:
            {
                "access": "external_jwt_token",
                "refresh": "external_refresh_token",
                "user": {
                    "id": 1,
                    "external_id": 123,
                    "username": "user",
                    "email": "user@example.com",
                    "worker_id": "WORKER123",
                    "is_ptb_admin": true,
                    "first_name": "First",
                    "last_name": "Last"
                }
            }
        """
        username = request.data.get("username")
        password = request.data.get("password")

        if not username or not password:
            return Response({"error": "Username and password are required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Login to external API
            auth_data = ExternalAuthService.login(username, password)

            # Get user info from external API
            user_info = ExternalAuthService.get_user_info(auth_data["access"])

            # Get or create user
            user, _ = ExternalUser.objects.get_or_create(external_id=user_info["id"], defaults={"username": user_info["username"], "email": user_info.get("email", ""), "date_joined": user_info.get("date_joined", timezone.now())})

            # Check if user is deactivated in our system (not the external system)
            if not user.is_active:
                logger.warning(f"Deactivated external user login attempt: {username}")
                return Response({"error": "Your account has been deactivated. Please contact your administrator."}, status=status.HTTP_401_UNAUTHORIZED)

            # Update user data from external API
            user.update_from_external_api(user_info)
            # update_from_external_api() already calls self.save(); update last_login
            # in the same save to avoid a redundant second write.
            ExternalUser.objects.filter(pk=user.pk).update(
                last_login=timezone.now(),
                updated_at=timezone.now(),
            )

            # Deactivate old sessions for this user
            UserSession.objects.filter(user=user, is_active=True).update(is_active=False)

            # Create new session
            payload = ExternalAuthService.decode_token_payload(auth_data["access"])
            UserSession.objects.create(
                user=user,
                access_token=auth_data["access"],
                refresh_token=auth_data.get("refresh", ""),
                token_issued_at=datetime.fromtimestamp(payload.get("iat", 0), tz=timezone.get_current_timezone()),
                token_expires_at=datetime.fromtimestamp(payload.get("exp", 0), tz=timezone.get_current_timezone()),
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )

            logger.info(f"External user logged in successfully: {username}")

            # Log user login activity
            UserActivityLog.log_activity(user=user, action="login", details={"login_type": "external", "username": username}, request=request)

            # Helper function to safely format datetime fields
            def format_datetime(dt_value):
                """Safely convert datetime to ISO format string"""
                if dt_value is None:
                    return None
                if isinstance(dt_value, str):
                    return dt_value  # Already a string
                if hasattr(dt_value, "isoformat"):
                    return dt_value.isoformat()
                return str(dt_value)

            response = Response(
                {
                    "user": {
                        "id": user.id,
                        "username": user.username,
                        "email": user.email or "",
                        "first_name": user.first_name or "",
                        "last_name": user.last_name or "",
                        "is_staff": user.is_staff,
                        "is_superuser": user.is_superuser,
                        "is_active": user.is_active,
                        "is_ptb_admin": user.is_ptb_admin,
                        "role": getattr(user, "role", "user"),
                        "worker_id": user.worker_id or "",
                        "menu_permissions": user.menu_permissions,  # CRITICAL: Include menu permissions
                        "event_reminders_enabled": user.event_reminders_enabled,
                        "preferred_language": getattr(user, "preferred_language", "en"),
                        "permission_updated_at": format_datetime(user.permission_updated_at),  # For force logout detection
                        "employee_id": None,
                        "department_id": None,
                        "date_joined": format_datetime(user.date_joined),
                        "last_login": format_datetime(user.last_login),
                        "groups": user.groups_cache,
                        "permissions": user.permissions_cache,
                    },
                },
                status=status.HTTP_200_OK,
            )

            # Deliver tokens exclusively via httpOnly cookies
            set_auth_cookies(response, auth_data["access"], auth_data.get("refresh"))
            return response

        except AuthenticationFailed as e:
            logger.warning(f"External login failed for user {username}: {str(e)}")
            return Response({"error": "Authentication failed."}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            logger.error(f"External login error for user {username}: {str(e)}")
            return Response({"error": "Authentication service error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @staticmethod
    def _get_client_ip(request):
        """Get client IP address"""
        x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
        if x_forwarded_for:
            ip = x_forwarded_for.split(",")[0].strip()
        else:
            ip = request.META.get("REMOTE_ADDR")
        return ip


class TokenVerifyView(APIView):
    """
    Token verification endpoint
    Determines if a token is from local API or external API
    """

    permission_classes = [AllowAny]

    @swagger_auto_schema(
        operation_summary="Verify token source",
        operation_description="Verify if a token is from local API or external API, and check if it's valid",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["token"],
            properties={
                "token": openapi.Schema(type=openapi.TYPE_STRING, description="JWT token to verify"),
            },
        ),
        responses={
            200: openapi.Response(
                description="Token verification result",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "valid": openapi.Schema(type=openapi.TYPE_BOOLEAN, description="Whether the token is valid"),
                        "source": openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Token source: 'local', 'external', or 'unknown'",
                            enum=["local", "external", "unknown"],
                        ),
                        "details": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            description="Additional token details if valid",
                            properties={
                                "user_id": openapi.Schema(type=openapi.TYPE_INTEGER),
                                "username": openapi.Schema(type=openapi.TYPE_STRING),
                                "exp": openapi.Schema(type=openapi.TYPE_INTEGER, description="Token expiration timestamp"),
                                "iat": openapi.Schema(type=openapi.TYPE_INTEGER, description="Token issued at timestamp"),
                            },
                        ),
                    },
                ),
            ),
            400: "Bad request - token missing",
        },
        tags=["auth"],
    )
    def post(self, request):
        """
        Verify token and determine its source

        Request Body:
            {
                "token": "jwt_token_here"
            }

        Response:
            {
                "valid": true,
                "source": "external",
                "details": {
                    "user_id": 123,
                    "username": "user",
                    "exp": 1759831574,
                    "iat": 1759745114
                }
            }
        """
        token = request.data.get("token")

        if not token:
            return Response({"error": "Token is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Try to verify as local JWT token first
        try:
            from rest_framework_simplejwt.tokens import AccessToken

            # Validate local token
            validated_token = AccessToken(token)

            # Get user from token
            from django.contrib.auth import get_user_model

            User = get_user_model()

            try:
                user = User.objects.get(id=validated_token["user_id"])

                return Response(
                    {
                        "valid": True,
                        "source": "local",
                        "details": {
                            "user_id": user.id,
                            "username": user.username,
                            "exp": validated_token.get("exp"),
                            "iat": validated_token.get("iat"),
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            except User.DoesNotExist:
                logger.warning(f"Local token valid but user not found: {validated_token.get('user_id')}")
                return Response(
                    {"valid": False, "source": "local", "details": {"error": "User not found"}},
                    status=status.HTTP_200_OK,
                )

        except Exception as e:
            # Not a valid local token, try external
            logger.debug(f"Not a valid local token: {str(e)}")

        # Try to verify as external token
        try:
            # First check if token exists in our session database
            from api.models import UserSession

            try:
                session = UserSession.objects.select_related("user").get(access_token=token, is_active=True)

                # Check if token is expired
                if not session.is_token_expired():
                    payload = ExternalAuthService.decode_token_payload(token)

                    return Response(
                        {
                            "valid": True,
                            "source": "external",
                            "details": {
                                "user_id": session.user.external_id,
                                "username": session.user.username,
                                "exp": payload.get("exp"),
                                "iat": payload.get("iat"),
                            },
                        },
                        status=status.HTTP_200_OK,
                    )
                else:
                    # Token expired in our database
                    session.deactivate()
                    return Response(
                        {"valid": False, "source": "external", "details": {"error": "Token expired"}},
                        status=status.HTTP_200_OK,
                    )

            except UserSession.DoesNotExist:
                # Not in our database, verify with external API
                pass

            # Verify with external API
            is_valid = ExternalAuthService.verify_token(token)

            if is_valid:
                # Decode payload for details
                payload = ExternalAuthService.decode_token_payload(token)

                return Response(
                    {
                        "valid": True,
                        "source": "external",
                        "details": {
                            "user_id": payload.get("user_id"),
                            "exp": payload.get("exp"),
                            "iat": payload.get("iat"),
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {"valid": False, "source": "external", "details": {"error": "Token is invalid or expired"}},
                    status=status.HTTP_200_OK,
                )

        except AuthenticationFailed as e:
            logger.error(f"External token verification failed: {str(e)}")
            return Response(
                {"valid": False, "source": "unknown", "details": {"error": "Token verification failed."}},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            logger.error(f"Token verification error: {str(e)}", exc_info=True)
            return Response(
                {"valid": False, "source": "unknown", "details": {"error": "Verification failed"}},
                status=status.HTTP_200_OK,
            )


class TokenRefreshView(APIView):
    """
    Token refresh endpoint for both local and external tokens
    """

    permission_classes = [AllowAny]

    @swagger_auto_schema(
        operation_summary="Refresh access token",
        operation_description="Refresh access token using refresh token. Supports both local and external tokens.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["refresh"],
            properties={
                "refresh": openapi.Schema(type=openapi.TYPE_STRING, description="Refresh token"),
                "auth_type": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    enum=["local", "external"],
                    description="Token type (optional, auto-detected if not provided)",
                ),
            },
        ),
        responses={
            200: openapi.Response(
                description="Token refreshed successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "access": openapi.Schema(type=openapi.TYPE_STRING, description="New access token"),
                        "refresh": openapi.Schema(type=openapi.TYPE_STRING, description="New refresh token (local only)"),
                    },
                ),
            ),
            400: "Bad request - refresh token missing or invalid auth_type",
            401: "Unauthorized - invalid or expired refresh token",
        },
        tags=["auth"],
    )
    def post(self, request):
        """
        Refresh access token using refresh token

        Request Body:
            {
                "refresh": "refresh_token",
                "auth_type": "local" | "external"  # optional, auto-detected if not provided
            }

        Response:
            {
                "access": "new_access_token",
                "refresh": "new_refresh_token"  # only for local tokens with rotation
            }
        """
        refresh_token = request.data.get("refresh")
        auth_type = request.data.get("auth_type")  # 'local' or 'external'

        # Fallback: read refresh token from httpOnly cookie
        if not refresh_token:
            refresh_token = request.COOKIES.get("refresh_token")

        if not refresh_token:
            return Response({"error": "Refresh token is required", "detail": "No refresh token provided. Please login first.", "code": "no_refresh_token"}, status=status.HTTP_400_BAD_REQUEST)

        # If auth_type not specified, try to detect
        if not auth_type:
            # Try local first (SimplJWT tokens are typically longer)
            try:
                refresh = RefreshToken(refresh_token)
                auth_type = "local"
            except (InvalidToken, TokenError):
                # Not a local token, assume external
                auth_type = "external"

        # Handle local token refresh
        if auth_type == "local":
            try:
                refresh = RefreshToken(refresh_token)
                new_access_token = str(refresh.access_token)

                # If token rotation is enabled, get new refresh token
                new_refresh = str(refresh) if hasattr(refresh, "access_token") else None

                logger.info("Local token refreshed successfully")
                response = Response({"message": "Token refreshed"}, status=status.HTTP_200_OK)

                # Deliver tokens exclusively via httpOnly cookies
                set_auth_cookies(response, new_access_token, new_refresh)
                return response

            except (InvalidToken, TokenError) as e:
                logger.warning(f"Local token refresh failed: {str(e)}")
                return Response({"error": "Invalid or expired refresh token", "detail": "Your session has expired. Please login again.", "code": "invalid_refresh_token"}, status=status.HTTP_401_UNAUTHORIZED)

        # Handle external token refresh
        elif auth_type == "external":
            try:
                new_access_token = ExternalAuthService.refresh_token(refresh_token)

                # Update session if exists
                try:
                    session = UserSession.objects.get(refresh_token=refresh_token, is_active=True)

                    payload = ExternalAuthService.decode_token_payload(new_access_token)
                    session.access_token = new_access_token
                    session.token_expires_at = datetime.fromtimestamp(payload.get("exp", 0), tz=timezone.get_current_timezone())
                    session.save()

                    logger.info(f"External token refreshed for user: {session.user.username}")

                except UserSession.DoesNotExist:
                    logger.warning("Session not found for external token refresh")

                response = Response({"message": "Token refreshed"}, status=status.HTTP_200_OK)

                # Deliver token exclusively via httpOnly cookie
                set_auth_cookies(response, new_access_token)
                return response

            except AuthenticationFailed as e:
                logger.warning(f"External token refresh failed: {str(e)}")
                return Response({"error": "Authentication failed.", "detail": "Your session has expired. Please login again.", "code": "invalid_refresh_token"}, status=status.HTTP_401_UNAUTHORIZED)

        else:
            return Response({"error": 'Invalid auth_type. Must be "local" or "external"', "code": "invalid_auth_type"}, status=status.HTTP_400_BAD_REQUEST)


class LogoutView(APIView):
    """
    Logout endpoint - deactivates sessions
    """

    @swagger_auto_schema(
        operation_summary="Logout user",
        operation_description="Logout current user and deactivate sessions. For external users, deactivates all active sessions. For local users, client should discard tokens.",
        responses={
            200: openapi.Response(
                description="Logout successful",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "message": openapi.Schema(type=openapi.TYPE_STRING, description="Success message"),
                    },
                ),
            ),
        },
        tags=["auth"],
    )
    def post(self, request):
        """
        Logout current user
        Deactivates session for external users
        For local users, client should discard tokens
        """
        user = request.user

        if isinstance(user, ExternalUser):
            # Log logout activity before deactivating sessions
            UserActivityLog.log_activity(user=user, action="logout", details={"logout_type": "manual"}, request=request)
            # Deactivate all active sessions for external user
            UserSession.objects.filter(user=user, is_active=True).update(is_active=False)
            logger.info(f"External user logged out: {user.username}")
        else:
            # For local users, just log the logout
            logger.info(f"Local user logged out: {user.username}")

        response = Response({"message": "Successfully logged out"}, status=status.HTTP_200_OK)

        # Clear httpOnly auth cookies
        clear_auth_cookies(response)
        return response


class ExchangeExternalTokenView(APIView):
    """
    Exchange an external access token for httpOnly session cookies.
    Used when an external app redirects the user with a non-httpOnly
    access_token cookie; the frontend reads it and POSTs here to set
    secure httpOnly cookies for all subsequent API calls.
    """

    permission_classes = [AllowAny]

    @swagger_auto_schema(
        operation_summary="Exchange external token for httpOnly cookies",
        operation_description="Validates an external access token and sets httpOnly session cookies.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["token"],
            properties={
                "token": openapi.Schema(type=openapi.TYPE_STRING, description="External access token"),
            },
        ),
        responses={
            200: "Cookie set + user data returned",
            401: "Invalid token",
        },
        tags=["auth"],
    )
    def post(self, request):
        token = request.data.get("token")
        if not token:
            return Response({"error": "Token is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user_info = ExternalAuthService.get_user_info(token)
        except Exception:
            return Response({"error": "Invalid token"}, status=status.HTTP_401_UNAUTHORIZED)

        # Get or create user
        user, _ = ExternalUser.objects.get_or_create(
            external_id=user_info["id"],
            defaults={
                "username": user_info["username"],
                "email": user_info.get("email", ""),
                "date_joined": user_info.get("date_joined", timezone.now()),
            },
        )
        user.update_from_external_api(user_info)

        # Ensure an active session exists
        try:
            UserSession.objects.get(access_token=token, is_active=True)
        except UserSession.DoesNotExist:
            payload = ExternalAuthService.decode_token_payload(token)
            UserSession.objects.create(
                user=user,
                access_token=token,
                refresh_token="",
                token_issued_at=datetime.fromtimestamp(payload.get("iat", 0), tz=timezone.get_current_timezone()),
                token_expires_at=datetime.fromtimestamp(payload.get("exp", 0), tz=timezone.get_current_timezone()),
                ip_address=request.META.get("REMOTE_ADDR", ""),
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )

        def format_datetime(dt_value):
            if dt_value is None:
                return None
            if isinstance(dt_value, str):
                return dt_value
            if hasattr(dt_value, "isoformat"):
                return dt_value.isoformat()
            return str(dt_value)

        response = Response(
            {
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email or "",
                    "first_name": user.first_name or "",
                    "last_name": user.last_name or "",
                    "is_staff": user.is_staff,
                    "is_superuser": user.is_superuser,
                    "is_active": user.is_active,
                    "is_ptb_admin": user.is_ptb_admin,
                    "role": getattr(user, "role", "user"),
                    "worker_id": user.worker_id or "",
                    "menu_permissions": user.menu_permissions,
                    "event_reminders_enabled": user.event_reminders_enabled,
                    "preferred_language": getattr(user, "preferred_language", "en"),
                    "permission_updated_at": format_datetime(user.permission_updated_at),
                    "employee_id": None,
                    "department_id": None,
                    "date_joined": format_datetime(user.date_joined),
                    "last_login": format_datetime(user.last_login),
                    "groups": user.groups_cache,
                    "permissions": user.permissions_cache,
                },
            },
            status=status.HTTP_200_OK,
        )

        set_auth_cookies(response, token)
        return response


class CurrentUserView(APIView):
    """
    Get current authenticated user information
    """

    @swagger_auto_schema(
        operation_summary="Get current user",
        operation_description="Get current authenticated user information. Works for both local and external users.",
        responses={
            200: openapi.Response(
                description="User information retrieved successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                        "username": openapi.Schema(type=openapi.TYPE_STRING),
                        "email": openapi.Schema(type=openapi.TYPE_STRING),
                        "first_name": openapi.Schema(type=openapi.TYPE_STRING),
                        "last_name": openapi.Schema(type=openapi.TYPE_STRING),
                        "is_staff": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        "is_superuser": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        "employee_id": openapi.Schema(type=openapi.TYPE_INTEGER, nullable=True),
                        "department_id": openapi.Schema(type=openapi.TYPE_INTEGER, nullable=True),
                    },
                ),
            ),
            401: "Unauthorized - authentication required",
            500: "Internal server error",
        },
        tags=["auth"],
    )
    def get(self, request):
        """
        Get current user details
        Works for both local and external users
        """
        try:
            user = request.user

            # Check if user is authenticated
            if not user or not user.is_authenticated:
                logger.warning(f"Unauthenticated user trying to access /auth/me/: {user}")
                return Response({"error": "auth required"}, status=status.HTTP_401_UNAUTHORIZED)

            logger.info(f"CurrentUserView called for user: {user.username}, type: {type(user).__name__}")

            # Check if external user
            if isinstance(user, ExternalUser):
                logger.info(f"Processing external user: {user.username} (ID: {user.id})")
                try:
                    # Build response safely with ALL required fields
                    permission_updated = getattr(user, "permission_updated_at", None)
                    user_data = {
                        "id": getattr(user, "id", None),
                        "username": getattr(user, "username", ""),
                        "email": getattr(user, "email", "") or "",
                        "first_name": getattr(user, "first_name", "") or "",
                        "last_name": getattr(user, "last_name", "") or "",
                        "is_staff": getattr(user, "is_staff", False),
                        "is_superuser": getattr(user, "is_superuser", False),
                        "is_active": getattr(user, "is_active", True),
                        # CRITICAL: Include external user specific fields
                        "worker_id": getattr(user, "worker_id", None),
                        "is_ptb_admin": getattr(user, "is_ptb_admin", False),
                        "role": getattr(user, "role", "user"),
                        "menu_permissions": getattr(user, "menu_permissions", []),
                        "event_reminders_enabled": getattr(user, "event_reminders_enabled", True),
                        "preferred_language": getattr(user, "preferred_language", "en"),
                        "permission_updated_at": permission_updated.isoformat() if permission_updated else None,
                        "employee_id": None,
                        "department_id": None,
                    }
                    logger.info(f"Successfully built external user data for: {user.username}, worker_id={user_data['worker_id']}, is_ptb_admin={user_data['is_ptb_admin']}, menu_permissions_type={type(user_data['menu_permissions']).__name__}")
                    return Response(user_data, status=status.HTTP_200_OK)
                except Exception as e:
                    logger.error(f"Error building external user response for {user.username}: {str(e)}", exc_info=True)
                    return Response({"error": "Failed to build user data", "detail": "Check server logs for details."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                # Local Django user
                logger.info(f"Processing local user: {user.username} (ID: {user.id})")
                try:
                    # For local users, check if matches superadmin criteria
                    # This ensures superadmin-role users can access SuperAdmin pages even as a local user
                    is_super_admin = is_superadmin_user(user)
                    permission_updated = getattr(user, "permission_updated_at", None)
                    user_data = {
                        "id": getattr(user, "id", None),
                        "username": getattr(user, "username", ""),
                        "email": getattr(user, "email", "") or "",
                        "first_name": getattr(user, "first_name", "") or "",
                        "last_name": getattr(user, "last_name", "") or "",
                        "is_active": getattr(user, "is_active", True),
                        "is_staff": getattr(user, "is_staff", False),
                        "is_superuser": getattr(user, "is_superuser", False),
                        # Include worker_id, is_ptb_admin, menu_permissions for consistency with external users
                        "worker_id": getattr(user, "worker_id", None),
                        "is_ptb_admin": getattr(user, "is_ptb_admin", False) or is_super_admin,
                        "role": getattr(user, "role", "developer" if is_super_admin else "user"),
                        "menu_permissions": getattr(user, "menu_permissions", {}),
                        "event_reminders_enabled": getattr(user, "event_reminders_enabled", True),
                        "preferred_language": getattr(user, "preferred_language", "en"),
                        "permission_updated_at": permission_updated.isoformat() if permission_updated else None,
                        "employee_id": None,
                        "department_id": None,
                    }
                    logger.info(f"Successfully built local user data for: {user.username}, is_super_admin={is_super_admin}")
                    return Response(user_data, status=status.HTTP_200_OK)
                except Exception as e:
                    logger.error(f"Error building local user response for {user.username}: {str(e)}", exc_info=True)
                    return Response({"error": "Failed to build user data", "detail": "Check server logs for details."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            logger.error(f"Unexpected error in CurrentUserView.get(): {str(e)}", exc_info=True)
            return Response({"error": "Internal server error", "detail": "Check server logs for details."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def patch(self, request):
        """Update current user preferences (e.g. event_reminders_enabled)"""
        try:
            user = request.user
            if not user or not user.is_authenticated:
                return Response({"error": "auth required"}, status=status.HTTP_401_UNAUTHORIZED)

            if not isinstance(user, ExternalUser):
                return Response({"error": "Only external users can update preferences"}, status=status.HTTP_400_BAD_REQUEST)

            allowed_fields = {"event_reminders_enabled", "preferred_language"}
            updated_fields = []
            for field in allowed_fields:
                if field in request.data:
                    setattr(user, field, request.data[field])
                    updated_fields.append(field)

            if updated_fields:
                user.save(update_fields=updated_fields + ["updated_at"])
                logger.info(f"User {user.username} updated preferences: {updated_fields}")

            return Response({"status": "ok", "updated": updated_fields}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error updating user preferences: {str(e)}", exc_info=True)
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UserAccessViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing user access control (PTB Admin, Superuser, Staff)
    Only accessible by users with the 'developer' or 'superadmin' role.
    """

    queryset = ExternalUser.objects.all()
    serializer_class = UserAccessSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    @swagger_auto_schema(
        operation_summary="List all users with access control",
        responses={200: UserAccessSerializer(many=True)},
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="Get user access details",
        responses={200: UserAccessSerializer()},
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_summary="Update user access permissions",
        request_body=UserAccessUpdateSerializer,
        responses={200: UserAccessSerializer()},
    )
    def partial_update(self, request, *args, **kwargs):
        from django.utils import timezone

        instance = self.get_object()

        # Prevent modifying developer's permissions (highest privilege, protected)
        if is_developer_user(instance):
            return Response({"error": "Cannot modify developer permissions"}, status=status.HTTP_403_FORBIDDEN)

        # Non-developer superadmins cannot change role field (only developer can)
        if "role" in request.data and not is_developer_user(request.user):
            return Response({"error": "Only developer can change user roles"}, status=status.HTTP_403_FORBIDDEN)

        # Track if we need to force logout the user
        should_force_logout = False

        # Handle role change (only developer can do this)
        if "role" in request.data:
            new_role = request.data["role"]
            if new_role not in ("developer", "superadmin", "user"):
                return Response({"error": "Invalid role"}, status=status.HTTP_400_BAD_REQUEST)
            # Cannot grant developer role to others
            if new_role == "developer":
                return Response({"error": "Cannot grant developer role"}, status=status.HTTP_403_FORBIDDEN)
            if instance.role != new_role:
                should_force_logout = True
            instance.role = new_role

        # Update only the allowed fields
        if "is_ptb_admin" in request.data:
            if instance.is_ptb_admin != request.data["is_ptb_admin"]:
                should_force_logout = True
            instance.is_ptb_admin = request.data["is_ptb_admin"]
        if "is_superuser" in request.data:
            if instance.is_superuser != request.data["is_superuser"]:
                should_force_logout = True
            instance.is_superuser = request.data["is_superuser"]
        if "is_staff" in request.data:
            if instance.is_staff != request.data["is_staff"]:
                should_force_logout = True
            instance.is_staff = request.data["is_staff"]
        if "is_active" in request.data:
            if instance.is_active != request.data["is_active"]:
                should_force_logout = True
            instance.is_active = request.data["is_active"]
        if "menu_permissions" in request.data:
            if instance.menu_permissions != request.data["menu_permissions"]:
                should_force_logout = True
            instance.menu_permissions = request.data["menu_permissions"]
        if "event_reminders_enabled" in request.data:
            instance.event_reminders_enabled = request.data["event_reminders_enabled"]
            # Preference-only change: still push WebSocket so UI updates in real-time
            should_force_logout = True

        # If any permission-related field changed, update the timestamp and notify via WebSocket
        if should_force_logout:
            instance.permission_updated_at = timezone.now()
            logger.info(f"User {instance.username} permissions updated at {instance.permission_updated_at}")

        instance.save()

        serializer = self.get_serializer(instance)

        # Send real-time permission update via WebSocket
        if should_force_logout:
            try:
                from .consumers import send_permission_update_to_user

                send_permission_update_to_user(instance.id, serializer.data)
                logger.info(f"WebSocket permission update sent to user {instance.username} (ID: {instance.id})")
            except Exception as e:
                logger.warning(f"Failed to send WebSocket permission update: {e}")

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        """Prevent deletion of user records"""
        return Response({"error": "User records cannot be deleted"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


class NotificationPagination(PageNumberPagination):
    """
    Pagination for notifications.
    Supports 'limit' query param for different page sizes.
    """

    page_size = 20
    page_size_query_param = "limit"
    max_page_size = 100

    def get_paginated_response(self, data):
        return Response({"count": self.page.paginator.count, "next": self.get_next_link(), "previous": self.get_previous_link(), "total_pages": self.page.paginator.num_pages, "current_page": self.page.number, "results": data})


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for user notifications.
    Supports list, retrieve, mark as read, archive, and delete.

    Query Parameters:
    - limit: Number of notifications per page (default: 20, max: 100)
    - page: Page number for pagination
    - no_pagination: If 'true', returns all notifications without pagination (for dropdown menu)
    - include_archived: If 'true', includes archived notifications
    - archived_only: If 'true', returns only archived notifications
    """

    http_method_names = ["get", "post", "delete", "head", "options"]

    permission_classes = [IsAuthenticated]
    serializer_class = NotificationSerializer
    pagination_class = NotificationPagination

    def get_queryset(self):
        # Fix for swagger schema generation and unauthenticated access
        if getattr(self, "swagger_fake_view", False) or not self.request.user.is_authenticated:
            return Notification.objects.none()

        # Filter notifications for the current user
        queryset = Notification.objects.filter(recipient=self.request.user)

        # Detail actions (archive, unarchive, destroy, retrieve) must see ALL
        # notifications regardless of archive status, otherwise archived items
        # return 404 when the user tries to unarchive or delete them.
        if self.action in ("archive", "unarchive", "destroy", "retrieve", "mark_read"):
            return queryset.select_related("recipient", "event").only(
                "id", "title", "message", "is_read", "is_archived", "event_type", "created_at", "recipient__id", "recipient__username", "event__id", "event__title", "event__meeting_url", "event__event_type"
            )

        # Exclude archived by default unless explicitly requested
        include_archived = self.request.query_params.get("include_archived", "").lower() == "true"
        archived_only = self.request.query_params.get("archived_only", "").lower() == "true"
        if archived_only:
            queryset = queryset.filter(is_archived=True)
        elif not include_archived:
            queryset = queryset.filter(is_archived=False)

        # Use select_related for efficient queries
        return queryset.select_related("recipient", "event").only(
            "id",
            "title",
            "message",
            "is_read",
            "is_archived",
            "event_type",
            "created_at",
            "recipient__id",
            "recipient__username",
            "event__id",
            "event__title",
            "event__meeting_url",
            "event__event_type",
        )

    def list(self, request, *args, **kwargs):
        """
        List notifications with optional pagination.
        Use ?no_pagination=true to get all notifications without pagination.
        Use ?limit=10 to get specific number of notifications.
        """
        queryset = self.filter_queryset(self.get_queryset())

        # Check if no_pagination is requested (for notification dropdown)
        no_pagination = request.query_params.get("no_pagination", "").lower() == "true"
        limit = request.query_params.get("limit")

        if no_pagination:
            # For dropdown menu - return limited results without pagination structure
            if limit:
                try:
                    limit_int = int(limit)
                    queryset = queryset[:limit_int]
                except ValueError:
                    pass
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        # Standard paginated response
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="unread-count")
    def unread_count(self, request):
        """Get unread notification count"""
        count = Notification.get_unread_count(request.user)
        return Response({"unread_count": count})

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        """Mark a single notification as read"""
        notification = self.get_object()
        if not notification.is_read:
            notification.is_read = True
            notification.save(update_fields=["is_read"])
        return Response({"status": "marked as read"})

    @action(detail=False, methods=["post"], url_path="mark-all-read")
    def mark_all_read(self, request):
        """Mark all notifications for the user as read"""
        updated = self.get_queryset().filter(is_read=False).update(is_read=True)
        return Response({"status": "all marked as read", "count": updated})

    @action(detail=False, methods=["post"], url_path="archive-old")
    def archive_old(self, request):
        """Archive old notifications (admin only)"""
        if not is_superadmin_user(request.user):
            return Response({"detail": "Only super admin can archive notifications."}, status=status.HTTP_403_FORBIDDEN)
        days = int(request.data.get("days", 90))
        archived_count = Notification.archive_old_notifications(days=days)
        return Response({"status": "archived", "count": archived_count})

    @action(detail=True, methods=["post"], url_path="archive")
    def archive(self, request, pk=None):
        """Archive a single notification (idempotent)"""
        notification = self.get_object()
        if not notification.is_archived:
            notification.is_archived = True
            notification.save(update_fields=["is_archived"])
        return Response({"status": "archived"})

    @action(detail=True, methods=["post"], url_path="unarchive")
    def unarchive(self, request, pk=None):
        """Unarchive a single notification (idempotent)"""
        notification = self.get_object()
        if notification.is_archived:
            notification.is_archived = False
            notification.save(update_fields=["is_archived"])
        return Response({"status": "unarchived"})

    def destroy(self, request, *args, **kwargs):
        """Delete a single notification"""
        notification = self.get_object()
        notification.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SystemConfigurationView(APIView):
    """
    Get or Update System Configuration.
    Singleton logic.
    GET is public (app name/version is non-sensitive, needed on login page).
    PATCH requires Super Admin.
    POST tab-icon/ for favicon upload.
    """

    permission_classes = [AllowAny]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get(self, request):
        config, created = SystemConfiguration.objects.get_or_create(pk=1)  # Ensure existing
        serializer = SystemConfigurationSerializer(config, context={"request": request})
        return Response(serializer.data)

    def patch(self, request):
        # Strict Super Admin Check
        user = request.user
        is_super_admin = is_superadmin_user(user)

        if not is_super_admin:
            return Response({"detail": "Only Super Admin can edit configuration."}, status=status.HTTP_403_FORBIDDEN)

        config, created = SystemConfiguration.objects.get_or_create(pk=1)

        # Handle tab_icon file upload from multipart form data
        data = request.data.copy() if hasattr(request.data, "copy") else dict(request.data)
        if "tab_icon" in request.FILES:
            data["tab_icon"] = request.FILES["tab_icon"]

        serializer = SystemConfigurationSerializer(config, data=data, partial=True, context={"request": request})
        if serializer.is_valid():
            serializer.save()
            # Log activity
            UserActivityLog.log_activity(user=request.user, action="update", resource="system_configuration", resource_id=1, details={"changes": {k: str(v) for k, v in request.data.items()}}, request=request)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request):
        """Remove the tab icon (reset to default)."""
        user = request.user
        if not is_superadmin_user(user):
            return Response({"detail": "Only Super Admin can edit configuration."}, status=status.HTTP_403_FORBIDDEN)

        config, _ = SystemConfiguration.objects.get_or_create(pk=1)
        if config.tab_icon:
            config.tab_icon.delete(save=False)
            config.tab_icon = None
            config.save()
        serializer = SystemConfigurationSerializer(config, context={"request": request})
        return Response(serializer.data)


class UserActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only view for user activity logs.
    Only accessible to super admins.
    Includes a custom action for frontend page-view tracking.
    """

    permission_classes = [IsAuthenticated]
    queryset = UserActivityLog.objects.select_related("user").all()
    serializer_class = UserActivityLogSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False) or not self.request.user.is_authenticated:
            return UserActivityLog.objects.none()

        user = self.request.user
        is_super_admin = is_superadmin_user(user)

        if not is_super_admin:
            # Return empty queryset for non-super admins
            return UserActivityLog.objects.none()

        queryset = UserActivityLog.objects.select_related("user").all()

        # Filter by user if specified
        user_id = self.request.query_params.get("user_id")
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        # Filter by action if specified
        action = self.request.query_params.get("action")
        if action:
            queryset = queryset.filter(action=action)

        # Filter by resource if specified
        resource = self.request.query_params.get("resource")
        if resource:
            queryset = queryset.filter(resource__icontains=resource)

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")

        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        if end_date:
            queryset = queryset.filter(timestamp__lte=end_date)

        return queryset

    @action(detail=False, methods=["post"], url_path="log-page-view")
    def log_page_view(self, request):
        """
        Log a page view from the frontend.
        Any authenticated user can log their own page views.
        Payload: { "page": "/dashboard", "title": "Dashboard" }
        """
        user = request.user
        ext_user = None

        if isinstance(user, ExternalUser):
            ext_user = user
        else:
            worker_id = getattr(user, "worker_id", None)
            if worker_id:
                ext_user = ExternalUser.objects.filter(worker_id=worker_id).first()
            if not ext_user:
                username = getattr(user, "username", None)
                if username:
                    ext_user = ExternalUser.objects.filter(username__iexact=username).first()

        if not ext_user:
            return Response({"detail": "User not resolvable"}, status=status.HTTP_400_BAD_REQUEST)

        page = request.data.get("page", "")
        title = request.data.get("title", "")

        if not page:
            return Response({"detail": "page is required"}, status=status.HTTP_400_BAD_REQUEST)

        UserActivityLog.log_activity(
            user=ext_user,
            action="page_view",
            resource=page,
            details={"title": title} if title else {},
            request=request,
        )

        return Response({"status": "ok"}, status=status.HTTP_201_CREATED)


class TaskCommentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task comments with threading and mentions support.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskCommentSerializer
    pagination_class = None  # Task sub-resources are small per-task lists

    def get_queryset(self):
        queryset = TaskComment.objects.select_related("author", "task").prefetch_related(
            "mentions",
            models.Prefetch(
                "replies",
                queryset=TaskComment.objects.select_related("author").prefetch_related("mentions"),
            ),
        )

        # Require task_id to prevent enumerating all comments
        task_id = self.request.query_params.get("task_id")
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        elif self.action == "list":
            return queryset.none()

        # Only return top-level comments by default (replies are nested)
        if self.request.query_params.get("top_level", "true").lower() == "true":
            queryset = queryset.filter(parent__isnull=True)

        return queryset

    def perform_create(self, serializer):
        # Get or create the employee for the current user
        employee = get_employee_for_user(self.request.user)

        comment = serializer.save(author=employee)

        # Create activity log for comment
        TaskActivity.objects.create(task=comment.task, actor=employee, action="comment_added", extra_data={"comment_id": comment.id, "preview": comment.content[:100]})

        # Create notifications for mentioned users
        mentioned_employees = comment.mentions.all()
        if mentioned_employees.exists():
            task = comment.task

            # Determine valid mention targets based on group or assigned_to
            if task.group_id:
                valid_employee_ids = set(task.group.members.values_list("id", flat=True))
            else:
                valid_employee_ids = set(task.assigned_to.values_list("id", flat=True))

            from .consumers import send_notification_to_user

            eligible_mentions = [mentioned_emp for mentioned_emp in mentioned_employees if mentioned_emp.id in valid_employee_ids and mentioned_emp.id != employee.id]

            emp_ids = [m.emp_id for m in eligible_mentions if m.emp_id]
            ext_users = ExternalUser.objects.filter(worker_id__in=emp_ids, is_active=True)
            ext_user_map = {u.worker_id.lower(): u for u in ext_users if u.worker_id}

            for mentioned_emp in eligible_mentions:
                worker_id = (mentioned_emp.emp_id or "").lower()
                ext_user = ext_user_map.get(worker_id)
                if not ext_user:
                    logger.warning(f"No ExternalUser found for employee {mentioned_emp.name} (emp_id={mentioned_emp.emp_id})")
                    continue

                title = "You were mentioned in a comment"
                message = f'{employee.name} mentioned you in a comment on "{task.title}"'

                notif = Notification.objects.create(
                    recipient=ext_user,
                    title=title,
                    message=message,
                    event=task,
                    event_type="task_mention",
                )
                send_notification_to_user(
                    ext_user.id,
                    {
                        "id": notif.id,
                        "title": title,
                        "message": message,
                        "event_type": "task_mention",
                        "event_id": task.id,
                        "is_read": False,
                        "created_at": notif.created_at.isoformat(),
                    },
                )

        return comment

    def perform_update(self, serializer):
        comment = self.get_object()
        employee = get_employee_for_user(self.request.user)
        if comment.author_id != employee.id:
            raise PermissionDenied("You can only edit your own comments.")
        serializer.save(is_edited=True, edited_at=timezone.now())

    def perform_destroy(self, instance):
        employee = get_employee_for_user(self.request.user)
        # Only the comment author or a PTB admin can delete comments
        if instance.author_id != employee.id and not getattr(self.request.user, "is_ptb_admin", False):
            raise PermissionDenied("You can only delete your own comments.")
        instance.delete()

    @action(detail=True, methods=["get"])
    def replies(self, request, pk=None):
        """Get all replies for a comment"""
        comment = self.get_object()
        replies = comment.replies.all()
        serializer = self.get_serializer(replies, many=True)
        return Response(serializer.data)


class TaskActivityViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only ViewSet for task activity logs.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskActivitySerializer
    pagination_class = None

    def get_queryset(self):
        queryset = TaskActivity.objects.select_related("actor", "task").all()

        # Require task_id to prevent enumerating all activities
        task_id = self.request.query_params.get("task_id")
        if task_id:
            queryset = queryset.filter(task_id=task_id)
        elif self.action == "list":
            return queryset.none()

        # Limit results for performance (cap at 200)
        limit = self.request.query_params.get("limit", 50)
        try:
            limit = min(int(limit), 200)
        except ValueError:
            limit = 50

        # Use proper filtering instead of slicing to keep queryset operations safe
        pks = list(queryset.order_by("-created_at").values_list("pk", flat=True)[:limit])
        return queryset.filter(pk__in=pks).order_by("-created_at")


class TaskSubtaskViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task subtasks/checklist items.
    Restricted to users who are the creator or an assignee of the parent task.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskSubtaskSerializer
    pagination_class = None

    def _check_task_membership(self, task):
        """Verify the current user is the creator, an assignee, or a PTB admin of the task."""
        if getattr(self.request.user, "is_ptb_admin", False):
            return
        employee = get_employee_for_user(self.request.user)
        is_creator = task.created_by_id == employee.id
        is_assigned = task.assigned_to.filter(id=employee.id).exists() if hasattr(task.assigned_to, "filter") else False
        if not is_creator and not is_assigned:
            raise PermissionDenied("You can only manage subtasks on tasks you created or are assigned to.")

    def get_queryset(self):
        queryset = TaskSubtask.objects.select_related("task", "created_by", "completed_by")

        # Filter by task if specified
        task_id = self.request.query_params.get("task_id")
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset.order_by("order", "created_at")

    def perform_create(self, serializer):
        employee = get_employee_for_user(self.request.user)
        task = serializer.validated_data.get("task")
        if task:
            self._check_task_membership(task)

        subtask = serializer.save(created_by=employee)

        # Create activity log
        TaskActivity.objects.create(task=subtask.task, actor=employee, action="updated", extra_data={"type": "subtask_added", "subtask_title": subtask.title})

        return subtask

    def perform_destroy(self, instance):
        employee = get_employee_for_user(self.request.user, raise_if_not_found=False)

        if employee:
            TaskActivity.objects.create(task=instance.task, actor=employee, action="updated", extra_data={"type": "subtask_removed", "subtask_title": instance.title})

        instance.delete()

    @action(detail=True, methods=["post"])
    def toggle(self, request, pk=None):
        """Toggle subtask completion status"""
        subtask = self.get_object()
        employee = get_employee_for_user(request.user, raise_if_not_found=False)

        is_completed = subtask.toggle_complete(employee)

        # Create activity log
        if employee:
            TaskActivity.objects.create(task=subtask.task, actor=employee, action="updated", extra_data={"type": "subtask_toggled", "subtask_title": subtask.title, "is_completed": is_completed})

        serializer = self.get_serializer(subtask)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def reorder(self, request):
        """Reorder subtasks within a task"""
        task_id = request.data.get("task_id")
        subtask_ids = request.data.get("subtask_ids", [])

        if not task_id or not subtask_ids:
            return Response({"error": "task_id and subtask_ids required"}, status=400)

        # Bulk update order using Case/When (single query instead of N queries)
        from django.db.models import Case, IntegerField, Value, When

        cases = [When(id=sid, then=Value(idx)) for idx, sid in enumerate(subtask_ids)]
        TaskSubtask.objects.filter(id__in=subtask_ids, task_id=task_id).update(order=Case(*cases, output_field=IntegerField()))

        # Return updated subtasks
        subtasks = TaskSubtask.objects.filter(task_id=task_id).order_by("order")
        serializer = self.get_serializer(subtasks, many=True)
        return Response(serializer.data)


class TaskTimeLogViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task time logs and timer functionality.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskTimeLogSerializer
    pagination_class = None

    def get_queryset(self):
        queryset = TaskTimeLog.objects.select_related("task", "employee")

        # Filter by task if specified
        task_id = self.request.query_params.get("task_id")
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        # Filter by employee if specified
        employee_id = self.request.query_params.get("employee_id")
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        return queryset.order_by("-started_at")

    def perform_create(self, serializer):
        employee = get_employee_for_user(self.request.user)

        # Stop any running timers for this employee first (use stop_timer for proper duration calc)
        running_timers = TaskTimeLog.objects.filter(employee=employee, is_running=True)
        for timer in running_timers:
            timer.stop_timer()

        time_log = serializer.save(employee=employee)

        # Create activity log
        TaskActivity.objects.create(task=time_log.task, actor=employee, action="updated", extra_data={"type": "time_logged", "duration": time_log.duration_minutes})

        return time_log

    def perform_destroy(self, instance):
        employee = get_employee_for_user(self.request.user)
        # Only the time log owner or a PTB admin can delete entries
        if instance.employee_id != employee.id and not getattr(self.request.user, "is_ptb_admin", False):
            raise PermissionDenied("You can only delete your own time entries.")
        instance.delete()

    @action(detail=False, methods=["post"])
    def start_timer(self, request):
        """Start a new timer for a task"""
        task_id = request.data.get("task_id")
        description = request.data.get("description", "")

        if not task_id:
            return Response({"error": "task_id required"}, status=400)

        employee = get_employee_for_user(request.user)

        # Stop any existing running timer for this user
        existing_running = TaskTimeLog.objects.filter(employee=employee, is_running=True).first()
        if existing_running:
            existing_running.stop_timer()

        # Create new timer
        time_log = TaskTimeLog.objects.create(task_id=task_id, employee=employee, description=description, started_at=timezone.now(), is_running=True)

        serializer = self.get_serializer(time_log)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=["post"])
    def stop_timer(self, request, pk=None):
        """Stop a running timer"""
        time_log = self.get_object()

        if not time_log.is_running:
            return Response({"error": "Timer is not running"}, status=400)

        time_log.stop_timer()

        # Create activity log
        employee = get_employee_for_user(request.user, raise_if_not_found=False)
        if employee:
            TaskActivity.objects.create(task=time_log.task, actor=employee, action="updated", extra_data={"type": "time_logged", "duration": time_log.duration_minutes})

        serializer = self.get_serializer(time_log)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def active(self, request):
        """Get currently running timer for the user"""
        employee = get_employee_for_user(request.user, raise_if_not_found=False)

        if not employee:
            return Response(None)

        running = TaskTimeLog.objects.filter(employee=employee, is_running=True).first()
        if running:
            serializer = self.get_serializer(running)
            return Response(serializer.data)
        return Response(None)

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Get time summary for a task"""
        task_id = request.query_params.get("task_id")

        if not task_id:
            return Response({"error": "task_id required"}, status=400)

        from django.db.models import Sum

        logs = TaskTimeLog.objects.filter(task_id=task_id)
        total_minutes = logs.aggregate(total=Sum("duration_minutes"))["total"] or 0

        # Get task estimated hours
        try:
            task = CalendarEvent.objects.get(id=task_id)
            estimated_hours = float(task.estimated_hours or 0)
        except CalendarEvent.DoesNotExist:
            estimated_hours = 0

        return Response({"task_id": task_id, "total_minutes": total_minutes, "total_hours": round(total_minutes / 60, 2), "estimated_hours": estimated_hours, "log_count": logs.count()})


class BoardPresenceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for tracking board presence (who's viewing/editing).
    """

    permission_classes = [IsAuthenticated]
    serializer_class = BoardPresenceSerializer

    def get_queryset(self):
        # Only return presence records from last 5 minutes
        from datetime import timedelta

        cutoff = timezone.now() - timedelta(minutes=5)
        return BoardPresence.objects.filter(last_seen__gte=cutoff).select_related("user", "editing_task")

    @action(detail=False, methods=["post"])
    def heartbeat(self, request):
        """
        Update presence heartbeat. Called periodically by clients.
        """
        employee = get_employee_for_user(request.user, raise_if_not_found=False)

        if not employee:
            return Response({"error": "Employee not found"}, status=status.HTTP_400_BAD_REQUEST)

        editing_task_id = request.data.get("editing_task_id")
        channel_name = request.data.get("channel_name", "")

        # Update or create presence record
        presence, created = BoardPresence.objects.update_or_create(user=employee, defaults={"editing_task_id": editing_task_id, "channel_name": channel_name})

        # Get all current viewers (excluding self)
        from datetime import timedelta

        cutoff = timezone.now() - timedelta(minutes=5)
        viewers = BoardPresence.objects.filter(last_seen__gte=cutoff).exclude(user=employee).select_related("user", "editing_task")

        serializer = self.get_serializer(viewers, many=True)
        return Response({"status": "ok", "viewers": serializer.data})

    @action(detail=False, methods=["post"])
    def leave(self, request):
        """
        Called when user leaves the board.
        """
        employee = get_employee_for_user(request.user, raise_if_not_found=False)

        if employee:
            BoardPresence.objects.filter(user=employee).delete()

        return Response({"status": "ok"})


class PersonalNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet for personal notes (sticky notes / personal board).
    Users can only see and manage their own notes.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = PersonalNoteSerializer
    pagination_class = None  # Return all notes without pagination

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return PersonalNote.objects.none()
        if not getattr(self.request, "user", None) or not self.request.user.is_authenticated:
            return PersonalNote.objects.none()
        return PersonalNote.objects.filter(owner=self.request.user).order_by("-is_pinned", "order", "-created_at")

    def get_object(self):
        """Enforce object-level ownership — users can only access their own notes."""
        obj = super().get_object()
        if obj.owner_id != self.request.user.id:
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("You do not have permission to access this note.")
        return obj

    def perform_create(self, serializer):
        # Get max order for user's notes
        max_order = PersonalNote.objects.filter(owner=self.request.user).aggregate(max_order=models.Max("order"))["max_order"] or 0
        serializer.save(owner=self.request.user, order=max_order + 1)

    @action(detail=False, methods=["post"])
    def reorder(self, request):
        """
        Reorder notes. Expects: { "order": [id1, id2, id3, ...] }
        """
        order_list = request.data.get("order", [])

        if order_list:
            from django.db.models import Case, IntegerField, Value, When

            cases = [When(id=nid, then=Value(idx)) for idx, nid in enumerate(order_list)]
            PersonalNote.objects.filter(id__in=order_list, owner=request.user).update(order=Case(*cases, output_field=IntegerField()))

        return Response({"status": "ok"})

    @action(detail=True, methods=["post"])
    def toggle_pin(self, request, pk=None):
        """Toggle pinned status of a note."""
        note = self.get_object()
        note.is_pinned = not note.is_pinned
        note.save(update_fields=["is_pinned"])
        return Response(self.get_serializer(note).data)

    @action(detail=True, methods=["post"])
    def toggle_complete(self, request, pk=None):
        """Toggle completed status of a note."""
        note = self.get_object()
        note.is_completed = not note.is_completed
        note.save(update_fields=["is_completed"])
        return Response(self.get_serializer(note).data)


class TaskGroupViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task groups (folders/categories for organizing tasks).
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskGroupSerializer
    pagination_class = None

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return TaskGroup.objects.none()
        user = self.request.user
        if not user or not user.is_authenticated:
            return TaskGroup.objects.none()
        # User can see groups they created or groups they are a member of (if not private)
        worker_id = getattr(user, "worker_id", None)
        employee = None
        if worker_id:
            employee = Employee.objects.filter(emp_id=worker_id).first()

        base_qs = TaskGroup.objects.select_related("created_by", "department").prefetch_related("members").annotate(_task_count=models.Count("tasks"))

        if employee:
            return base_qs.filter(models.Q(created_by=user) | models.Q(is_private=False) | models.Q(members=employee)).distinct().order_by("order", "name")
        else:
            return base_qs.filter(models.Q(created_by=user) | models.Q(is_private=False)).distinct().order_by("order", "name")

    def perform_create(self, serializer):
        max_order = TaskGroup.objects.filter(created_by=self.request.user).aggregate(max_order=models.Max("order"))["max_order"] or 0
        instance = serializer.save(created_by=self.request.user, order=max_order + 1)
        # Notify members that they were added to the new group
        self._notify_group_members(instance, list(instance.members.all()))

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.is_department_group:
            raise PermissionDenied("Department groups cannot be edited. Use 'Sync Departments' to update them.")
        serializer.save()

    def perform_destroy(self, instance):
        if instance.is_department_group:
            raise PermissionDenied("Department groups cannot be deleted.")
        instance.delete()

    def _notify_group_members(self, group, members):
        """Notify employees that they were added to a task group."""
        from .signals import send_websocket_notification

        emp_ids = [e.emp_id for e in members if e.emp_id]
        if not emp_ids:
            return

        user_map = {}
        for eu in ExternalUser.objects.filter(worker_id__in=emp_ids):
            user_map[eu.worker_id.lower()] = eu

        creator = self.request.user
        creator_name = getattr(creator, "username", str(creator))

        notifs_to_create = []
        ws_payloads = []
        for member in members:
            if not member.emp_id:
                continue
            external_user = user_map.get(member.emp_id.lower())
            if not external_user:
                continue
            # Don't notify the creator themselves
            if external_user.id == creator.id:
                continue

            title = f"Added to Group: {group.name}"
            message = f"{creator_name} added you as a member of the group '{group.name}'."

            notifs_to_create.append(
                Notification(
                    recipient=external_user,
                    title=title,
                    message=message,
                    event_type="task_group",
                )
            )
            ws_payloads.append(
                (
                    external_user.id,
                    {
                        "title": title,
                        "message": message,
                        "event_type": "task_group",
                        "is_read": False,
                    },
                )
            )

        if notifs_to_create:
            try:
                created = Notification.objects.bulk_create(notifs_to_create)
                for notif, (uid, payload) in zip(created, ws_payloads, strict=True):
                    payload["id"] = notif.id
                    payload["created_at"] = notif.created_at.isoformat()
                    send_websocket_notification(uid, payload)
            except Exception as e:
                logger.error(f"Error creating group member notifications: {e}")

    @action(detail=True, methods=["post"])
    def add_member(self, request, pk=None):
        """Add a member to the group. Only the group creator or admin can modify membership."""
        group = self.get_object()
        if group.created_by != request.user and not getattr(request.user, "is_ptb_admin", False):
            raise PermissionDenied("Only the group creator or an admin can modify membership.")
        employee_id = request.data.get("employee_id")

        if not employee_id:
            return Response({"error": "employee_id required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id)
            group.members.add(employee)
            # Notify the newly added member
            self._notify_group_members(group, [employee])
            return Response(self.get_serializer(group).data)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=["post"])
    def remove_member(self, request, pk=None):
        """Remove a member from the group. Only the group creator or admin can modify membership."""
        group = self.get_object()
        if group.created_by != request.user and not getattr(request.user, "is_ptb_admin", False):
            raise PermissionDenied("Only the group creator or an admin can modify membership.")
        employee_id = request.data.get("employee_id")

        if not employee_id:
            return Response({"error": "employee_id required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = Employee.objects.get(id=employee_id)
            group.members.remove(employee)
            return Response(self.get_serializer(group).data)
        except Employee.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=["post"])
    def reorder(self, request):
        """Reorder groups. Expects: { "order": [id1, id2, id3, ...] }"""
        order_list = request.data.get("order", [])

        if order_list:
            from django.db.models import Case, IntegerField, Value, When

            cases = [When(id=gid, then=Value(idx)) for idx, gid in enumerate(order_list)]
            TaskGroup.objects.filter(id__in=order_list, created_by=request.user).update(order=Case(*cases, output_field=IntegerField()))

        return Response({"status": "ok"})

    @action(detail=False, methods=["post"])
    def sync_departments(self, request):
        """
        Create or update groups for each department with their employees as members.
        Uses department code/name as the group name.
        Admin-only action.
        """
        if not (getattr(request.user, "is_ptb_admin", False) or is_superadmin_user(request.user)):
            raise PermissionDenied("Only administrators can sync department groups.")

        from .models import Department

        # Define colors for departments (cycling through a palette)
        colors = [
            "#3B82F6",  # Blue
            "#10B981",  # Emerald
            "#F59E0B",  # Amber
            "#EF4444",  # Red
            "#8B5CF6",  # Violet
            "#EC4899",  # Pink
            "#06B6D4",  # Cyan
            "#84CC16",  # Lime
            "#F97316",  # Orange
            "#6366F1",  # Indigo
        ]

        departments = Department.objects.prefetch_related("employees").all()
        created_count = 0
        updated_count = 0

        for idx, dept in enumerate(departments):
            # Check if a department group already exists
            existing_group = TaskGroup.objects.filter(is_department_group=True, department=dept).first()

            # Get all employee IDs for this department
            employee_ids = list(dept.employees.values_list("id", flat=True))

            if existing_group:
                # Update existing group's members and name (use dept code)
                existing_group.name = dept.code
                existing_group.save(update_fields=["name"])
                existing_group.members.set(employee_ids)
                updated_count += 1
            else:
                # Create new department group
                color = colors[idx % len(colors)]
                group = TaskGroup.objects.create(name=dept.code, description=f"Default group for {dept.name} department", color=color, created_by=request.user, is_department_group=True, department=dept, is_private=False, order=idx)
                group.members.set(employee_ids)
                created_count += 1

        return Response({"status": "ok", "created": created_count, "updated": updated_count})


class TaskAttachmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task attachments (files attached to tasks).
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskAttachmentSerializer
    parser_classes = [MultiPartParser, FormParser]
    pagination_class = None

    def get_queryset(self):
        queryset = TaskAttachment.objects.select_related("task", "uploaded_by")

        # Filter by task if provided
        task_id = self.request.query_params.get("task_id")
        if task_id:
            queryset = queryset.filter(task_id=task_id)

        return queryset.order_by("-created_at")

    # 10 MB upload limit
    MAX_UPLOAD_SIZE = 10 * 1024 * 1024
    ALLOWED_CONTENT_TYPES = {
        # Documents
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        # Text
        "text/plain",
        "text/csv",
        "text/xml",
        "application/json",
        "application/xml",
        # Archives
        "application/zip",
        "application/x-rar-compressed",
        "application/x-7z-compressed",
        "application/gzip",
        # Images
        "image/png",
        "image/jpeg",
        "image/gif",
        "image/svg+xml",
        "image/webp",
        "image/bmp",
    }

    def perform_create(self, serializer):
        user = self.request.user
        worker_id = getattr(user, "worker_id", None)
        employee = None
        if worker_id:
            # Note: Employee model uses emp_id field, not worker_id
            employee = Employee.objects.filter(emp_id=worker_id).first()

        # Get file info
        file = self.request.FILES.get("file")
        if file:
            # Validate file size
            if file.size > self.MAX_UPLOAD_SIZE:
                raise serializers.ValidationError({"file": f"File size exceeds {self.MAX_UPLOAD_SIZE // (1024 * 1024)} MB limit."})
            # Validate content type
            if file.content_type and file.content_type not in self.ALLOWED_CONTENT_TYPES:
                raise serializers.ValidationError({"file": f'File type "{file.content_type}" is not allowed.'})
            serializer.save(uploaded_by=employee, filename=file.name, file_size=file.size, file_type=file.content_type or "")
        else:
            serializer.save(uploaded_by=employee)

    def perform_destroy(self, instance):
        employee = get_employee_for_user(self.request.user, raise_if_not_found=False)
        # Only the uploader or a PTB admin can delete attachments
        is_admin = getattr(self.request.user, "is_ptb_admin", False)
        if employee and instance.uploaded_by_id != employee.id and not is_admin:
            raise PermissionDenied("You can only delete your own attachments.")
        instance.delete()

    @action(detail=False, methods=["get"])
    def by_task(self, request):
        """Get all attachments for a specific task."""
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response({"error": "task_id required"}, status=status.HTTP_400_BAD_REQUEST)

        attachments = self.get_queryset().filter(task_id=task_id)
        serializer = self.get_serializer(attachments, many=True)
        return Response(serializer.data)


class TaskReminderViewSet(viewsets.ModelViewSet):
    """
    ViewSet for task reminders/alerts.
    Users can only see and manage their own reminders.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = TaskReminderSerializer
    pagination_class = None

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return TaskReminder.objects.none()
        if not getattr(self.request, "user", None) or not self.request.user.is_authenticated:
            return TaskReminder.objects.none()
        return TaskReminder.objects.filter(user=self.request.user).select_related("task").order_by("remind_at")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"])
    def pending(self, request):
        """Get pending (not triggered and not dismissed) reminders."""
        reminders = self.get_queryset().filter(is_triggered=False, is_dismissed=False, remind_at__gte=timezone.now())
        serializer = self.get_serializer(reminders, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def due(self, request):
        """Get reminders that are due (remind_at <= now and not yet triggered)."""
        reminders = self.get_queryset().filter(is_triggered=False, is_dismissed=False, remind_at__lte=timezone.now())
        serializer = self.get_serializer(reminders, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    def trigger(self, request, pk=None):
        """Mark a reminder as triggered."""
        reminder = self.get_object()
        reminder.is_triggered = True
        reminder.save(update_fields=["is_triggered"])
        return Response(self.get_serializer(reminder).data)

    @action(detail=True, methods=["post"])
    def dismiss(self, request, pk=None):
        """Dismiss a reminder."""
        reminder = self.get_object()
        reminder.is_dismissed = True
        reminder.save(update_fields=["is_dismissed"])
        return Response(self.get_serializer(reminder).data)

    @action(detail=False, methods=["get"])
    def by_task(self, request):
        """Get all reminders for a specific task."""
        task_id = request.query_params.get("task_id")
        if not task_id:
            return Response({"error": "task_id required"}, status=status.HTTP_400_BAD_REQUEST)

        reminders = self.get_queryset().filter(task_id=task_id)
        serializer = self.get_serializer(reminders, many=True)
        return Response(serializer.data)


class PurchaseRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing purchase requests.
    Supports CRUD operations, filtering, and import/export.
    """

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "purchasing"
    queryset = PurchaseRequest.objects.all()
    serializer_class = PurchaseRequestSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["owner", "doc_id", "part_no", "description_spec", "pr_no", "remarks"]
    ordering_fields = ["id", "request_date", "owner", "doc_id", "part_no", "pr_no", "status", "created_at"]
    ordering = ["-request_date", "-id"]

    def get_queryset(self):
        queryset = PurchaseRequest.objects.select_related("owner_employee", "owner_employee__department").all()

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            queryset = queryset.filter(request_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(request_date__lte=end_date)

        return queryset

    @swagger_auto_schema(
        operation_summary="Import purchase requests from CSV/Excel",
        manual_parameters=[
            openapi.Parameter(name="file", in_=openapi.IN_FORM, type=openapi.TYPE_FILE, required=True),
        ],
    )
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Import purchase requests from CSV or Excel file.
        Supports multi-sheet Excel files with sheets:
        - 'List of Purchase' or active sheet -> status=pending
        - 'Done' -> status=done
        - 'Cancel Purchase' -> status=canceled
        """
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        import csv
        import io
        from datetime import datetime

        def parse_date(val):
            if not val:
                return None
            for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(str(val), fmt).date()
                except (ValueError, TypeError):
                    continue
            return None

        def process_rows(rows, default_status="pending"):
            batch = []
            for row in rows:
                pr_data = {
                    "request_date": parse_date(row.get("Request Date")),
                    "owner": row.get("Owner") or None,
                    "doc_id": row.get("Doc ID") or None,
                    "part_no": row.get("Part No.") or None,
                    "description_spec": row.get("Description-Spec") or None,
                    "material_category": row.get("Material Category") or None,
                    "purpose_desc": row.get("Purpose/Desc.") or None,
                    "qty": row.get("Qty") or 1,
                    "plant": row.get("Plant") or None,
                    "project_code": row.get("Project Code") or None,
                    "pr_type": row.get("PR Type") or None,
                    "mrp_id": row.get("MRPID") or None,
                    "purch_org": row.get("Purch. Org.") or None,
                    "sourcer_price": row.get("Sourcer Price") or None,
                    "pr_no": row.get("PR No.") or None,
                    "remarks": row.get("Remarks") or None,
                    "status": default_status,
                }
                batch.append(PurchaseRequest(**pr_data))
            if batch:
                PurchaseRequest.objects.bulk_create(batch, batch_size=200)
            return len(batch)

        def read_sheet(ws):
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row, strict=True))
                # Skip empty rows
                if any(v for v in row_dict.values() if v is not None):
                    rows.append(row_dict)
            return rows

        try:
            if file.name.endswith(".csv"):
                content = file.read().decode("utf-8")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
                created_count = process_rows(rows, "pending")
                return Response({"message": f"Successfully imported {created_count} purchase requests"})

            elif file.name.endswith(".xlsx") or file.name.endswith(".xls"):
                import openpyxl

                wb = openpyxl.load_workbook(file)
                sheet_names = wb.sheetnames

                # Sheet-to-status mapping
                sheet_status_map = {
                    "List of Purchase": "pending",
                    "Done": "done",
                    "Cancel Purchase": "canceled",
                }

                total_created = 0
                sheets_processed = []

                # Check if workbook has our expected sheet names
                has_named_sheets = any(name in sheet_names for name in sheet_status_map)

                if has_named_sheets:
                    # Multi-sheet import
                    for sheet_name, default_status in sheet_status_map.items():
                        if sheet_name in sheet_names:
                            ws = wb[sheet_name]
                            rows = read_sheet(ws)
                            if rows:
                                count = process_rows(rows, default_status)
                                total_created += count
                                sheets_processed.append(f"{sheet_name}: {count}")
                else:
                    # Fallback: read active sheet as pending
                    ws = wb.active
                    rows = read_sheet(ws)
                    total_created = process_rows(rows, "pending")
                    sheets_processed.append(f"{ws.title}: {total_created}")

                detail = ", ".join(sheets_processed)
                return Response({"message": f"Successfully imported {total_created} purchase requests ({detail})", "created": total_created})
            else:
                return Response({"error": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)

        except Exception:
            return Response({"error": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """Update a purchase request and notify on status change."""
        from .signals import notify_purchase_request_status_change

        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        old_status = instance.status

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Check for status change and notify
        new_status = serializer.validated_data.get("status", instance.status)
        if old_status != new_status and new_status in ["done", "canceled"]:
            notify_purchase_request_status_change(instance, old_status, new_status)

        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        """Partial update a purchase request."""
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """Delete multiple purchase requests at once."""
        if not is_ptb_admin(request.user) and not is_superadmin_user(request.user):
            return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = PurchaseRequest.objects.filter(id__in=ids).delete()
        return Response({"message": f"Successfully deleted {deleted_count} purchase requests", "deleted": deleted_count})

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        """Update status of multiple purchase requests at once."""
        if not is_ptb_admin(request.user) and not is_superadmin_user(request.user):
            return Response({"error": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get("ids", [])
        new_status = request.data.get("status", "")
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        if new_status not in ["pending", "done", "canceled"]:
            return Response({"error": "Invalid status"}, status=status.HTTP_400_BAD_REQUEST)
        updated_count = PurchaseRequest.objects.filter(id__in=ids).update(status=new_status)
        return Response({"message": f"Successfully updated {updated_count} purchase requests to {new_status}", "updated": updated_count})


class AssetViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing assets.
    Supports CRUD operations, filtering by department, and import/export.
    """

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "assets"
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["asset_id", "part_number", "product_name", "keeper_name", "cost_center"]
    ordering_fields = ["id", "asset_id", "part_number", "product_name", "keeper_name", "status", "receive_date", "created_at"]
    ordering = ["-receive_date", "asset_id"]

    def get_queryset(self):
        queryset = Asset.objects.select_related("department").all()

        # Filter by department
        department_id = self.request.query_params.get("department")
        if department_id:
            queryset = queryset.filter(department_id=department_id)

        # Filter by cost center
        cost_center = self.request.query_params.get("cost_center")
        if cost_center:
            queryset = queryset.filter(cost_center=cost_center)

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return AssetSummarySerializer
        return AssetSerializer

    @swagger_auto_schema(
        operation_summary="Get assets grouped by department",
        responses={200: openapi.Response("Assets grouped by department")},
    )
    @action(detail=False, methods=["get"])
    def by_department(self, request):
        """Get all assets grouped by department, with optional search"""
        departments = Department.objects.filter(is_enabled=True).order_by("code")

        # Fetch all assets once and group in Python instead of N+1 queries
        all_assets_qs = Asset.objects.select_related("department").all()

        # Apply search filter if provided
        search_query = request.query_params.get("search", "").strip()
        if search_query:
            from django.db.models import Q

            all_assets_qs = all_assets_qs.filter(Q(asset_id__icontains=search_query) | Q(part_number__icontains=search_query) | Q(product_name__icontains=search_query) | Q(keeper_name__icontains=search_query) | Q(cost_center__icontains=search_query))

        all_assets = list(all_assets_qs)

        # Build lookup: dept_code -> list of assets
        dept_assets_map = {}
        unassigned_assets = []
        dept_codes = {dept.code: dept for dept in departments}

        for asset in all_assets:
            matched_dept = None
            if asset.department and asset.department.code in dept_codes:
                matched_dept = asset.department.code
            elif asset.cost_center and asset.cost_center in dept_codes:
                matched_dept = asset.cost_center
            elif asset.keeper_dept and asset.keeper_dept in dept_codes:
                matched_dept = asset.keeper_dept

            if matched_dept:
                dept_assets_map.setdefault(matched_dept, []).append(asset)
            else:
                unassigned_assets.append(asset)

        result = []
        for dept in departments:
            assets = dept_assets_map.get(dept.code, [])
            serializer = AssetSummarySerializer(assets, many=True)
            result.append({"department_id": dept.id, "department_code": dept.code, "department_name": dept.name, "asset_count": len(assets), "assets": serializer.data})

        if unassigned_assets:
            serializer = AssetSummarySerializer(unassigned_assets, many=True)
            result.append({"department_id": None, "department_code": "UNASSIGNED", "department_name": "Unassigned Assets", "asset_count": len(unassigned_assets), "assets": serializer.data})

        return Response(result)

    @swagger_auto_schema(
        operation_summary="Import assets from CSV/Excel",
        manual_parameters=[
            openapi.Parameter(name="file", in_=openapi.IN_FORM, type=openapi.TYPE_FILE, required=True),
        ],
    )
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Import assets from CSV or Excel file"""
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        import csv
        import io

        try:
            # Read file content
            if file.name.endswith(".csv"):
                content = file.read().decode("utf-8")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif file.name.endswith(".xlsx") or file.name.endswith(".xls"):
                import openpyxl

                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row, strict=True)))
            else:
                return Response({"error": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0

            # Pre-load departments for batch lookup
            dept_map = {d.code: d for d in Department.objects.all()}

            # Collect all asset_ids from the rows and pre-load existing assets
            row_asset_ids = [row.get("Asset Id") for row in rows if row.get("Asset Id")]
            existing_assets = {a.asset_id: a for a in Asset.objects.filter(asset_id__in=row_asset_ids)}

            to_create = []
            to_update = []

            for row in rows:
                asset_id = row.get("Asset Id")
                if not asset_id:
                    continue

                # Parse boolean values
                def parse_bool(val):
                    if isinstance(val, bool):
                        return val
                    return str(val).upper() in ["Y", "YES", "TRUE", "1"] if val else False

                # Parse date values
                def parse_date(val):
                    if not val:
                        return None
                    from datetime import datetime

                    for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                        try:
                            return datetime.strptime(str(val), fmt).date()
                        except ValueError:
                            continue
                    return None

                # Parse decimal values
                def parse_decimal(val):
                    if not val:
                        return None
                    try:
                        return float(str(val).replace(",", ""))
                    except ValueError:
                        return None

                # Map CSV headers to model fields
                asset_data = {
                    "company_code": row.get("Company Code") or None,
                    "fixed_asset_id": row.get("Fixed Asset Id") or None,
                    "is_fixed_asset": parse_bool(row.get("Is Fixed Asset")),
                    "is_customs_control": parse_bool(row.get("Is Customs Control")),
                    "part_number": row.get("Part Number") or None,
                    "group_3": row.get("Group 3") or None,
                    "product_name": row.get("Product Name") or None,
                    "spec": row.get("Spec") or None,
                    "quantity": parse_decimal(row.get("Quantity")) or 1,
                    "receive_date": parse_date(row.get("Receive Date")),
                    "status": row.get("Status") or None,
                    "cost_center": row.get("Cost Center") or None,
                    "cost_center_name": row.get("Cost Center Name") or None,
                    "keeper_dept": row.get("Keeper Dept.") or None,
                    "keeper_dept_name": row.get("Keeper Dept. Name") or None,
                    "keeper": row.get("Keeper") or None,
                    "keeper_name": row.get("Keeper Name") or None,
                    "group_1": row.get("Group 1") or None,
                    "group_2": row.get("Group 2") or None,
                    "storage": row.get("Storage") or None,
                    "location_code": row.get("Location Code") or None,
                    "storage_desc": row.get("Storage Desc.") or None,
                    "consign": row.get("Consign") or None,
                    "vendor": row.get("Vendor") or None,
                    "pr_no": row.get("PR No.") or None,
                    "pr_sequence": row.get("PR Sequence") or None,
                    "po_no": row.get("PO No.") or None,
                    "po_sequence": row.get("PO Sequence") or None,
                    "dn_no": row.get("DN No.") or None,
                    "dn_sequence": row.get("DN Sequence") or None,
                    "dn_date": parse_date(row.get("DN Date")),
                    "application_number": row.get("Application Number") or None,
                    "sequence": row.get("Sequence") or None,
                    "import_number": row.get("Import Number") or None,
                    "picking_no": row.get("Picking No.") or None,
                    "picking_sequence": row.get("Picking Sequence") or None,
                    "picking_year": row.get("Picking Year") or None,
                    "picking_date": parse_date(row.get("Picking Date")),
                    "chinese_product_name": row.get("Chinese Product Name") or None,
                    "hs_code": row.get("HS Code") or None,
                    "declaration_number": row.get("Declaration Number") or None,
                    "declaration_date": parse_date(row.get("Declaration Date")),
                    "control_end_date": parse_date(row.get("Control End Date")),
                    "outsource_number": row.get("Outsource Number") or None,
                    "price": parse_decimal(row.get("Price")),
                    "currency": row.get("Currency") or None,
                    "local_price": parse_decimal(row.get("Local Price")),
                    "price_level": row.get("Price Level") or None,
                    "sn": row.get("SN") or None,
                    "is_qualified": parse_bool(row.get("Is Qualified")),
                    "itc_end_date": parse_date(row.get("ITC End Date")),
                    "elec_declaration_number": row.get("Elec Declaration Number") or None,
                    "national_inspection_certification": row.get("National Inspection Certification") or None,
                    "note1": row.get("Note1") or None,
                    "note2": row.get("Note2") or None,
                    "note3": row.get("Note3") or None,
                    "note4": row.get("Note4") or None,
                    "note5": row.get("Note5") or None,
                    "note6": row.get("Note6") or None,
                    "note7": row.get("Note7") or None,
                    "note8": row.get("Note8") or None,
                    "note9": row.get("Note9") or None,
                    "note10": row.get("Note10") or None,
                }

                # Link to department using pre-loaded map
                dept_code = asset_data.get("cost_center") or asset_data.get("keeper_dept")
                dept = dept_map.get(dept_code) if dept_code else None

                if asset_id in existing_assets:
                    # Update existing
                    existing = existing_assets[asset_id]
                    for field, value in asset_data.items():
                        setattr(existing, field, value)
                    if dept:
                        existing.department = dept
                    to_update.append(existing)
                    updated_count += 1
                else:
                    # Create new
                    new_asset = Asset(asset_id=asset_id, **asset_data)
                    if dept:
                        new_asset.department = dept
                    to_create.append(new_asset)
                    created_count += 1

            if to_create:
                Asset.objects.bulk_create(to_create, batch_size=200)
            if to_update:
                update_fields = list(asset_data.keys()) + ["department"]
                Asset.objects.bulk_update(to_update, update_fields, batch_size=200)

            return Response({"message": "Successfully imported assets", "created": created_count, "updated": updated_count})

        except Exception:
            import traceback

            logger.error(f"Asset import error: {traceback.format_exc()}")
            return Response(
                {
                    "error": "An error occurred while importing assets. Please check the file format.",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """Delete multiple assets at once."""
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = Asset.objects.filter(id__in=ids).delete()
        return Response({"message": f"Successfully deleted {deleted_count} assets", "deleted": deleted_count})

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        """Update status of multiple assets at once."""
        ids = request.data.get("ids", [])
        new_status = request.data.get("status", "")
        if not ids:
            return Response({"error": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        if not new_status:
            return Response({"error": "No status provided"}, status=status.HTTP_400_BAD_REQUEST)
        updated_count = Asset.objects.filter(id__in=ids).update(status=new_status)
        return Response({"message": f"Successfully updated {updated_count} assets to {new_status}", "updated": updated_count})


# ===========================================================================
# SMB Configuration (Super Admin only — multi-config with active selection)
# ===========================================================================


class SMBConfigurationViewSet(viewsets.ModelViewSet):
    """
    CRUD for SMB configurations. Super admin only.
    Supports multiple configs — one marked `is_active` is used for file uploads.
    Custom actions: activate (POST), test (POST).
    """

    permission_classes = [IsAuthenticated]
    serializer_class = SMBConfigurationSerializer
    queryset = SMBConfiguration.objects.all()

    def list(self, request, *args, **kwargs):
        if not is_superadmin_user(request.user):
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        if not is_superadmin_user(request.user):
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
        return super().retrieve(request, *args, **kwargs)

    def perform_create(self, serializer):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can create SMB configs.")
        instance = serializer.save()
        # Invalidate cached SMB config in ExcelGenerator
        from .utils.excel_generator import ExcelGenerator

        ExcelGenerator.invalidate_smb_cache()
        UserActivityLog.log_activity(
            user=self.request.user,
            action="create",
            resource="smb_configuration",
            resource_id=instance.id,
            details={"name": instance.name, "server": instance.server},
            request=self.request,
        )

    def perform_update(self, serializer):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can update SMB configs.")
        instance = serializer.save()
        from .utils.excel_generator import ExcelGenerator

        ExcelGenerator.invalidate_smb_cache()
        UserActivityLog.log_activity(
            user=self.request.user,
            action="update",
            resource="smb_configuration",
            resource_id=instance.id,
            details={"fields": list(self.request.data.keys())},
            request=self.request,
        )

    def perform_destroy(self, instance):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can delete SMB configs.")
        from .utils.excel_generator import ExcelGenerator

        ExcelGenerator.invalidate_smb_cache()
        UserActivityLog.log_activity(
            user=self.request.user,
            action="delete",
            resource="smb_configuration",
            resource_id=instance.id,
            details={"name": instance.name},
            request=self.request,
        )
        instance.delete()

    @action(detail=True, methods=["post"], url_path="activate")
    def activate(self, request, pk=None):
        """Set this config as the active one (deactivates all others)."""
        if not is_superadmin_user(request.user):
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
        config = self.get_object()
        config.is_active = True
        config.save()  # save() deactivates others
        from .utils.excel_generator import ExcelGenerator

        ExcelGenerator.invalidate_smb_cache()
        UserActivityLog.log_activity(
            user=request.user,
            action="update",
            resource="smb_configuration",
            resource_id=config.id,
            details={"action": "activate", "name": config.name},
            request=request,
        )
        return Response(SMBConfigurationSerializer(config).data)

    @action(detail=True, methods=["post"], url_path="test")
    def test_connection(self, request, pk=None):
        """Test SMB connection for a specific config."""
        if not is_superadmin_user(request.user):
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

        config = self.get_object()
        password = config.get_password()
        if not config.server or not config.username:
            return Response({"success": False, "error": "Missing host or username"})
        if not password:
            return Response({"success": False, "error": "Password could not be decrypted. Please re-enter the password and save before testing."})

        try:
            import socket

            from smb.SMBConnection import SMBConnection

            client_name = socket.gethostname()
            conn = SMBConnection(
                config.username,
                password,
                client_name,
                config.server,
                use_ntlm_v2=True,
                is_direct_tcp=True,
            )
            connected = conn.connect(config.server, int(config.port or 445), timeout=10)
            if connected:
                shares = conn.listShares()
                share_names = [s.name for s in shares]
                conn.close()
                return Response(
                    {
                        "success": True,
                        "message": f"Connected successfully. Found {len(shares)} shares.",
                        "shares": share_names,
                    }
                )
            else:
                return Response({"success": False, "error": "Connection refused"})
        except Exception:
            return Response({"success": False, "error": "Connection test failed. Check configuration and try again."})


# ===========================================================================
# User Reports (Bug reports & Feature requests)
# ===========================================================================


class UserReportViewSet(viewsets.ModelViewSet):
    """
    Users can create and view their own reports.
    Super admin can see all reports and update status.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = UserReportSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False) or not self.request.user.is_authenticated:
            return UserReport.objects.none()

        user = self.request.user
        qs = UserReport.objects.select_related("reporter")

        if is_superadmin_user(user):
            # Super admin sees all
            report_type = self.request.query_params.get("report_type")
            if report_type:
                qs = qs.filter(report_type=report_type)

            report_status = self.request.query_params.get("status")
            if report_status:
                qs = qs.filter(status=report_status)

            priority = self.request.query_params.get("priority")
            if priority:
                qs = qs.filter(priority=priority)
        else:
            # Regular users see only their own reports
            qs = qs.filter(reporter=user)

        return qs

    def get_serializer_class(self):
        if self.action in ("partial_update", "update") and is_superadmin_user(self.request.user):
            return UserReportAdminSerializer
        return UserReportSerializer

    def perform_create(self, serializer):
        serializer.save(reporter=self.request.user)
        UserActivityLog.log_activity(
            user=self.request.user,
            action="create",
            resource="user_report",
            resource_id=serializer.instance.id,
            details={"type": serializer.instance.report_type, "title": serializer.instance.title},
            request=self.request,
        )

        # Notify superadmin(s) about the new report
        from .consumers import send_notification_to_user

        report = serializer.instance
        reporter_name = self.request.user.username
        type_label = report.get_report_type_display()
        title = f"New {type_label}"
        message = f'{reporter_name} submitted a {type_label.lower()}: "{report.title}"'

        # Find superadmin users and create DB notifications + WS push
        superadmin_users = list(ExternalUser.objects.filter(is_active=True, role__in=("developer", "superadmin")))
        if superadmin_users:
            admin_notifs = [Notification(recipient=admin, title=title, message=message, event_type="user_report") for admin in superadmin_users]
            created_notifs = Notification.objects.bulk_create(admin_notifs)
            for notif, admin in zip(created_notifs, superadmin_users, strict=True):
                send_notification_to_user(
                    admin.id,
                    {
                        "id": notif.id,
                        "title": title,
                        "message": message,
                        "event_type": "user_report",
                        "event_id": None,
                        "is_read": False,
                        "created_at": notif.created_at.isoformat(),
                    },
                )

    def perform_update(self, serializer):
        instance = serializer.instance
        user = self.request.user
        is_admin = is_superadmin_user(user)
        # Non-admin users can only edit their own open reports
        if not is_admin:
            if instance.reporter_id != user.id:
                raise PermissionDenied("You can only edit your own reports.")
            if instance.status != "open":
                raise PermissionDenied("You can only edit reports that are still open.")

        # Track what changed before saving (for notification message)
        old_status = instance.status
        old_admin_notes = instance.admin_notes

        serializer.save()
        UserActivityLog.log_activity(
            user=self.request.user,
            action="update",
            resource="user_report",
            resource_id=serializer.instance.id,
            details={"status": serializer.instance.status},
            request=self.request,
        )

        # If superadmin updated the report, notify the reporter
        if is_admin and instance.reporter_id != user.id:
            from .consumers import send_notification_to_user

            report = serializer.instance
            changes = []
            if report.status != old_status:
                changes.append(f'status changed to "{report.get_status_display()}"')
            if report.admin_notes != old_admin_notes and report.admin_notes:
                changes.append("admin added a note")
            change_text = " and ".join(changes) if changes else "updated"

            title = "Report Updated"
            message = f'Your {report.get_report_type_display().lower()} "{report.title}" was {change_text} by admin.'

            notif = Notification.objects.create(
                recipient_id=report.reporter_id,
                title=title,
                message=message,
                event_type="user_report",
            )
            send_notification_to_user(
                report.reporter_id,
                {
                    "id": notif.id,
                    "title": title,
                    "message": message,
                    "event_type": "user_report",
                    "event_id": None,
                    "is_read": False,
                    "created_at": notif.created_at.isoformat(),
                },
            )

    def perform_destroy(self, instance):
        user = self.request.user
        # Non-admin users can only delete their own open reports
        if not is_superadmin_user(user):
            if instance.reporter_id != user.id:
                raise PermissionDenied("You can only delete your own reports.")
            if instance.status != "open":
                raise PermissionDenied("You can only delete reports that are still open.")
        UserActivityLog.log_activity(
            user=user,
            action="delete",
            resource="user_report",
            resource_id=instance.id,
            details={"type": instance.report_type, "title": instance.title},
            request=self.request,
        )
        instance.delete()

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        """Return aggregate stats for super admin dashboard."""
        if not is_superadmin_user(request.user):
            return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

        from django.db.models import Count

        qs = UserReport.objects.all()
        by_status = dict(qs.values_list("status").annotate(c=Count("id")).values_list("status", "c"))
        by_type = dict(qs.values_list("report_type").annotate(c=Count("id")).values_list("report_type", "c"))
        return Response(
            {
                "total": qs.count(),
                "by_status": by_status,
                "by_type": by_type,
            }
        )


# ===========================================================================
# Release Notes / Version History
# ===========================================================================


class ReleaseNoteViewSet(viewsets.ModelViewSet):
    """
    Public read for published release notes.
    Super admin can create / update / delete.
    Auto-syncs SystemConfiguration version + build_date from the latest release.
    """

    permission_classes = [IsAuthenticated]
    serializer_class = ReleaseNoteSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False) or not self.request.user.is_authenticated:
            return ReleaseNote.objects.none()

        qs = ReleaseNote.objects.select_related("created_by")
        if not is_superadmin_user(self.request.user):
            qs = qs.filter(published=True)
        return qs

    @staticmethod
    def _sync_system_version():
        """Update SystemConfiguration version + build_date from the latest published ReleaseNote."""
        latest = ReleaseNote.objects.filter(published=True).order_by("-release_date", "-created_at").values("version", "release_date").first()
        if latest:
            from calendar import month_name

            rd = latest["release_date"]
            build_date = f"{month_name[rd.month]} {rd.year}"
            SystemConfiguration.objects.update_or_create(
                pk=1,
                defaults={"version": latest["version"], "build_date": build_date},
            )

    def perform_create(self, serializer):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can create release notes.")
        serializer.save(created_by=self.request.user)
        self._sync_system_version()

    def perform_update(self, serializer):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can update release notes.")
        serializer.save()
        self._sync_system_version()

    def perform_destroy(self, instance):
        if not is_superadmin_user(self.request.user):
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Only Super Admin can delete release notes.")
        instance.delete()
        self._sync_system_version()
