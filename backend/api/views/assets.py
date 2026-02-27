import logging
import traceback

from django.contrib.auth import get_user_model
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ..models import (
    Asset,
    Department,
    PurchaseRequest,
)
from ..permissions import ResourcePermission
from ..serializers import (
    AssetSerializer,
    AssetSummarySerializer,
    PurchaseRequestSerializer,
)
from .helpers import get_employee_for_user, is_developer_user, is_ptb_admin, is_superadmin_user  # noqa: F401

logger = logging.getLogger(__name__)
User = get_user_model()


class PurchaseRequestViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing purchase requests.
    Supports CRUD operations, filtering, and import/export.
    """

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "purchasing"
    queryset = PurchaseRequest.objects.all()
    serializer_class = PurchaseRequestSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["owner", "doc_id", "part_no", "description_spec", "pr_no", "remarks"]
    ordering_fields = ["id", "request_date", "owner", "doc_id", "part_no", "pr_no", "status", "created_at"]
    ordering = ["-request_date", "-id"]

    def get_queryset(self):
        queryset = PurchaseRequest.objects.select_related("owner_employee", "owner_employee__department").all()

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by date range
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            queryset = queryset.filter(request_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(request_date__lte=end_date)

        return queryset

    @swagger_auto_schema(
        operation_summary="Import purchase requests from CSV/Excel",
        manual_parameters=[
            openapi.Parameter(name="file", in_=openapi.IN_FORM, type=openapi.TYPE_FILE, required=True),
        ],
    )
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Import purchase requests from CSV or Excel file.
        Supports multi-sheet Excel files with sheets:
        - 'List of Purchase' or active sheet -> status=pending
        - 'Done' -> status=done
        - 'Cancel Purchase' -> status=canceled
        """
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        import csv
        import io
        from datetime import datetime

        def parse_date(val):
            if not val:
                return None
            for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(str(val), fmt).date()
                except (ValueError, TypeError):
                    continue
            return None

        def process_rows(rows, default_status="pending"):
            batch = []
            for row in rows:
                pr_data = {
                    "request_date": parse_date(row.get("Request Date")),
                    "owner": row.get("Owner") or None,
                    "doc_id": row.get("Doc ID") or None,
                    "part_no": row.get("Part No.") or None,
                    "description_spec": row.get("Description-Spec") or None,
                    "material_category": row.get("Material Category") or None,
                    "purpose_desc": row.get("Purpose/Desc.") or None,
                    "qty": row.get("Qty") or 1,
                    "plant": row.get("Plant") or None,
                    "project_code": row.get("Project Code") or None,
                    "pr_type": row.get("PR Type") or None,
                    "mrp_id": row.get("MRPID") or None,
                    "purch_org": row.get("Purch. Org.") or None,
                    "sourcer_price": row.get("Sourcer Price") or None,
                    "pr_no": row.get("PR No.") or None,
                    "remarks": row.get("Remarks") or None,
                    "status": default_status,
                }
                batch.append(PurchaseRequest(**pr_data))
            if batch:
                PurchaseRequest.objects.bulk_create(batch, batch_size=200)
            return len(batch)

        def read_sheet(ws):
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row, strict=True))
                # Skip empty rows
                if any(v for v in row_dict.values() if v is not None):
                    rows.append(row_dict)
            return rows

        try:
            if file.name.endswith(".csv"):
                content = file.read().decode("utf-8")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
                created_count = process_rows(rows, "pending")
                return Response({"message": f"Successfully imported {created_count} purchase requests"})

            elif file.name.endswith(".xlsx") or file.name.endswith(".xls"):
                import openpyxl

                wb = openpyxl.load_workbook(file)
                sheet_names = wb.sheetnames

                # Sheet-to-status mapping
                sheet_status_map = {
                    "List of Purchase": "pending",
                    "Done": "done",
                    "Cancel Purchase": "canceled",
                }

                total_created = 0
                sheets_processed = []

                # Check if workbook has our expected sheet names
                has_named_sheets = any(name in sheet_names for name in sheet_status_map)

                if has_named_sheets:
                    # Multi-sheet import
                    for sheet_name, default_status in sheet_status_map.items():
                        if sheet_name in sheet_names:
                            ws = wb[sheet_name]
                            rows = read_sheet(ws)
                            if rows:
                                count = process_rows(rows, default_status)
                                total_created += count
                                sheets_processed.append(f"{sheet_name}: {count}")
                else:
                    # Fallback: read active sheet as pending
                    ws = wb.active
                    rows = read_sheet(ws)
                    total_created = process_rows(rows, "pending")
                    sheets_processed.append(f"{ws.title}: {total_created}")

                detail = ", ".join(sheets_processed)
                return Response({"message": f"Successfully imported {total_created} purchase requests ({detail})", "created": total_created})
            else:
                return Response({"detail": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)

        except Exception:
            return Response({"detail": "An unexpected error occurred. Please try again or contact support."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """Update a purchase request and notify on status change."""
        from ..signals import notify_purchase_request_status_change

        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        old_status = instance.status

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Check for status change and notify
        new_status = serializer.validated_data.get("status", instance.status)
        if old_status != new_status and new_status in ["done", "canceled"]:
            notify_purchase_request_status_change(instance, old_status, new_status)

        return Response(serializer.data)

    def partial_update(self, request, *args, **kwargs):
        """Partial update a purchase request."""
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """Delete multiple purchase requests at once."""
        if not is_ptb_admin(request.user) and not is_superadmin_user(request.user):
            return Response({"detail": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"detail": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = PurchaseRequest.objects.filter(id__in=ids).delete()
        return Response({"message": f"Successfully deleted {deleted_count} purchase requests", "deleted": deleted_count})

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        """Update status of multiple purchase requests at once."""
        if not is_ptb_admin(request.user) and not is_superadmin_user(request.user):
            return Response({"detail": "Admin access required"}, status=status.HTTP_403_FORBIDDEN)
        ids = request.data.get("ids", [])
        new_status = request.data.get("status", "")
        if not ids:
            return Response({"detail": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        if new_status not in ["pending", "done", "canceled"]:
            return Response({"detail": "Invalid status"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch requests that will actually change status so we can notify
        requests_to_update = list(PurchaseRequest.objects.filter(id__in=ids).exclude(status=new_status))
        updated_count = PurchaseRequest.objects.filter(id__in=ids).update(status=new_status)

        # Send notifications for status changes to done/canceled
        if new_status in ["done", "canceled"]:
            from ..signals import notify_purchase_request_status_change

            for pr in requests_to_update:
                old_status = pr.status
                pr.status = new_status  # Update in-memory for the notification
                try:
                    notify_purchase_request_status_change(pr, old_status, new_status)
                except Exception:
                    logger.warning("Failed to notify PR %s status change", pr.id, exc_info=True)

        return Response({"message": f"Successfully updated {updated_count} purchase requests to {new_status}", "updated": updated_count})


class AssetViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing assets.
    Supports CRUD operations, filtering by department, and import/export.
    """

    permission_classes = [IsAuthenticated, ResourcePermission]
    resource_name = "assets"
    queryset = Asset.objects.all()
    serializer_class = AssetSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["asset_id", "part_number", "product_name", "keeper_name", "cost_center"]
    ordering_fields = ["id", "asset_id", "part_number", "product_name", "keeper_name", "status", "receive_date", "created_at"]
    ordering = ["-receive_date", "asset_id"]

    def get_queryset(self):
        queryset = Asset.objects.select_related("department").all()

        # Filter by department
        department_id = self.request.query_params.get("department")
        if department_id:
            queryset = queryset.filter(department_id=department_id)

        # Filter by cost center
        cost_center = self.request.query_params.get("cost_center")
        if cost_center:
            queryset = queryset.filter(cost_center=cost_center)

        # Filter by status
        status_filter = self.request.query_params.get("status")
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return AssetSummarySerializer
        return AssetSerializer

    @swagger_auto_schema(
        operation_summary="Get assets grouped by department",
        responses={200: openapi.Response("Assets grouped by department")},
    )
    @action(detail=False, methods=["get"])
    def by_department(self, request):
        """Get all assets grouped by department, with optional search"""
        departments = Department.objects.filter(is_enabled=True).order_by("code")

        # Fetch only the fields needed by AssetSummarySerializer to reduce memory
        all_assets_qs = Asset.objects.select_related("department").only(
            "id",
            "asset_id",
            "part_number",
            "product_name",
            "spec",
            "quantity",
            "receive_date",
            "status",
            "cost_center",
            "keeper_dept",
            "department_id",
            "department__code",
            "keeper_name",
        )

        # Apply search filter if provided
        search_query = request.query_params.get("search", "").strip()
        if search_query:
            from django.db.models import Q

            all_assets_qs = all_assets_qs.filter(Q(asset_id__icontains=search_query) | Q(part_number__icontains=search_query) | Q(product_name__icontains=search_query) | Q(keeper_name__icontains=search_query) | Q(cost_center__icontains=search_query))

        # Group assets by department using iterator to reduce peak memory usage
        dept_assets_map = {}
        unassigned_assets = []
        dept_codes = {dept.code: dept for dept in departments}

        for asset in all_assets_qs.iterator(chunk_size=500):
            matched_dept = None
            if asset.department and asset.department.code in dept_codes:
                matched_dept = asset.department.code
            elif asset.cost_center and asset.cost_center in dept_codes:
                matched_dept = asset.cost_center
            elif asset.keeper_dept and asset.keeper_dept in dept_codes:
                matched_dept = asset.keeper_dept

            if matched_dept:
                dept_assets_map.setdefault(matched_dept, []).append(asset)
            else:
                unassigned_assets.append(asset)

        result = []
        for dept in departments:
            assets = dept_assets_map.get(dept.code, [])
            serializer = AssetSummarySerializer(assets, many=True)
            result.append({"department_id": dept.id, "department_code": dept.code, "department_name": dept.name, "asset_count": len(assets), "assets": serializer.data})

        if unassigned_assets:
            serializer = AssetSummarySerializer(unassigned_assets, many=True)
            result.append({"department_id": None, "department_code": "UNASSIGNED", "department_name": "Unassigned Assets", "asset_count": len(unassigned_assets), "assets": serializer.data})

        return Response(result)

    @swagger_auto_schema(
        operation_summary="Import assets from CSV/Excel",
        manual_parameters=[
            openapi.Parameter(name="file", in_=openapi.IN_FORM, type=openapi.TYPE_FILE, required=True),
        ],
    )
    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def import_data(self, request):
        """Import assets from CSV or Excel file"""
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        import csv
        import io

        try:
            # Read file content
            if file.name.endswith(".csv"):
                content = file.read().decode("utf-8")
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif file.name.endswith(".xlsx") or file.name.endswith(".xls"):
                import openpyxl

                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row, strict=True)))
            else:
                return Response({"detail": "Unsupported file format"}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0

            # Pre-load departments for batch lookup
            dept_map = {d.code: d for d in Department.objects.all()}

            # Collect all asset_ids from the rows and pre-load existing assets
            row_asset_ids = [row.get("Asset Id") for row in rows if row.get("Asset Id")]
            existing_assets = {a.asset_id: a for a in Asset.objects.filter(asset_id__in=row_asset_ids)}

            to_create = []
            to_update = []

            for row in rows:
                asset_id = row.get("Asset Id")
                if not asset_id:
                    continue

                # Parse boolean values
                def parse_bool(val):
                    if isinstance(val, bool):
                        return val
                    return str(val).upper() in ["Y", "YES", "TRUE", "1"] if val else False

                # Parse date values
                def parse_date(val):
                    if not val:
                        return None
                    from datetime import datetime

                    for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                        try:
                            return datetime.strptime(str(val), fmt).date()
                        except ValueError:
                            continue
                    return None

                # Parse decimal values
                def parse_decimal(val):
                    if not val:
                        return None
                    try:
                        return float(str(val).replace(",", ""))
                    except ValueError:
                        return None

                # Map CSV headers to model fields
                asset_data = {
                    "company_code": row.get("Company Code") or None,
                    "fixed_asset_id": row.get("Fixed Asset Id") or None,
                    "is_fixed_asset": parse_bool(row.get("Is Fixed Asset")),
                    "is_customs_control": parse_bool(row.get("Is Customs Control")),
                    "part_number": row.get("Part Number") or None,
                    "group_3": row.get("Group 3") or None,
                    "product_name": row.get("Product Name") or None,
                    "spec": row.get("Spec") or None,
                    "quantity": parse_decimal(row.get("Quantity")) or 1,
                    "receive_date": parse_date(row.get("Receive Date")),
                    "status": row.get("Status") or None,
                    "cost_center": row.get("Cost Center") or None,
                    "cost_center_name": row.get("Cost Center Name") or None,
                    "keeper_dept": row.get("Keeper Dept.") or None,
                    "keeper_dept_name": row.get("Keeper Dept. Name") or None,
                    "keeper": row.get("Keeper") or None,
                    "keeper_name": row.get("Keeper Name") or None,
                    "group_1": row.get("Group 1") or None,
                    "group_2": row.get("Group 2") or None,
                    "storage": row.get("Storage") or None,
                    "location_code": row.get("Location Code") or None,
                    "storage_desc": row.get("Storage Desc.") or None,
                    "consign": row.get("Consign") or None,
                    "vendor": row.get("Vendor") or None,
                    "pr_no": row.get("PR No.") or None,
                    "pr_sequence": row.get("PR Sequence") or None,
                    "po_no": row.get("PO No.") or None,
                    "po_sequence": row.get("PO Sequence") or None,
                    "dn_no": row.get("DN No.") or None,
                    "dn_sequence": row.get("DN Sequence") or None,
                    "dn_date": parse_date(row.get("DN Date")),
                    "application_number": row.get("Application Number") or None,
                    "sequence": row.get("Sequence") or None,
                    "import_number": row.get("Import Number") or None,
                    "picking_no": row.get("Picking No.") or None,
                    "picking_sequence": row.get("Picking Sequence") or None,
                    "picking_year": row.get("Picking Year") or None,
                    "picking_date": parse_date(row.get("Picking Date")),
                    "chinese_product_name": row.get("Chinese Product Name") or None,
                    "hs_code": row.get("HS Code") or None,
                    "declaration_number": row.get("Declaration Number") or None,
                    "declaration_date": parse_date(row.get("Declaration Date")),
                    "control_end_date": parse_date(row.get("Control End Date")),
                    "outsource_number": row.get("Outsource Number") or None,
                    "price": parse_decimal(row.get("Price")),
                    "currency": row.get("Currency") or None,
                    "local_price": parse_decimal(row.get("Local Price")),
                    "price_level": row.get("Price Level") or None,
                    "sn": row.get("SN") or None,
                    "is_qualified": parse_bool(row.get("Is Qualified")),
                    "itc_end_date": parse_date(row.get("ITC End Date")),
                    "elec_declaration_number": row.get("Elec Declaration Number") or None,
                    "national_inspection_certification": row.get("National Inspection Certification") or None,
                    "notes": {
                        k: v
                        for k, v in {
                            "note1": row.get("Note1") or None,
                            "note2": row.get("Note2") or None,
                            "note3": row.get("Note3") or None,
                            "note4": row.get("Note4") or None,
                            "note5": row.get("Note5") or None,
                            "note6": row.get("Note6") or None,
                            "note7": row.get("Note7") or None,
                            "note8": row.get("Note8") or None,
                            "note9": row.get("Note9") or None,
                            "note10": row.get("Note10") or None,
                        }.items()
                        if v is not None
                    },
                }

                # Link to department using pre-loaded map
                dept_code = asset_data.get("cost_center") or asset_data.get("keeper_dept")
                dept = dept_map.get(dept_code) if dept_code else None

                if asset_id in existing_assets:
                    # Update existing
                    existing = existing_assets[asset_id]
                    for field, value in asset_data.items():
                        setattr(existing, field, value)
                    if dept:
                        existing.department = dept
                    to_update.append(existing)
                    updated_count += 1
                else:
                    # Create new
                    new_asset = Asset(asset_id=asset_id, **asset_data)
                    if dept:
                        new_asset.department = dept
                    to_create.append(new_asset)
                    created_count += 1

            if to_create:
                Asset.objects.bulk_create(to_create, batch_size=200)
            if to_update:
                update_fields = list(asset_data.keys()) + ["department"]
                Asset.objects.bulk_update(to_update, update_fields, batch_size=200)

            return Response({"message": "Successfully imported assets", "created": created_count, "updated": updated_count})

        except Exception:
            logger.error("Asset import error: %s", traceback.format_exc())
            return Response(
                {
                    "detail": "An error occurred while importing assets. Please check the file format.",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """Delete multiple assets at once."""
        ids = request.data.get("ids", [])
        if not ids:
            return Response({"detail": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = Asset.objects.filter(id__in=ids).delete()
        return Response({"message": f"Successfully deleted {deleted_count} assets", "deleted": deleted_count})

    @action(detail=False, methods=["post"])
    def bulk_update_status(self, request):
        """Update status of multiple assets at once."""
        ids = request.data.get("ids", [])
        new_status = request.data.get("status", "")
        if not ids:
            return Response({"detail": "No IDs provided"}, status=status.HTTP_400_BAD_REQUEST)
        if not new_status:
            return Response({"detail": "No status provided"}, status=status.HTTP_400_BAD_REQUEST)
        updated_count = Asset.objects.filter(id__in=ids).update(status=new_status)
        return Response({"message": f"Successfully updated {updated_count} assets to {new_status}", "updated": updated_count})


# ===========================================================================
# SMB Configuration (Super Admin only — multi-config with active selection)
# ===========================================================================
